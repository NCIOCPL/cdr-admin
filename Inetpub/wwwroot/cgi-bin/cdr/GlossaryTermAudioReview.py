#!/usr/bin/env python

"""Review glossary term audio files for approval of import.

The pronunciation files for the GlossaryTermName documents are produced
by a contractor (currently Vanessa Richardson) and sent to CIAT in a
zipfile using naming conventions established for this purpose. CIAT
uploads the zipfiles to the CIPSFTP server. A separate script is used
to transfer the zipfiles to the CDR server. A pair of database tables
is used to track the zipfiles and the review disposition of each of
the MP3 pronunciation files in each zipfile. In addition to the MP3
files, the zipfile contains a spreadsheet cataloging the audio files,
with the CDR ID for the glossary term to which an audio file will be
attached, a term name from the CDR document, the language for that
name, an optional pronunciation, and some other columns. Rows are
added to the database tables the first time a set of audio files is
selected for review.

The processing for this script happens in several phases.  When the
user first invokes the script, the list of all sets of audio files
in the file system is presented, with their statuses (completed,
started, unreviewed). Each set which is not completed has a hyperlink
to the next page, which contains a form for recording the review
decision and optional reviewer notes. When this form is submitted,
the new decisions are recorded in the database, and if there are
still audio files to be reviewed, the form is redisplayed. Otherwise,
the user it taken back to the list of zipfile sets. If there are any
audio files which were rejected after all of the reviews have finished,
a new spreadsheet is generated and a link provided for its download,
to be used as the start of another round of reviews after replacement
MP3 files have been created by the contractor for the ones which had
been rejected.

The script is complicated somewhat by the need to track separately
the information in the database tables, and the information in the
files on the disk, some of which are not yet reflected in the database.
Here are the properties containing the information needed about the
audio file sets at various stages of processing:

   zipfiles

        This is the catalog of all of the zipfiles which have at least
        started review, and are therefore relected in the database
        tables. It has no information about the individual audio files,
        but just the information needed for the first page, including
        the date and name for the zipfile, and whether the review of
        the set has finished. This catalog will contain zipfiles which
        are no longer available on disk in the audio files directory,
        because they have been archived by the scheduled Files Sweeper
        job. Only those zipfiles which are still in the audio files
        directory are shown on the landing page for this script.

    zipfiles_on_disk

        The catalog of all of the zipfiles in the file system, some of
        which may not yet be represented in the database tables.

    audio_set

        This contains the detailed information about the particular
        zipfile selected by the user for review, including particulars
        about each of the MP3 audio pronunciation files and their
        review dispositions.
"""

from cdr import run_command
from cdrcgi import Controller, Excel
from cdrapi.settings import Tier
from datetime import datetime
from functools import cached_property
from io import BytesIO
from os import scandir
from sys import stdout
from re import compile
from zipfile import ZipFile
from openpyxl import load_workbook


class Control(Controller):
    """Logic control center for the script."""

    SUBTITLE = "Glossary Term Audio Review"
    SAVE = "Save"
    BASEDIR = Tier().basedir
    LOGNAME = "GlossaryTermAudioReview"
    ZIPDIR = f"{BASEDIR}/Audio_from_CIPSFTP"
    REVDIR = f"{ZIPDIR}/GeneratedRevisionSheets"
    IGNORE = "__MACOSX"
    NOTEPAT = compile(r"[\r\n]+")
    NAMEPAT = compile(r"(?i)(?P<base>Week_\d{4}_\d\d)(?P<rev>_Rev\d)*.zip")
    REVPAT = compile(r"(?i)_Rev(?P<num>\d+)")
    MAXNOTE = 2040
    MAXFILE = 250
    MAXTERM = 250
    MAXNAME = 250
    PERMISSION = "REVIEW TERM AUDIO"
    FIXNAME_INSTRUCTIONS = [
        "Please correct the name to reflect one of the following formats "
        "or contact programming support staff for assistance.",
        "Week_YYYY_WW.zip or Week_YYYY_WW_RevN.zip",
        "... where 'Y', 'W', and 'N' represent decimal digits.",
    ]

    def run(self):
        """Provide custom routing."""

        args = self.request, self.name, self.id
        self.logger.debug("request=%s name=%s, id=%s", *args)
        try:
            if self.book:
                return self.send_book()
            if self.mp3:
                return self.send_mp3()
            elif self.request == self.SAVE:
                return self.save()
        except Exception as e:
            self.logger.exception("Failure")
            self.bail(e)
        Controller.run(self)

    def populate_form(self, page):
        """Show the review form for a set, or the set list if none selected.

        The landing page for this script shows the list of audio file sets
        on the disk. If the user selects one of the sets, we draw the
        form for rewviewing the audio files in that set.

        Pass:
            page - HTMLPage object on which we draw the form
        """

        # Show the review form for an audio file set if one has been picked.
        rules = []
        if self.audio_set:

            # Assemble the instructions for using the form.
            instructions = (
                "Click a hyperlinked mp3 filename to play the sound in "
                "your browser-configured mp3 player (files which have "
                "already been reviewed files are at the bottom of the "
                "list of files).",
                "Use the radio buttons to approve or reject a file.",
                "When finished, click 'Save' to save any changes to "
                "the database. If all files in the set have been reviewed "
                "and any have been rejected, a spreadsheet containing "
                "rejected terms will be created and displayed on your "
                "workstation. Please save it for future use.",
            )
            fieldset = page.fieldset("Instructions")
            for paragraph in instructions:
                fieldset.append(page.B.P(paragraph))

            # Force the page object to assemble the DOM elements for the main
            # block. We need to move the very wide form out from the grid
            # container.
            grid_container = None
            if page.main is not None:
                grid_container = page.form.getparent()
            page.form.append(page.hidden_field("id", self.audio_set.id))
            if grid_container is not None:
                grid_container.append(fieldset)
                grid_container.addnext(page.form)
            else:
                page.form.append(fieldset)
                message = "Unable to widen the form."
                header = "Internal error"
                alert = dict(header=header, message=message, type="warning")
                self.alerts.append(alert)
            page.form.append(page.button(self.SAVE))
            page.form.append(self.audio_set.table)
            page.form.set("target", "_self")
            rules += (
                "#primary-form table { width: 100%; }",
                ".usa-table td.status-buttons { padding: 0 1rem .8rem; }",
                "#primary-form { width: 90%; margin: 2rem; }",
            )

        # Otherwise, show the list of all the sets on the disk.
        else:
            fieldset = page.fieldset("Instructions")
            instructions = (
                "Click a link to a zip file to review from the table below. "
                "Only those files that have not yet been completely reviewed "
                "are hyperlinked."
            )
            fieldset.append(page.B.P(instructions))
            page.form.append(fieldset)
            columns = "File name", "Review status", "Date modified"
            columns = [page.B.TH(column) for column in columns]
            thead = page.B.THEAD(page.B.TR(*columns))
            classes = "usa-table usa-table--borderless"
            table = page.B.TABLE(thead, page.B.CLASS(classes))
            for zipfile in self.zipfiles_on_disk:
                table.append(zipfile.row)
            fieldset = page.fieldset("Audio Zip Files")
            fieldset.append(table)
            page.form.append(fieldset)
            color = f"color: {page.LINK_COLOR};"
            rules += [
                f".usa-form td a {{ text-decoration: None; {color} }}",
                f".usa-form td a:visited {{ {color} }}",
            ]
        page.add_css("\n".join(rules))

    def save(self):
        """Save review results and show another form.

        If the user has not completed the review of this set, redisplay
        its review form. Otherwise, go back to the display of all the
        sets on the disk.
        """

        if not self.session.can_do(self.PERMISSION):
            self.bail("User not authorized to review term audio files")
        updates = 0
        for mp3 in self.audio_set.audio_files:
            status = self.fields.getvalue(f"status-{mp3.id}") or "U"
            note = self.fields.getvalue(f"note-{mp3.id}") or ""
            note = self.NOTEPAT.sub("\n", note.strip())[:self.MAXNOTE]
            if note != mp3.reviewer_note or status != mp3.review_status:
                mp3.update(status, note)
                updates += 1

        # If there have been any changes, commit them and refresh the set.
        if updates:
            self.logger.info("updated %d mp3 rows", updates)
            self.conn.commit()
        if self.audio_set.done:
            self.suppress_sidenav = False
            book_name = self.audio_set.close()
            legend = f"Audio Set {self.audio_set.name} Review Complete"
            fieldset = self.form_page.fieldset(legend)
            args = ["All of the audio files in this set have been reviewed. "]
            if book_name:
                url = self.make_url(self.script, book=book_name)
                label = "the workbook for these rejected audio files"
                link = self.form_page.B.A(label, href=url)
                args += [
                    "Some of the audio files were rejected. You can retrieve ",
                    link,
                    ", which can be used for the next round of audio files.",
                ]
            else:
                args.append(
                    "None of the files in the set were rejected, so there "
                    "is no new workbook for a subsequent round of files."
                )
            paragraph = self.form_page.B.P(*args)
            fieldset.append(paragraph)
            self.form_page.form.append(fieldset)
            self.audio_set = None
        else:
            if updates:
                message = f"Saved updates for {updates} recording(s)."
                self.alerts.append(dict(message=message, type="success"))
        self.show_form()

    def send_book(self):
        """Serve up the new workbook with rejected audio files."""

        with open(f"{self.REVDIR}/{self.book}.xlsx", "rb") as fp:
            book_bytes = fp.read()
        mime_type = f"application/{Excel.MIME_SUBTYPE}"
        self.send_bytes(book_bytes, f"{self.book}.xlsx", mime_type)

    def send_bytes(self, payload, name, mime_type):
        """Return a binary file to the browser.

        Used by `send_book()` and `send_mp3()`.

        Pass:
            payload - the bytes to return
            name - string for the content disposition's filename
            mime_type - standard RFC6838 type/subtype string
        """

        headers = (
            f"Content-Type: {mime_type}",
            f"Content-disposition: inline; filename={name}",
            f"Content-Length: {len(payload):d}",
        )
        for header in headers:
            stdout.buffer.write(header.encode("utf-8"))
            stdout.buffer.write(b"\n")
        stdout.buffer.write(b"\n")
        stdout.buffer.write(payload)

    def send_mp3(self):
        """Let the reviewer listen to the audio file."""

        query = self.Query("term_audio_mp3 m", "m.mp3_name", "z.filename")
        query.join("term_audio_zipfile z", "z.id = m.zipfile_id")
        query.where(query.Condition("m.id", self.mp3))
        mp3_name, filename = query.execute(self.cursor).fetchone()
        with ZipFile(f"{self.ZIPDIR}/{filename}") as zipfile:
            mp3_bytes = zipfile.read(mp3_name)
        self.send_bytes(mp3_bytes, mp3_name, "audio/mpeg")

    @cached_property
    def audio_set(self):
        """Information about the set of MP3 files being reviewed."""

        if self.name:
            return AudioSet(self, name=self.name)
        elif self.id:
            return AudioSet(self, id=self.id)
        return None

    @cached_property
    def book(self):
        """Name of new workbook with rejected audio files.

        Used by the callback to fetch the new Excel file.
        """
        return self.fields.getvalue("book")

    @cached_property
    def buttons(self):
        """We'll put the button at the top in populate_form()."""
        return []

    @cached_property
    def id(self):
        """ID of the MP3 file set's row in the database table."""
        return self.fields.getvalue("id")

    @cached_property
    def mp3(self):
        """ID of the MP3 file the reviewer wishes to hear."""
        return self.fields.getvalue("mp3")

    @cached_property
    def name(self):
        """File name for the selected MP3 file set to be reviewed."""
        return self.fields.getvalue("name")

    @cached_property
    def name_counts(self):
        """Index of integers for new MP3 names.

        This is used to prevent name collisions in the event there
        are multiple Spanish names for the same term.
        """

        return {}

    @cached_property
    def suppress_sidenav(self):
        """Use the full grid container width for the second form."""
        return True if self.audio_set else False

    @cached_property
    def user_id(self):
        """Account ID for the current CDR user."""
        return self.session.user_id

    @cached_property
    def zipfiles(self):
        """Load the complete set of term audio zipfiles from the database.

        Does not include zipfiles which are in the file system but have
        not yet been reviewed. See the `zipfiles_on_disk` property for
        the list of all file in the zipfile directory which match our
        conventional filename pattern for audio zipfiles.

        This set contains zipfiles which are no longer in the audio
        files directory in the file system (because they have been
        archived by the scheduler file sweeper). Only those zipfiles
        which are still in the file system are shown in the list of
        zipfiles on this script's initial page.
        """

        class ZipFiles:
            """ID and name indexes to the term audio zipfiles."""

            def __init__(self, control):
                """Save the reference to the control object.

                Pass:
                    control - access to the DB and the HTML builder class
                """

                self.control = control

            @cached_property
            def files(self):
                """Sequence of `ZipFile` objects."""

                query = self.control.Query("term_audio_zipfile", "*")
                rows = query.execute(self.control.cursor).fetchall()
                return [self.ZipFile(self.control, row) for row in rows]

            @cached_property
            def ids(self):
                """Dictionary of zipfiles by primary key."""
                return dict([(file.id, file) for file in self.files])

            @cached_property
            def names(self):
                """Dictionary of zipfiles by primary key."""
                return dict([(f.filename, f) for f in self.files])

            class ZipFile:
                """Information about a single archive of audio files.

                This is a simpler class than the global `AudioSet`
                class. That class has information about the audio
                files in the zip file. This class has just enough
                information to meet the needs of the page which
                displays all of the zipfiles.

                Properties:
                    id - integer primary key for the zipfile record
                    filename - string for the zipfile's name
                    filedate - date/time stamp for the zipfile
                    complete - Boolean indicating whether reviews are done
                """

                PROPS = "id", "filename", "filedate", "complete"

                def __init__(self, control, row):
                    """Capture the caller's information.

                    Pass:
                        control - access to the HTML builder class
                        row - result set row from the SQL query
                    """

                    self.control = control
                    self.row = row

                def __getattr__(self, name):
                    """Return the other properties directly."""
                    return getattr(self.row, name)

                def __str__(self):
                    """String for debugging/logging."""

                    props = [f"{n}={getattr(self, n)}" for n in self.PROPS]
                    return " ".join(props)

                @cached_property
                def complete(self):
                    """True if all the audio files have been reviewed."""
                    return self.row.complete == "Y"

        zipfiles = {}
        for zipfile in ZipFiles(self).files:
            zipfiles[zipfile.filename.lower()] = zipfile
        return zipfiles

    @cached_property
    def zipfile_names(self):
        """Index by name of all the audio set zipfiles on the disk."""

        by_name = {}
        for zipfile in self.zipfiles_on_disk:
            by_name[zipfile.key] = zipfile
        return by_name

    @cached_property
    def zipfiles_on_disk(self):
        """Zipfiles in the file system."""

        class DiskFile:
            STARTED = "Started"
            UNREVIEWED = "Unreviewed"
            COMPLETED = "Completed"
            STATUS_SORT = {STARTED: 1, UNREVIEWED: 2, COMPLETED: 3}

            def __init__(self, control, entry):
                self.control = control
                self.entry = entry

            def __lt__(self, other):
                """Sort by status then by filename."""
                return self.sortkey < other.sortkey

            @cached_property
            def datetime(self):
                """When was the file last modified?"""
                return datetime.fromtimestamp(self.entry.stat().st_mtime)

            @cached_property
            def db_info(self):
                """Information about this file from the database."""
                return self.control.zipfiles.get(self.key)

            @cached_property
            def name(self):
                """Base name for the file."""
                return self.entry.name

            @cached_property
            def path(self):
                """Location of the file."""
                return self.entry.path.replace("\\", "/")

            @cached_property
            def key(self):
                """Used by the sort key."""
                return self.name.lower()

            @cached_property
            def sortkey(self):
                "Major sort by status, subsort by filename"
                return self.STATUS_SORT[self.status], self.key

            @cached_property
            def status(self):
                """Review status for the set."""

                if not self.db_info:
                    return self.UNREVIEWED
                elif self.db_info.complete:
                    return self.COMPLETED
                else:
                    return self.STARTED

            @cached_property
            def row(self):
                """Table row for showing the available sets."""

                B = self.control.HTMLPage.B
                filename = self.entry.name
                if self.status != self.COMPLETED:
                    script = self.control.script
                    if self.status == self.UNREVIEWED:
                        params = dict(name=self.name)
                    else:
                        params = dict(id=self.db_info.id)
                    url = self.control.make_url(script, **params)
                    filename = B.A(filename, href=url, target=self.target)
                filename = B.TD(filename)
                status = B.TD(self.status)
                modified = str(self.datetime)[:19]
                modified = B.TD(modified)
                return B.TR(filename, status, modified)

            @cached_property
            def target(self):
                """Don't open new browser tabs indefinitely."""

                if self.control.request == Control.SAVE:
                    return "_self"
                return "_blank"

        files = []
        for entry in scandir(self.ZIPDIR):
            key = entry.name.lower()
            if key.startswith("week") and key.endswith(".zip"):
                if self.NAMEPAT.match(entry.name):
                    files.append(DiskFile(self, entry))
                else:
                    message = f"Found file {entry.name!r}."
                    self.logger.warning(message)
                    self.bail(message, extra=self.FIXNAME_INSTRUCTIONS)
        return sorted(files)


class AudioSet:
    """Zip archive of glossary pronunciation audio files."""

    TABLE = "term_audio_mp3"
    FIELDS = (
        "zipfile_id",
        "cdr_id",
        "term_name",
        "language",
        "pronunciation",
        "mp3_name",
        "reader_note",
        "reviewer_note",
        "reuse_media_id",
        "reviewer_id",
        "review_date",
    )
    HEADERS = (
        "Disposition",
        "CDR ID",
        "Term name",
        "Lang",
        "Pronunciation",
        "MP3 file",
        "Reader note",
        "Reviewer note",
    )
    COLUMNS = (
        ("CDR ID", 10),
        ("Term Name", 30),
        ("Language", 10),
        ("Pronunciation", 30),
        ("Filename", 30),
        ("Notes (Vanessa)", 20),
        ("Notes (NCI)", 30),
        ("Reuse Media ID", 15),
    )
    FIELD_LIST = ", ".join(FIELDS)
    PLACEHOLDERS = ", ".join(["?"] * len(FIELDS))
    INSERT = f"INSERT INTO {TABLE} ({FIELD_LIST}) VALUES ({PLACEHOLDERS})"
    UPDATE = "UPDATE term_audio_zipfile SET complete = 'Y' WHERE id = ?"
    LANGUAGE_COL = 3
    LANGUAGES = "English", "Spanish"

    def __init__(self, control, **opts):
        """Save caller's values.

        Required positional argument:
            control - access to database

        Keyword arguments (one must be supplied):
            id - integer for primary key to term_audio_zipfile db table
            name - string for zip file name
        """

        self.__control = control
        self.__opts = opts
        self.__cache = {}

    def close(self):
        """Update the database row save any rejects to a new workbook."""

        self.cursor.execute(self.UPDATE, self.id)
        self.control.conn.commit()
        if self.rejects:
            path = self.new_workbook.save(Control.REVDIR).replace("/", "\\")
            process = run_command(f"fix-permissions {path}")
            if process.stderr:
                self.control.bail(f"Unable to fix permissions for {path}",
                                  extra=[process.stderr])
            return self.new_name
        return None

    @property
    def audio_files(self):
        """Sequence of `MP3` files in this set (manually cached)."""

        if "audio_files" not in self.__cache:
            query = self.control.Query("term_audio_mp3", "*")
            query.order("cdr_id", "mp3_name")
            query.where(query.Condition("zipfile_id", self.id))
            rows = query.execute(self.cursor).fetchall()
            audio_files = [self.MP3(self.control, row) for row in rows]
            self.__cache["audio_files"] = audio_files
        return self.__cache["audio_files"]

    @cached_property
    def control(self):
        """Access to the database, runtime parameters, and logging."""
        return self.__control

    @cached_property
    def cursor(self):
        """Access to the database."""
        return self.control.cursor

    @property
    def done(self):
        """Boolean: have all of the files been reviewed?

        We get asked this question after changes to the reviews have
        been committed to the database, so we need to refresh the
        cache of our audio file information.
        """

        del self.__cache["audio_files"]
        return all([(mp3.review_status in "AR") for mp3 in self.audio_files])

    @cached_property
    def filename(self):
        """The name of the zip file (without its directory)."""
        return self.row.filename

    @cached_property
    def id(self):
        """Primary key for the set's row in the term_audio_zipfile table."""
        return self.row.id

    @cached_property
    def name(self):
        """Alias for self.filename."""
        return self.filename

    @cached_property
    def new_name(self):
        """Filename for a new Excel workbook for rejected audio files."""

        match = Control.NAMEPAT.match(self.filename)
        base = match.group("base")
        suffix = match.group("rev")
        revision = 1
        if suffix:
            match = Control.REVPAT.match(suffix)
            revision = int(match.group("num")) + 1
        return f"{base}_Rev{revision:d}"

    @cached_property
    def new_workbook(self):
        """New Excel workbook for the rejected audio files."""

        book = Excel(self.new_name)
        book.add_sheet("Term Names")
        styles = dict(alignment=book.center, font=book.bold)
        col = 1
        for name, width in self.COLUMNS:
            book.set_width(col, width)
            book.write(1, col, name, styles)
            col += 1
        row = 2
        for mp3 in self.rejects:
            book.write(row, 1, mp3.cdr_id)
            book.write(row, 2, mp3.term_name)
            book.write(row, 3, mp3.language)
            book.write(row, 4, mp3.pronunciation)
            book.write(row, 5, mp3.new_mp3_name)
            book.write(row, 6, mp3.reader_note)
            book.write(row, 7, mp3.reviewer_note)
            if mp3.reuse_media_id:
                book.write(row, 8, mp3.reuse_media_id)
            row += 1
        return book

    @cached_property
    def rejects(self):
        """Audio files which need to be re-done."""

        rejects = []
        for mp3 in self.audio_files:
            if mp3.review_status == "R":
                rejects.append(mp3)
        return rejects

    @cached_property
    def row(self):
        """Database row for set (may have to create it first)."""

        id = self.__opts.get("id")
        if not id:
            name = self.__opts.get("name")
            if not name:
                self.control.bail("No ID or name for MP3 set")
            query = self.control.Query("term_audio_zipfile", "id")
            query.where(query.Condition("filename", name))
            rows = query.execute(self.cursor).fetchall()
            if rows:
                id = rows[0].id
            else:
                try:
                    id = self.__install_set(name)
                except Exception as e:
                    self.control.logger.exception("Installing set")
                    self.control.bail(e)
        query = self.control.Query("term_audio_zipfile", "*")
        query.where(query.Condition("id", id))
        return query.execute(self.cursor).fetchall()[0]

    @cached_property
    def table(self):
        """Display the MP3 files for this set, with decision form fields."""

        B = self.control.HTMLPage.B
        cols = [B.TH(col) for col in self.HEADERS]
        rows = [mp3.row for mp3 in sorted(self.audio_files)]
        classes = "usa-table usa-table--borderless"
        return B.TABLE(B.THEAD(B.TR(*cols)), B.TBODY(*rows), B.CLASS(classes))

    def __install_set(self, name):
        """Add the rows for this set to the database.

        Pass:
            name - string identifying the zipfile for this set

        Return:
            integer for the new term_audio_zipfile row
        """

        diskfile = self.control.zipfile_names.get(name.lower())
        if not diskfile:
            self.control.bail(f"{name} has disappeared")
        with ZipFile(diskfile.path) as zipfile:
            book_name = self.__get_book_name(zipfile)
            if book_name is None:
                self.control.bail("Excel workbook not found")
            with zipfile.open(book_name) as fp:
                book_bytes = fp.read()
        opts = dict(read_only=True, data_only=True)
        book = load_workbook(BytesIO(book_bytes), **opts)
        args = book_name, len(book_bytes)
        self.control.logger.info("opened %s with %d bytes", *args)
        sheet = book.active
        args = sheet.max_row, sheet.max_column
        self.control.logger.info("sheet has %d rows and %d columns", *args)
        args = diskfile.name, diskfile.datetime
        self.control.logger.info("inserting %s, %s", *args)
        self.cursor.execute(
            "INSERT INTO term_audio_zipfile (filename, filedate, complete) "
            "VALUES (?, ?, 'N')", args)
        row = self.cursor.execute("SELECT @@IDENTITY id").fetchone()
        set_id = row.id
        self.control.logger.info("new row id is %d", set_id)
        last_values = [self.control.user_id, self.control.started]
        self.control.logger.info("query: %s", self.INSERT)
        row_number = 1
        for row in sheet:
            column_values = [column.value for column in row]
            if len(column_values) < 8 or column_values[0] == "CDR ID":
                continue
            try:
                values = [set_id, int(column_values[0])]
            except Exception:
                args = row_number, column_values[0]
                self.control.logger.warning("row %d has %r", *args)
                continue
            values += column_values[1:8]
            if values[self.LANGUAGE_COL] not in self.LANGUAGES:
                language = values[self.LANGUAGE_COL]
                message = f"Invalid language {language} in row {row_number}"
                self.control.bail(message)
            values += last_values
            self.control.logger.info("inserting values %s", values)
            self.cursor.execute(self.INSERT, values)
            row_number += 1
        self.control.conn.commit()
        return set_id

    @staticmethod
    def __get_book_name(zipfile):
        """Get the name of the Excel workbook from the zipfile's list."""

        for name in zipfile.namelist():
            if "MACOSX" not in name:
                if name.endswith(".xlsx"):
                    return name
        return None

    class MP3:
        """A glossary term name audio pronunciation file to be reviewed."""

        STATUSES = "Approved", "Rejected", "Unreviewed"
        STATUS_ORDER = dict(U=1, R=2, A=3)
        FIELDS = "review_status", "reviewer_note", "reviewer_id", "review_date"
        FIELDS = ", ".join(f"{name} = ?" for name in FIELDS)
        UPDATE = f"UPDATE term_audio_mp3 SET {FIELDS} WHERE id = ?"

        def __init__(self, control, row):
            """Remember the caller's values.

            Pass:
                control - access to the database and logging
                row - database query result set row with values for this mp3
            """

            self.__control = control
            self.__row = row

        def __getattr__(self, name):
            """Return most properties from the database row."""
            return getattr(self.__row, name)

        def __lt__(self, other):
            """Support sorting of the audio files by status, then doc ID."""
            return self.sortkey < other.sortkey

        def __str__(self):
            """String for debugging/logging."""

            names = AudioSet.FIELDS
            props = [f"{n}={getattr(self, n)}" for n in names]
            props.insert(0, f"id={self.id}")
            props.insert(2, f"review_status={self.review_status}")
            return " ".join(props)

        def update(self, status, note):
            """Save change to status and/or reviewer note.

            Pass:
                status - string for status from the review form (A, R, or U)
                note - string for the reviewer's note (from the form)
            """

            values = [status, note, self.control.user_id, self.control.started]
            args = self.id, values
            self.control.logger.info("mp3 %d updating with values %s", *args)
            args = values + [self.id]
            self.control.cursor.execute(self.UPDATE, args)

        @cached_property
        def control(self):
            """Access to the database and logging."""
            return self.__control

        @cached_property
        def langcode(self):
            """Used for generating a new name for a rejected pronunciation."""
            return "en" if self.language == "English" else "es"

        @cached_property
        def new_mp3_name(self):
            """Name to be added to the spreadsheet of rejected MP3 files."""

            book_name = self.control.audio_set.new_name
            name = f"{self.cdr_id:d}_{self.langcode}"
            n = ""
            if name in self.control.name_counts:
                self.control.name_counts[name] += 1
                n = self.control.name_counts[name]
            else:
                self.control.name_counts[name] = 1
            return f"{book_name}/{name}{n}.mp3"

        @cached_property
        def row(self):
            """Table columns for this MP3 file."""

            Page = self.control.HTMLPage
            B = Page.B
            buttons = []
            review_status = self.review_status.upper() or "U"
            for status in self.STATUSES:
                name = f"status-{self.id:d}"
                opts = dict(value=status[0], label=status)
                if review_status == status[0]:
                    opts["checked"] = True
                buttons.append(Page.radio_button(name, **opts))
            url = self.control.make_url(self.control.script, mp3=self.id)
            note_opts = dict(rows="4", cols="30", name=f"note-{self.id:d}")
            cells = [
                B.TD(*buttons, B.CLASS("status-buttons")),
                B.TD(str(self.cdr_id), B.CLASS("center")),
                B.TD(self.term_name),
                B.TD(self.language, B.CLASS("center")),
                B.TD(self.pronunciation or ""),
                B.TD(B.A(self.mp3_name, href=url, target="_blank")),
                B.TD(self.reader_note or ""),
                B.TD(B.TEXTAREA(self.reviewer_note or "", **note_opts)),
            ]
            return B.TR(*cells)

        @cached_property
        def sortkey(self):
            """Sort by status, then by glossary term ID, then by language."""

            status_order = self.STATUS_ORDER[self.review_status]
            return status_order, self.cdr_id, self.language, self.mp3_name


if __name__ == "__main__":
    """Don't execute script if loaded as a module."""
    Control().run()
