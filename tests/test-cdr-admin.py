#!/usr/bin/env python

"""Test the CDR Administration web pages.

Make sure you're on the NCI network when you run these tests. Otherwise,
the tests will all fail with "unknown error: net::ERR_NAME_NOT_RESOLVED.

You will need to set up a Python (3.10 or higher) environment in which the
third-party modules lxml (4.9.3+), openpyxl (3.1.2+), and selenium (4.16.0+)
are installed. A Python virtual environment is the most effective way to
achieve this. You will also need to install Chrome for Testing (120.0+) from
https://googlechromelabs.github.io/chrome-for-testing/. The most efficient
way to run the tests involves launching Chrome for Testing before invoking
the test script.

Finally, you will need to use a CDR account which has all of the permissions
needed to use every page in the administrative menu system. Any account which
has membership in the Developers group should work.

The DNS name for the CDR servers which are being tested must be provided, as
well as the name of a currently active session on that server. For obvious
reasons, the tests cannot be run against the production server.

$ ./test-cdr-admin.py --help
usage: test-cdr-admin.py [-h] --host HOST --api API --session SESSION
                         [--verbose]
                         [--tests [TESTS ...]]

options:
  -h, --help            show this help message and exit
  --host HOST
  --api API
  --session SESSION
  --verbose, -v
  --tests [TESTS ...], -t [TESTS ...]

Some examples of code which can be added (usually temporarily) when trying
to troubleshoot unexplained failures of new or modified tests:

        self.driver.set_window_size(some_width_integer, some_height_integer)

        self.driver.maximize_window()

        self.driver.save_screenshot("./name-of-page.png")

        self.save_pdf("report.pdf")

It's a good idea to leave the machine on which the tests are running alone
during the run. It is possible to cause some of the tests to fail just by
inadvertently clicking on the browser, thus changing the position of the
cursor/focus (and therefore the behavior of the testing actions). To this
end I prefer to run the tests in a Windows VM, so that I can continue
working on my own local system without risking any disturbance of the tests.

Because these tests rely on outside services over which we have little or no
control, there is no guarantee that all of the tests will pass every time.
If NLM is having a bad day, for example, and a request to PubMed times out,
the best we can do is try again. When this happens, a reasonable attempt is
made to make the test or the CGI script which it is testing more resilient
by detecting the problem at runtime and submitting the request again after a
delay. Then we repeatedly try the test which failed. If we get (say) ten
successful attempts in a row, we'll be satisfied that we've done our best,
with the realization that it's still possible that the test will fail at some
point in the future. The most fragile link in the chain would be the Drupal
CMS servers. We have absolutely no guarantees that no one will start
modifying or rebuilding any of the non-production servers at any time.
Not much we can do about that, aside from cheating and writing the publish
preview test reports to use the production server, but then if we did that,
we wouldn't be testing the same behavior that will go into production.
"""

from argparse import ArgumentParser
from base64 import b64decode
from collections import defaultdict
from datetime import date, datetime, timedelta
from functools import cached_property
from html import escape as html_escape
from io import BytesIO
from json import loads as load_from_json
from logging import getLogger, Formatter, FileHandler
from pathlib import Path
from re import search as re_search, escape as re_escape, sub as re_sub
from re import compile as re_compile
from ssl import _create_unverified_context
from sys import argv
from time import sleep
from unittest import TestCase, TextTestRunner, TextTestResult, main
from urllib.parse import urlencode
from urllib.request import urlopen, Request
from lxml import etree
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select, WebDriverWait


class Tester(TestCase):
    """Base class from which the classes for the test groups are derivced."""

    DEFAULT_WAIT = 20
    HOST = "cdr-dev.cancer.gov"
    API = "cdrapi-dev.cancer.gov"
    SESSION = "guest"
    SUBMIT = "submit-button-submit"
    LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
    LOG_PATH = "test-cdr-admin.log"
    LOGGER = getLogger("cdr-test")
    HANDLER = FileHandler(LOG_PATH, encoding="utf-8")
    HANDLER.setFormatter(Formatter(LOG_FORMAT))
    LOGGER.setLevel("INFO")
    LOGGER.addHandler(HANDLER)
    TEST_DATA = {}
    CDR_ID_PATTERN = r"CDR\d+"
    ISO_DATE_PATTERN = r"\d{4}-\d\d-\d\d"
    DATETIME_PATTERN = rf"{ISO_DATE_PATTERN} \d\d:\d\d:\d\d"
    NS = "cips.nci.nih.gov/cdr"
    NSMAP = dict(cdr=NS)
    STARTED = datetime.now()
    SUCCESSES = FAILURES = ERRORS = 0
    VERBOSE = False
    GOOD_BODY = re_compile("(?s)<body[^>]*>[^<]*<.*</body>")
    del HANDLER, LOG_FORMAT, LOG_PATH

    def setUp(self):
        """This gets run at the start of every test."""

        self.started = datetime.now()
        options = webdriver.ChromeOptions()
        if not self.VERBOSE:
            exclude = ["enable-logging"]
            options.add_experimental_option("excludeSwitches", exclude)
        self.driver = webdriver.Chrome(options=options)
        self.driver.implicitly_wait(self.DEFAULT_WAIT)

    def tearDown(self):
        """This gets run at the end of every test."""

        self.driver.close()
        elapsed = datetime.now() - self.started
        _, test_set, method_name = self.id().split(".")
        args = elapsed, test_set, method_name
        self.logger.info("%s elapsed for %s.%s", *args)

    @cached_property
    def base(self):
        """Root URL for the site."""
        return f"https://{self.host}"

    @cached_property
    def cgi(self):
        """Path to the CGI scripts."""
        return f"{self.base}/cgi-bin/cdr"

    @cached_property
    def doctypes(self):
        """Dictionary of active document type IDs indexed by names."""
        return self.get_test_data("doctypes")

    @cached_property
    def host(self):
        """The DNS name for the instance of IIS serving up the pages."""
        return self.HOST

    @cached_property
    def logger(self):
        """Keep a record of what we do."""
        return self.LOGGER

    @cached_property
    def open_tabs(self):
        """The browser tabs which have been opened by the current test."""
        return set()

    @cached_property
    def session(self):
        """String for the CDR login session's ID."""
        return self.SESSION

    @cached_property
    def wait(self):
        """A WebDriverWait instance, used for interacting with popups."""
        return WebDriverWait(self.driver, timeout=15)

    def add_user_to_group(self, user, group):
        """Add the specified user to a CDR user group.

        Required positional arguments:
          user - name of the CDR user account
          group - string for the name of the CDR user group
        """

        self.navigate_to("EditUser.py", usr=user)
        normalized_group = group.replace(" ", "-").lower()
        id = f"group-{normalized_group}"
        checkbox = self.find(id, method=By.ID)
        if not checkbox.is_selected():
            self.click(id)
            self.click("submit-button-save-changes")

    def assert_table_caption(self, caption):
        """Verify that a specific table caption is present."""

        selector = "table caption span"
        for span in self.driver.find_elements(By.CSS_SELECTOR, selector):
            if span.text == caption:
                return True
        raise Exception(f"table caption {caption} not found")

    def assert_page_has(self, expected):
        """Assert that the source text for the page contains a specific value.

        Required positional argument:
          expected - string we should find
        """

        self.assertIn(str(expected), self.get_page_source())

    def assert_page_not_has(self, expected):
        """The reverse of the "page_has" assertion.

        Required positional argument:
          expected - string we should not find
        """

        self.assertNotIn(expected, self.get_page_source())

    def assert_plain_report(self):
        """Confirm that the report does not use the USWDS framework."""
        self.assertNotIn("uswds.min.css", self.get_page_source())

    def assert_multi_table_report(self):
        """Verify that the report has only one table."""

        tables = self.driver.find_elements(By.CSS_SELECTOR, "table")
        self.assertGreater(len(tables), 1)

    def assert_not_found(self, selector, method=By.TAG_NAME):
        """Verify that no matches are found for a given selector.

        Required positional argument:
          selector tag name, id, or CSS selector

        Optional keyword argument:
          method - defaults to By.TAG_NAME
        """

        self.driver.implicitly_wait(0)
        found = self.driver.find_elements(method, selector)
        self.driver.implicitly_wait(self.DEFAULT_WAIT)
        self.assertEqual(len(found), 0, f"No {selector}s should be present.")

    def assert_non_tabular_report(self):
        """Verify that the report has no tables."""
        self.assert_not_found("table")

    def assert_not_regex(self, expected):
        """Wrapper around the standard unittest assertNotRegex().

        Required positional argument:
          expected - regular expression string to be tested
        """

        self.assertNotRegex(self.get_page_source(), expected)

    def assert_regex(self, expected):
        """Wrapper around the standard unittest assertRegex().

        Required positional argument:
          expected - regular expression string to be tested
        """

        self.assertRegex(self.get_page_source(), expected)

    def assert_single_table_report(self):
        """Verify that the report has only one table."""

        tables = self.driver.find_elements(By.TAG_NAME, "table")
        self.assertEqual(len(tables), 1)

    def assert_tables_in_grid_container(self):
        """Confirm that the report tables are inside the page's grid."""

        selector = "main div.grid-container table"
        table = self.driver.find_element(By.CSS_SELECTOR, selector)
        self.assertIsNotNone(table)

    def assert_title(self, title):
        """Confirm that the title of the page matches the expected value.

        Required positional argument:
          title - string expected to be marked up as the page's title
        """

        self.assert_page_has(f"<h1>{title}</h1>")

    def assert_wide_report(self):
        """Assert that the report is given the full page width."""

        selector = "main > table"
        table = self.driver.find_element(By.CSS_SELECTOR, selector)
        self.assertIsNotNone(table)

    def click(self, field_id):
        """Send a click event to an element.

        Necessary because the USWDS framework is playing with magic
        behind the curtain for radio buttons.

        Required positional argument:
          field_id - string for the button's unique ID
        """

        script = f"""document.getElementById("{field_id}").click();"""
        try:
            self.logger.debug("submitting script %r", script)
            self.driver.execute_script(script)
        except Exception:
            self.logger.exception(f"click({field_id})")
            sleep(2)
            self.logger.warning("submitting %r again ...", script)
            self.driver.execute_script(script)

    def create_external_mapping(self, value, **opts):
        """Add a row to the external mapping table.

        Required positional argument:
          value - string for which a mapping is desired
          doc_id - CDR document ID

        Optional keyword argument:
          usage - defaults to "GlossaryTerm Phrases"
        """

        root = etree.Element("CdrCommandSet")
        etree.SubElement(root, "SessionId").text = self.session
        wrapper = etree.SubElement(root, "CdrCommand")
        command = etree.SubElement(wrapper, "CdrAddExternalMapping")
        usage = opts.get("usage", "GlossaryTerm Phrases")
        etree.SubElement(command, "Usage").text = usage
        etree.SubElement(command, "Value").text = value
        if opts.get("doc_id"):
            etree.SubElement(command, "CdrId").text = str(opts["doc_id"])
        url = f"https://{self.API}"
        request_xml = etree.tostring(root, encoding="utf-8")
        headers = {"Content-type": "application/xml"}
        request = Request(url, data=request_xml, headers=headers)
        response = self.fetch_from_url(request)
        root = etree.fromstring(response)
        self.assertEqual(root.tag, "CdrResponseSet")
        message = "server time for create_external_mapping: %s"
        self.logger.debug(message, root.get("Time"))
        cdr_response = root.find("CdrResponse")
        self.assertIsNotNone(cdr_response)
        elapsed = cdr_response.get("Elapsed")
        status = cdr_response.get("Status")
        message = "create_external_mapping server elapsed time: %s"
        self.logger.debug(message, elapsed)
        self.logger.debug(str(response, "utf-8"))
        if status != "success":
            message = "create_external_mapping response: %s"
            self.logger.error(message, str(response, "utf-8"))
        self.assertEqual(status, "success")
        mapping_response = cdr_response.find("CdrAddExternalMappingResp")
        mapping_id = mapping_response.get("MappingId")
        self.assertTrue(mapping_id and mapping_id.isdigit())
        self.logger.debug("created mapping %s", mapping_id)
        return int(mapping_id)

    def create_test_gtn(self, **opts):
        """Create a new GlossaryTernName document.

        Optional keyword arguments:
          name - defaults to "Test GTN Document"

        Return:
          string for the CDR ID (in canonical CDR9999999999 format)
        """

        root = etree.Element("GlossaryTermName", nsmap=self.NSMAP)
        term_name = etree.SubElement(root, "TermName")
        name = opts.get("name", "Test GTN Document")
        etree.SubElement(term_name, "TermNameString").text = name
        xml = etree.tostring(root, encoding="unicode")
        doctype = "GlossaryTermName"
        return self.save_doc(xml, doctype, unlock=True, version=True)

    def create_test_media_doc(self, **opts):
        """Create a new Media document.

        Optional keyword arguments:
          title - defaults to "Test Media Document"

        Return:
          string for the CDR ID (in canonical CDR9999999999 format)
        """

        root = etree.Element("Media", nsmap=self.NSMAP)
        title = opts.get("title", "Test Media Document")
        etree.SubElement(root, "MediaTitle").text = title
        xml = etree.tostring(root, encoding="unicode")
        doctype = "Media"
        return self.save_doc(xml, doctype, unlock=True, version=True)

    def create_test_summary(self, **opts):
        """Create a new English summary.

        Optional keyword arguments:
          title - defaults to "Test English Summary"
          description - defaults to "yada, yada"
          svpc - add SVPC attribute if True
          module - add ModuleOnly attribute if True

        Return:
          string for the CDR ID (in canonical CDR9999999999 format)
        """

        title = opts.get("title", "Test English Summary")
        description = opts.get("description", "yada, yada")
        root = etree.Element("Summary", nsmap=self.NSMAP)
        if opts.get("svpc"):
            root.set("SVPC", "Yes")
        if opts.get("module"):
            root.set("ModuleOnly", "Yes")
        metadata = etree.SubElement(root, "SummaryMetaData")
        etree.SubElement(metadata, "SummaryType").text = "Treatment"
        etree.SubElement(metadata, "SummaryAudience").text = "Patients"
        etree.SubElement(metadata, "SummaryLanguage").text = "English"
        etree.SubElement(metadata, "SummaryDescription").text = description
        etree.SubElement(root, "SummaryTitle").text = title
        xml = etree.tostring(root, encoding="unicode")
        return self.save_doc(xml, "Summary", unlock=True, version=True)

    def delete_doc(self, cdr_id, reason=None):
        """Delete a test document.

        Required positional argument:
          cdr_id - unique ID for the document

        Optional positional argument:
          reason - defaults to "Deleted by automated test"
        """

        self.logger.debug("delete_doc(%s)", cdr_id)
        self.navigate_to("del-some-docs.py")
        self.assert_title("CDR Document Deletion")
        self.set_field_value("ids", cdr_id)
        self.set_field_value("reason", reason)

        # The default for validation is ON, so we're turning it OFF.
        self.click("options-validate")
        self.submit_form(new_tab=False)
        self.assert_title("CDR Document Deletion")
        cdr_id = f"CDR{self.extract_id(cdr_id):010d}"
        self.assert_page_has(f"{cdr_id} has been deleted successfully.")

    def extract_id(self, cdr_id):
        """Get the integer ID from the caller's value.

        Required positional argument:
          cdr_id - string for the CDR ID (or possibly already an integer)

        Return:
          integer ID extracted from the caller's value
        """

        if isinstance(cdr_id, str):
            cdr_id = re_sub(r"[^\d]+", "", cdr_id)
        return int(cdr_id)

    def fetch_from_url(self, url):
        """Submit an HTTP request and return the bytes of the response.

        Required positional argument:
          url - string for the request (possibly wrapped in a Request object)

        Return:
          bytes returned from the HTTP server
        """

        try:
            context = _create_unverified_context()
            with urlopen(url, context=context) as response:
                return response.read()
        except Exception:
            self.logger.exception("failure fetching from %s", url)
            return None

    def fetch_workbook(self, script_or_url, params=None, save=False):
        """Get an Excel report from the web site.

        Required positional argument:
          script - string for the filename of the script, or a complete URL

        Optional keyword arguments:
          params - CGI named parameters needed for the test (dict or tuples)
          save - if True, save response as response.html for debugging

        Return:
          object for the Excel workbook
        """

        if params is None:
            url = script_or_url
        else:
            params = params or []
            if isinstance(params, dict):
                params["Session"] = self.session
            else:
                has_session = False
                for name, value in params:
                    if name == "Session":
                        has_session = True
                        break
                if not has_session:
                    params.append(("Session", self.session))
            url = f"{self.cgi}/{script_or_url}?{urlencode(params)}"
        self.logger.debug("loading workbook from %s", url)
        content = self.fetch_from_url(url)
        if content:
            if save:
                with open("response.html", "wb") as fp:
                    fp.write(content)
            try:
                return load_workbook(filename=BytesIO(content))
            except Exception:
                self.logger.exception(url)
                if not save:
                    with open("excel-failure.html", "wb") as fp:
                        fp.write(content)
                raise
        return None

    def find(self, selector, method=By.CSS_SELECTOR, all=False):
        """Find matching element(s).

        Required positional argument:
          selector - string identifying which elements we want

        Optional keyword arguments:
          method - default is By.CSS_SELECTOR
          all - set to True to return all matching elements
        """

        if all:
            return self.driver.find_elements(method, selector)
        return self.driver.find_element(method, selector)

    def find_test_citation_docs(self):
        """Find test Citation documents created from PMID 1."""

        sql = (
            "SELECT doc_id"
            "  FROM query_term"
            " WHERE path = '/Citation/PubmedArticle/MedlineCitation/PMID'"
            "   AND int_val = 1"
        )
        return [row[0] for row in self.run_query(sql)]

    def get_page_source(self):
        """Wait for the page to be completely loaded."""

        attempts = 3
        sleep(2)
        while attempts > 0:
            source = self.driver.page_source
            if "</body>" in source and "</html>" in source:
                if self.GOOD_BODY.search(source):
                    return source
            attempts -= 1
            args = source, attempts
            self.logger.warning("body missing in %s, %d tries left", *args)
            if attempts:
                sleep(3 - attempts)
        raise Exception("body element never found")

    def get_test_board(self, **opts):
        """Get a sample test board's information.

        Optional keyword arguments:
          type - default is "editorial"

        Return:
          dictionary of information about the requested test board
        """

        boards = self.get_test_data("boards")
        board_type = opts.get("type", "editorial")
        return boards[board_type][0]

    def get_test_board_member(self):
        """Get a sample test board member's information.

        Return:
          dictionary of information about the requested test board member
        """

        return self.get_test_data("board-members")[0]

    def get_test_citation(self):
        """Get a sample Citation document's information."""
        return self.get_test_data("citations")[0]

    def get_test_data(self, data_type):
        """Get the test values for a specific data type.

        We do this so we don't have to hard-code test values (which
        might change). If we don't have the values, fetch them.

        Required positional argument:
          data_type - for example, "summaries"
        """

        if data_type not in self.TEST_DATA:
            if data_type == "modules":
                url = f"{self.cgi}/get-summaries.py?modules-only=true&limit=2"
            else:
                url = f"{self.cgi}/get-{data_type}.py?limit=3"
            content = self.fetch_from_url(url)
            if not content:
                raise Exception(f"failure loading test {data_type}")
            try:
                self.TEST_DATA[data_type] = load_from_json(content)
            except Exception:
                self.logger.exception("loading %s", data_type)
                with open(f"{data_type}.json", "wb") as fp:
                    fp.write(content)
        return self.TEST_DATA[data_type]

    def get_test_drug_info(self):
        """Get a sample test DrugInformationSummary document's ID and title.

        Return:
          dictionary of information about the requested test drug
        """

        drugs = self.get_test_data("drug-info")
        return drugs[0]

    def get_test_glossary_term(self):
        """Get a sample test glossary term.

        Return:
          dictionary of information about a CDR glossary term
        """

        return self.get_test_data("glossary-terms")[0]

    def get_test_media_doc(self, **opts):
        """Get a sample test Media document's information.

        Optional keyword arguments:
          language - default is "en"
          format - default is "JPEG"

        Return:
          dictionary of information about the requested test media doc
        """

        docs = self.get_test_data("media")
        language = opts.get("language", "en")
        format = opts.get("format", "JPEG")
        return docs[language][format][0]

    def get_test_summary(self, **opts):
        """Get a sample test summary's information.

        Optional keyword arguments:
          language - default is "English"
          type - default is "Treatment"
          audience - default is "Health professionals"
          module - if True get module instead of standalong summary

        Return:
          dictionary of information about the requested test summary
        """

        group = "modules" if opts.get("module") else "summaries"
        docs = self.get_test_data(group)
        language = opts.get("language", "English")
        summary_type = opts.get("type", "Treatment")
        audience = opts.get("audience", "Health professionals")
        return docs[language][summary_type][audience][0]

    def get_test_term(self, **opts):
        """Get a sample test Term document's information.

        Optional keyword argument:
          type - default is 'Index term'

        Return:
          dictionary with term ID and name
        """

        type = opts.get("type") or "Index term"
        return self.get_test_data("terms")[type][0]

    def load_table(self, selector=None):
        """Load the first table matching the selector.

        Required positional argument:
          selector - CSS selector (if None, find the first table on the page)

        Return:
          Table object
        """

        if selector is None:
            node = self.driver.find_element(By.TAG_NAME, "table")
        else:
            node = self.driver.find_element(By.CSS_SELECTOR, selector)
        return self.Table(self, node)

    def load_tables(self, selector=None):
        """Load all tables matching the selector.

        Required positional argument:
          selector - CSS selector (if None, find all tables on the page)

        Return:
          sequence of Table objects
        """

        if selector is None:
            nodes = self.driver.find_elements(By.TAG_NAME, "table")
        else:
            nodes = self.driver.find_elements(By.CSS_SELECTOR, selector)
        return [self.Table(self, node) for node in nodes]

    def navigate_to(self, script, **params):
        """Open the page for a given CGI script.

        Required positional argument:
          script - string for the filename of the script

        Optional keyword arguments:
          params - CGI named parameters needed for the test

        Return:
          string for the handle of the browser tab for the page
        """

        params["Session"] = self.session
        url = f"{self.cgi}/{script}?{urlencode(params)}"
        self.logger.debug("navigating to %s", script)
        self.driver.get(url)
        handle = self.driver.current_window_handle
        self.open_tabs.add(handle)
        sleep(1)
        return handle

    def remove_user_from_group(self, user, group):
        """Remove the specified user from a CDR user group.

        Required positional arguments:
          user - name of the CDR user account
          group - string for the name of the CDR user group
        """

        self.navigate_to("EditUser.py", usr=user)
        normalized_group = group.replace(" ", "-").lower()
        id = f"group-{normalized_group}"
        checkbox = self.find(id, method=By.ID)
        if checkbox.is_selected:
            self.click(id)
        self.click("submit-button-save-changes")

    def run_query(self, sql):
        """Execute a SQL query and return the results.

        Required positional argument:
          sql - string for the query to execute

        Return:
          2-dimensional array of row/column values
        """

        self.navigate_to("CdrQueries.py", sql=sql, Request="Run")
        results = []
        if "<p>Retrieved 0 rows in" in self.get_page_source():
            return results
        table = self.load_table()
        for row in table.rows:
            results.append([col.text for col in row])
        return results

    def save_doc(self, xml, doctype, **opts):
        """Create or update a CDR document.

        Required positional arguments:
          xml - the serialized document to be saved
          doctype - string naming the document's type

        Optional keyword arguments:
          id - an integer if an existing document is being updated; else None
          title - string of the document's title
          unlock - if True, release the lock
          block - if True, mark the document as inactive
          version - if True, create a new version
          publishable - if True, make the version publishable (implies version)
          validate - if True, validate the document
          comment - defaults to "saved by an automated test"

        Return:
          integer ID for the summary
        """

        root = etree.Element("CdrCommandSet")
        etree.SubElement(root, "SessionId").text = self.session
        wrapper = etree.SubElement(root, "CdrCommand")
        tag = "CdrRepDoc" if opts.get("id") else "CdrAddDoc"
        command = etree.SubElement(wrapper, tag)
        unlock = "Y" if opts.get("unlock") else "N"
        etree.SubElement(command, "CheckIn").text = unlock
        version = etree.SubElement(command, "Version")
        if opts.get("version") or opts.get("publishable"):
            version.text = "Y"
            publishable = True if opts.get("publishable") else False
            version.set("Publishable", "Y" if publishable else "N")
        else:
            version.text = "N"
        validate = etree.SubElement(command, "Validate")
        validate.text = "Y" if opts.get("validate") else "N"
        cdr_doc = etree.SubElement(command, "CdrDoc")
        cdr_doc.set("Type", doctype)
        control = etree.SubElement(cdr_doc, "CdrDocCtl")
        if opts.get("id"):
            cdr_doc.set("Id", opts["id"])
            etree.SubElement(control, "DocId").text = opts["id"]
        title = opts.get("title", "Test Document For Automated Testing")
        etree.SubElement(control, "DocType").text = doctype
        etree.SubElement(control, "DocTitle").text = title
        if opts.get("block"):
            etree.SubElement(control, "DocActiveStatus").text = "I"
        if opts.get("comment"):
            etree.SubElement(command, "Reason").text = opts["comment"]
            etree.SubElement(control, "DocComment").text = opts["comment"]
        etree.SubElement(cdr_doc, "CdrDocXml").text = etree.CDATA(xml)
        url = f"https://{self.API}"
        request_xml = etree.tostring(root, encoding="utf-8")
        headers = {"Content-type": "application/xml"}
        request = Request(url, data=request_xml, headers=headers)
        response = self.fetch_from_url(request)
        root = etree.fromstring(response)
        self.assertEqual(root.tag, "CdrResponseSet")
        self.logger.debug("server time for save_doc: %s", root.get("Time"))
        cdr_response = root.find("CdrResponse")
        self.assertIsNotNone(cdr_response)
        elapsed = cdr_response.get("Elapsed")
        status = cdr_response.get("Status")
        self.logger.debug("save_doc server elapsed time: %s", elapsed)
        if status != "success":
            self.logger.error("save_doc response: %s", str(response, "utf-8"))
        self.assertEqual(status, "success")
        if opts.get("id"):
            self.logger.debug("updated %s", opts["id"])
            return opts["id"]
        doc_id = cdr_response.find("CdrAddDocResp/DocId")
        self.assertIsNotNone(doc_id)
        self.logger.debug("created %s", doc_id.text)
        return self.extract_id(doc_id.text)

    def save_pdf(self, filename="page.pdf"):
        """Create a PDF file of the current page.

        Optional keyword argument:
          filename - defaults to 'page.pdf'
        """

        with open(filename, "wb") as fp:
            fp.write(b64decode(self.driver.print_page()))

    def select_new_tab(self):
        """Move to the browser tab which was most recently opened.

        This is accomplished by keeping track of the browser tabs
        which were opened previously by the current test, and finding
        the tab whose handle isn't in that set. Of course we add
        the handle of the new tab to the set so this method will
        work correctly the next time we invoke it.

        We also take care of waiting until the browser catches up
        with the automated test.
        """

        for handle in self.driver.window_handles:
            if handle not in self.open_tabs:
                self.open_tabs.add(handle)
                self.driver.switch_to.window(handle)
                return handle
        self.logger.warning("can't find new browser tab; trying again")
        sleep(2)
        for handle in self.driver.window_handles:
            if handle not in self.open_tabs:
                self.open_tabs.add(handle)
                self.driver.switch_to.window(handle)
                return handle
        self.logger.warning("still can't find new browser tab; giving up")
        return None

    def select_values(self, field_id, *values):
        """Select a value from a picklist.

        Required positional arguments:
          field_id - string uniquely identifying the picklist field
          value - string for the option we want to select
        """

        select = Select(self.driver.find_element(By.ID, field_id))
        for value in values:
            select.select_by_value(str(value))

    def select_version(self, text, id="DocVersion"):
        """Find and select the most recent version with matching comment text.

        Required positional parameter:
          text - string to look for the in the display for the picklist option

        Optional keyword parameter:
          id - unique ID for the picklist

        Return:
          True if option found and selected; else False
        """

        select = Select(self.driver.find_element(By.ID, id))
        for option in select.options:
            if text in option.text:
                option.click()
                return True
        return False

    def set_field_value(self, field_id, value):
        """Replace a form field's value.

        Required positional arguments:
          field_id - string for the field's unique ID
          value - the new value to be associated with the field
        """

        field = self.driver.find_element(By.ID, field_id)
        if not field:
            self.logger.warning("can't find field %s", field_id)
        else:
            field.clear()
            field.send_keys(str(value))

    def submit_form(self, new_tab=True):
        """Submit the form on the current page.

        Optional keyword argument:
          new_tab - set to False if we expect to say in the same browser tab
        """

        self.driver.find_element(By.ID, self.SUBMIT).click()
        return self.select_new_tab() if new_tab else None

    def switch_to(self, tab, wait=DEFAULT_WAIT):
        """Make another browser tab the currently visible tab.

        Required positional argument:
          tab - string for the browser tab's handle
          wait - override for wait timeout
        """

        self.driver.switch_to.window(tab)
        self.logger.debug("switch_to(%s)", tab)

    @classmethod
    def setUpClass(cls):
        cls._started = datetime.now()
        cls.LOGGER.info("starting %s", cls.__name__)

    @classmethod
    def tearDownClass(cls):
        now = datetime.now()
        cls.LOGGER.info("%s elapsed for %s", now - cls._started, cls.__name__)
        if Tester.ERRORS or Tester.FAILURES:
            args = Tester.SUCCESSES, Tester.ERRORS, Tester.FAILURES
            message = "interim counts: succeeded=%d errors=%d failures=%d"
            Tester.LOGGER.info(message, *args)

    class Table:
        """Collects the body cells for the node's table."""

        def __init__(self, tester, node):
            """Remember the caller's values.

            Required positional argument:
              node - DOM node for the table
            """

            self.tester = tester
            self.node = node

        def check_headers(self, headers):
            """Verify that the table has the expected column headers.

            Required positional argument:
              columns - sequence of strings for expected column headers
            """

            for i, header in enumerate(headers):
                self.tester.assertEqual(self.headers[i].text, header)

        @cached_property
        def caption(self):
            """The CAPTION element of the table."""
            return self.node.find_element(By.TAG_NAME, "caption")

        @cached_property
        def headers(self):
            """Array of TH elements."""

            headers = []
            for row in self.node.find_elements(By.CSS_SELECTOR, "thead tr"):
                headers += row.find_elements(By.TAG_NAME, "th")
            return headers

        @cached_property
        def rows(self):
            """Two-dimensional array of TD elements."""

            rows = []
            for row in self.node.find_elements(By.CSS_SELECTOR, "tbody tr"):
                rows.append(row.find_elements(By.TAG_NAME, "td"))
            return rows


class CitationTests(Tester):
    """Tests of reports on Citation documents."""

    def test_advanced_search(self):
        """Test Citation Advanced Search.

        We are unable to test searching PubMed, because NLM refuses to
        allow our test client to connect.
        """

        # First thing to do is to clear out old test articles laying around.
        for doc_id in self.find_test_citation_docs():
            self.delete_doc(doc_id)

        # Test refresh of a CDR Citation document from PubMed.
        doc = self.get_test_citation()
        article_title = doc["title"]
        first_word = article_title.split()[0]
        pmid = doc["pmid"]
        doc_id = int(doc["id"])
        journal = doc["journal"]
        pub_year = doc["year"]
        self.navigate_to("CiteSearch.py")
        self.assert_title("Citation")
        button = self.find("submit-button-import", method=By.ID)
        self.assertIsNotNone(button)
        self.assertFalse(button.is_enabled())
        self.set_field_value("pmid", pmid)
        self.assertTrue(button.is_enabled())
        self.assertEqual(button.get_attribute("value"), "Import")
        self.set_field_value("cdrid", doc_id)
        self.assertTrue(button.is_enabled())
        self.assertEqual(button.get_attribute("value"), "Update")
        button.click()
        cdr_id = f"CDR{doc_id:010d}"
        self.assert_page_has(
            f"Updated {cdr_id} from PMID {pmid}"
            " (with a publishable version)"
        )

        # Test advanced search.
        self.set_field_value("title", f"{first_word}%")
        self.set_field_value("pub_in", journal)
        self.set_field_value("pub_year", pub_year)
        self.find("submit-button-search", method=By.ID).click()
        self.select_new_tab()
        self.assert_title("Citation")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        pattern = (
            rf"\d+ documents? match(es)? '{first_word}%' "
            f"and '{journal}' and '{pub_year}'"
        )
        table = self.load_table()
        self.assertRegex(table.caption.text, pattern)
        link = None
        for i, row in enumerate(table.rows, start=1):
            self.assertEqual(row[0].text, f"{i}.")
            candidate = row[1].find_element(By.TAG_NAME, "a")
            self.assertIsNotNone(candidate)
            if candidate.text == cdr_id:
                link = candidate
                self.assertEqual(row[2].text, article_title)
        self.assertIsNotNone(link)
        link.click()
        self.assert_plain_report()
        self.assert_page_has("<center>Citations<br>QC Report</center>")
        self.assert_page_has(f"CDR{doc_id}")
        self.assert_page_has(article_title)
        self.assert_page_has(journal)
        self.assert_page_has(pub_year)

        # Test import of a new Citation document.
        self.navigate_to("CiteSearch.py")
        self.assert_title("Citation")
        self.set_field_value("pmid", "1")
        self.click("submit-button-import")
        regex = r"Imported PMID 1 as CDR(\d+) \(with a publishable version\)"
        re_match = re_search(regex, self.get_page_source())
        self.assertIsNotNone(re_match)
        doc_id = int(re_match.group(1))

        # Attempt to import again (rather than refresh) should fail.
        self.set_field_value("pmid", "1")
        self.click("submit-button-import")
        expected = "Unable to import '1' from PubMed: PMID 1 already imported"
        self.assert_page_has(expected)

        # Don't leave test article behind.
        self.delete_doc(doc_id)

    def test_citations_in_summaries_report(self):
        """Test the Citations In Summaries report."""

        script = "CitationsInSummaries.py"
        self.navigate_to(script)
        self.assert_title("Citations In Summaries")
        self.assert_page_has("Instructions")
        self.assert_page_has("The full report generally takes about an hour")
        params = [
            ("options", "quick"),
            ("max-citations", "10"),
            ("max-seconds", "5"),
            ("Request", "Submit"),
        ]
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "Citations In Summaries")
        self.assertEqual(sheet["A1"].value, "Citations In Summaries")
        self.assertEqual(sheet["A3"].value, "Citation ID")
        self.assertEqual(sheet["B3"].value, "Citation Title")
        self.assertEqual(sheet["C3"].value, "PMID")
        self.assertEqual(sheet["D3"].value, "Summary ID")
        self.assertEqual(sheet["E3"].value, "Summary Title")
        self.assertEqual(sheet["F3"].value, "Summary Boards")

    def test_new_citations_report(self):
        """Test the New Citations report."""

        self.navigate_to("NewCitations.py")
        self.assert_title("New Citations Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title("New Citations Report")
        self.assert_plain_report()
        self.assert_single_table_report()
        table = self.load_table()
        expected = r"\d+ Documents Created Between 2020-01-01 and 2020-01-31"
        self.assertRegex(table.caption.text, expected)
        columns = (
            "CDR ID",
            "Document Title",
            "Created By",
            "Creation Date",
            "Last Version Pub?",
            "PMID",
        )
        table.check_headers(columns)

    def test_qc_report(self):
        """Test the Citation QC report."""

        self.navigate_to("QcReport.py", DocType="Citation")
        self.assert_title("QC Report")
        self.assert_page_has("Title or Document ID")
        self.set_field_value("DocTitle", "brain%")
        self.submit_form()
        self.assert_title("QC Report")
        self.assert_page_has("Multiple matches found for 'brain%'.")
        self.assert_page_has("Choose Document")
        radio_button = self.find('input[name="DocId"]')
        self.assertIsNotNone(radio_button)
        button_id = radio_button.get_attribute("id")
        self.click(button_id)
        self.submit_form()
        self.assert_plain_report()
        self.assert_page_has("<center>Citations<br>QC Report</center>")

    def test_unverified_citations_report(self):
        """Test the Unverified Citations report."""

        self.navigate_to("UnverifiedCitations.py")
        self.assert_title("Unverified Citations Report")
        self.assert_page_has("Instructions")
        self.assert_page_has("Press Submit to generate an HTML table")
        self.submit_form()
        self.assert_title("Unverified Citations Report")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        table.check_headers(("Doc ID", "Citation", "Comment"))

    def test_updates(self):
        """Test the Update Pre-Medline Citations utility."""

        # First thing to do is to clear out old test articles laying around.
        sql = (
            "SELECT doc_id"
            "  FROM query_term"
            " WHERE path = '/Citation/PubmedArticle/MedlineCitation/PMID'"
            "   AND int_val = 1"
        )
        for doc_id in [row[0] for row in self.run_query(sql)]:
            self.delete_doc(doc_id)

        # Create a test Citation document linked to PMID 1.
        root = etree.Element("Citation")
        verification_details = etree.SubElement(root, "VerificationDetails")
        etree.SubElement(verification_details, "Verified").text = "Yes"
        etree.SubElement(verification_details, "VerifiedIn").text = "PubMed"
        article = etree.SubElement(root, "PubmedArticle")
        citation = etree.SubElement(article, "MedlineCitation")
        citation.set("Status", "In-Process")
        etree.SubElement(citation, "PMID").text = "1"
        xml = etree.tostring(root, encoding="unicode")
        doc_id = self.save_doc(xml, "Citation", unlock=True, version=True)
        self.logger.debug("created CDR%s linked to PMID 1", doc_id)

        # Test the utility.
        self.navigate_to("UpdatePreMedlineCitations.py")
        self.assert_title("Citation Status Changes")
        self.assert_page_has("Instructions")
        self.assert_page_has("This utility checks to see if any")
        self.click("submit-button-check")
        self.assert_page_has("Available For Refresh")
        self.assert_page_has(
            f"Citation 1 (CDR{doc_id}): "
            "status In-Process will become MEDLINE."
        )
        self.click("submit-button-update")
        self.select_new_tab()
        self.assert_title("Citation Status Changes")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        columns = "PMID", "CDR ID", "Old Status", "New Status", "Notes"
        table.check_headers(columns)
        test_row = None
        for row in table.rows:
            if row[0].text == "1":
                test_row = row
                break
        self.assertIsNotNone(test_row)
        self.assertEqual(test_row[0].text, "1")
        self.assertEqual(test_row[1].text, str(doc_id))
        self.assertEqual(test_row[2].text, "In-Process")
        self.assertEqual(test_row[3].text, "MEDLINE")
        self.assertEqual(test_row[4].text, "updated")

        # Don't leave test trash lying around.
        self.delete_doc(doc_id)


class DeveloperTests(Tester):
    """Tests of tools used by the development tesam."""

    def test_client_file_tools(self):
        """Test the client file tools."""

        # Test installation of a client file.
        file_bytes = b"This is a test file\n"
        path = Path("dada.txt").resolve()
        path.write_bytes(file_bytes)
        self.navigate_to("InstallClientFile.py")
        self.assert_title("Install Client File")
        self.assert_page_has("Instructions")
        self.assert_page_has("Select the file to be installed")
        self.set_field_value("file", str(path))
        self.set_field_value("location", "dada.txt")
        self.submit_form(new_tab=False)
        self.assert_title("Install Client File")
        expected = r"D:\cdr\ClientFiles\dada.txt successfully installed"
        self.assert_page_has(expected)

        # Fetch the file.
        script = "FetchClientFile.py"
        self.navigate_to(script)
        self.assert_title("Fetch Client File")
        params = dict(path="dada.txt", Request="Submit")
        url = f"{self.cgi}/{script}?{urlencode(params)}"
        data = self.fetch_from_url(url)
        self.assertEqual(data, file_bytes)

        # Remove the file.
        self.navigate_to("RemoveClientFile.py")
        self.assert_title("Remove Client File")
        self.select_values("path", "dada.txt")
        self.submit_form(new_tab=False)
        self.assert_title("Remove Client File")
        self.assert_page_has("Processing Logs")
        self.assert_page_has(r"Removed D:\cdr\ClientFiles\dada.txt")
        self.assert_page_has("Running RefreshManifest.py ...")
        self.assert_page_has("fixing permissions...")
        self.assert_page_has("File removed successfully.")
        path.unlink()

    def test_configuration(self):
        """Test the interface for managing server configuration."""

        # Bring up the page and switch to the ports values.
        self.navigate_to("EditConfig.py")
        self.assert_title("CDR Settings")
        self.select_values("filename", "dbports")
        sleep(1)
        original = self.find("content", method=By.ID).text
        self.assertRegex(original, r"PROD:cdr:\d+")

        # Make a change and save it.
        insertion = "# This is a comment for automated testing.\n"
        self.set_field_value("content", insertion + original)
        self.click("submit-button-save")
        self.assert_title("CDR Settings")
        self.assert_page_has("Saved new values")
        modified = self.find("content", method=By.ID).text
        self.assertEqual(modified, insertion + original)

        # Back out the change.
        self.set_field_value("content", original)
        self.click("submit-button-save")
        self.assert_title("CDR Settings")
        self.assert_page_has("Saved new values")
        final = self.find("content", method=By.ID).text
        self.assertEqual(final, original)

    def test_control_values(self):
        """Test the utility for managing control values."""

        # Before we do anything else, make sure nothing was left behind
        # from a previous, failed test.
        form = self.navigate_to("EditControlValues.py")
        self.assert_title("Manage Control Values")
        self.assert_page_has("Instructions")
        self.assert_page_has(
            "Enter a group or value name (or both) in the New Value block to "
            "override where the value (and its comment) will be stored when "
            "the Save button is clicked."
        )
        test_group = "test-control-value-group"
        test_names = "test-control-value-name-1", "test-control-value-name-2"
        test_values = "dada", "yada, yada"
        test_comments = "This is a test control value", "Another test value"

        def load_values():
            params = dict(Session=self.session, Request="JSON")
            url = f"{self.cgi}/EditControlValues.py?{urlencode(params)}"
            json = self.fetch_from_url(url)
            return load_from_json(json)
        values = load_values()
        while test_group in values:
            self.select_values("group", test_group)
            self.click("submit-button-delete")
            self.assert_page_has("Value successfully dropped.")
            values = load_values()

        # Create some test values.
        for i, name in enumerate(test_names):
            self.set_field_value("new_group", test_group)
            self.set_field_value("new_name", name)
            self.set_field_value("value", test_values[i])
            self.set_field_value("comment", test_comments[i])
            self.click("submit-button-save")

        # Verify the values using the HTML table report.
        self.click("submit-button-show-all-values")
        self.select_new_tab()
        tables = self.load_tables()
        self.assertEqual(len(tables), 1)
        value_dictionary = defaultdict(dict)
        for row in tables[0].rows:
            (value, comment) = row[2].text, row[3].text
            value_dictionary[row[0].text][row[1].text] = (value, comment)
        test_group_values = value_dictionary[test_group]
        for i, name in enumerate(test_names):
            self.assertEqual(test_group_values[name][0], test_values[i])
            self.assertEqual(test_group_values[name][1], test_comments[i])

        # Change the values.
        new_test_values = "dada2", "yadissimo"
        new_test_comments = "I've changed my mind.", "Won't be around long."
        self.switch_to(form)
        for i, name in enumerate(test_names):
            self.select_values("group", test_group)
            self.select_values("name", name)
            self.set_field_value("value", new_test_values[i])
            self.set_field_value("comment", new_test_comments[i])
            self.click("submit-button-save")

        # Verify the values using the JSON report.
        group = load_values().get(test_group)
        self.assertIsNotNone(group)
        values = group["values"]
        for i, name in enumerate(test_names):
            self.assertEqual(values[name]["value"], new_test_values[i])
            self.assertEqual(values[name]["comment"], new_test_comments[i])

        # Remove the test control values.
        for i, name in enumerate(test_names):
            self.select_values("group", test_group)
            self.select_values("name", name)
            self.click("submit-button-delete")

        # Verify that they're gone.
        group = load_values().get(test_group)
        self.assertIsNone(group)

    def test_database_table_report(self):
        """Test the Database Tables and Views report."""

        self.navigate_to("db-tables.py")
        self.assert_title("Database Tables and Views")
        databases = "CDR", "CDR_ARCHIVED_VERSIONS"
        for db in databases:
            self.assert_page_has(f'<legend class="usa-legend">{db}</legend>')
        self.assert_page_has("<h3>TABLES</h3>")
        self.assert_page_has("<dt>all_docs</dt>")

    def test_doctypes(self):
        """Test the interface for managing CDR document types."""

        # Establish values used by the test.
        test_type_name = "DoctypeForRegressionTesting"
        del_msg = f"Successfully deleted document type {test_type_name!r}."
        add_msg = f"New document type {test_type_name!r} successfully added."

        # Create a local function we'll use more than once.
        def find_test_type_link():
            for link in self.find("form ul li a", all=True):
                if link.text == test_type_name:
                    return link
            return None

        # Clean out any leftover test data.
        self.navigate_to("EditDocTypes.py")
        self.assert_title("Manage Document Types")
        self.assert_page_has("Existing Document Types (click to edit)")
        link = find_test_type_link()
        if link is not None:
            link.click()
            self.select_new_tab()
            self.assert_title(f"Edit {test_type_name} Document Type")
            self.click("submit-button-delete-document-type")
            self.assert_title("Manage Document Types")
            self.assert_page_has(del_msg)

        # Create the test document type.
        self.click("submit-button-add-new-document-type")
        if link is None:
            self.select_new_tab()
        self.assert_title("Adding New Document Type")
        self.set_field_value("doctype", test_type_name)
        self.select_values("schema", "xxtest.xml")
        self.select_values("title_filter", "DocTitle for xxtest")
        self.set_field_value("comment", "this is a bogus test document type")
        self.click("submit-button-save-new-document-type")
        self.assert_title(f"Edit {test_type_name} Document Type")
        self.assert_page_has(add_msg)
        self.click("submit-button-document-type-menu")
        link = find_test_type_link()
        self.assertIsNotNone(link)

        # Remove the test type.
        link.click()
        self.assert_title(f"Edit {test_type_name} Document Type")
        self.click("submit-button-delete-document-type")
        self.assert_title("Manage Document Types")
        self.assert_page_has(del_msg)
        self.assertIsNone(find_test_type_link())

    def test_dtds(self):
        """Test the interface for posting a DTD to the CDR server."""

        # We'll need this a couple of times.
        def fetch_dtd(filename):
            url = f"{self.cgi}/get-dtd.py?dtd={filename}"
            return self.fetch_from_url(url).strip().replace(b"\r", b"") + b"\n"

        # Get a copy of the DTD and save it to disk.
        original_dtd = fetch_dtd("pdq.dtd")
        self.assertIn(b"<!ELEMENT", original_dtd)
        path = Path("pdq.dtd").resolve()
        path.write_bytes(original_dtd)

        # Post it to the server.
        self.navigate_to("PostDTD.py")
        self.assert_title("Post DTD")
        self.set_field_value("file", str(path))
        self.click("flavor-vendor")
        self.submit_form(new_tab=False)
        self.assert_title("Post DTD")
        self.assert_page_has("Successfully installed D:/cdr/licensee/pdq.dtd.")

        # Verify that it made the round trip intact.
        fetched_dtd = fetch_dtd("pdq.dtd")
        self.assertEqual(fetched_dtd, original_dtd)

        # Clean up after ourselves.
        path.unlink()

    def test_elements_report(self):
        """Test the Summary Elements report."""

        self.navigate_to("ShowSummaryIncludes.py")
        self.assert_title("Elements Included In PDQ Summaries")
        self.assert_page_has("Choose summary type")
        self.click("doctype-dis")
        self.submit_form()
        self.assert_title("Elements Included In PDQ Summaries")
        self.assert_plain_report()
        self.assert_non_tabular_report()
        self.assert_page_has("<h2>Report Elements</h2>")
        self.assert_page_has("<h2>Summaries</h2>")
        self.assertIsNotNone(self.driver.find_element(By.CSS_SELECTOR, "ul"))
        self.assertIsNotNone(self.driver.find_element(By.CSS_SELECTOR, "dl"))
        drug = self.get_test_drug_info()
        self.assert_page_has(drug["id"])
        self.assert_page_has(drug["name"])

    def test_link_types(self):
        """Test the interface for managing link type controls."""

        # Before we begin the real test, clear out leftover test residue.
        landing_page = self.navigate_to("EditLinkControl.py")
        test_link_type_name = "Link Type For Automated Testing"
        self.assert_title("Manage Link Types")
        self.assert_page_has("Existing Link Types (click to edit)")
        self.assertIsNotNone(self.find("form ul li a"))

        def find_test_type_link():
            for link in self.find("form ul li a", all=True):
                if link.text == test_link_type_name:
                    return link
            return None
        test_type_link = find_test_type_link()
        if test_type_link is not None:
            test_type_link.click()
            self.select_new_tab()
            self.click("submit-button-delete")
            alert = self.wait.until(expected_conditions.alert_is_present())
            alert.accept()
            sleep(1)
            self.assert_title("Manage Link Types")
            self.assert_page_has(
                f"Successfully deleted link type '{test_link_type_name}'."
            )
            self.switch_to(landing_page)
            self.driver.refresh()

        # Create our test rule.
        self.find("input[value='Add New Link Type']").click()
        self.select_new_tab()
        self.assert_title("Add Link Type")
        self.set_field_value("name", test_link_type_name)
        self.set_field_value("comment", "something insightful and intelligent")
        self.select_values("version", "V")
        self.select_values("doctype-1", "Person")
        self.set_field_value("element-1", "LinkingElement")
        self.find("#block-1 img").click()
        self.find("#block-2 img").click()
        self.select_values("doctype-3", "Organization")
        self.set_field_value("element-3", "AnotherLinkingElement")
        self.click("target-documentation")
        self.click("target-documentationtoc")
        self.select_values("ruletype-2", "LinkTargetContains")
        self.set_field_value("ruletext-2", '/Foo == "some value"')
        self.set_field_value("rulecomment-2", "yada yada")
        self.select_values("ruletype-4", "LinkTargetContains")
        self.set_field_value("ruletext-4", '//Bar == "another value"')
        self.set_field_value("rulecomment-4", "molto yada")
        self.click("submit-button-save")

        # Verify that the link type was recorded correctly.
        self.assert_title("Edit Link Type")
        self.assert_page_has(f"Added new link type {test_link_type_name}.")
        # landing_page = self.navigate_to("EditLinkControl.py")
        self.switch_to(landing_page)
        self.driver.refresh()
        self.assert_title("Manage Link Types")
        self.assert_page_has(test_link_type_name)
        self.find("input[value='Show All Link Types']").click()
        self.select_new_tab()
        self.assert_title("Show All Link Types")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected_caption = "All Available Linking Element Combinations"
        self.assertEqual(table.caption.text, expected_caption)
        columns = (
            "Link Type",
            "Source Doctype",
            "Linking Element",
            "Target Doctype",
            "Pub/Ver/Cwd",
        )
        table.check_headers(columns)
        test_rows = []
        for row in table.rows:
            if row[0].text == "Link Type For Automated Testing":
                test_rows.append([cell.text for cell in row[1:]])
        self.assertEqual(len(test_rows), 4)
        sources = (
            ("Organization", "AnotherLinkingElement"),
            ("Person", "LinkingElement"),
        )
        targets = "Documentation", "DocumentationToC"
        expected = []
        for doctype, element in sources:
            for target in targets:
                expected.append([doctype, element, target, "V"])
        self.assertEqual(test_rows, expected)

        # Remove the test link type.
        self.switch_to(landing_page)
        self.driver.refresh()
        self.assert_title("Manage Link Types")
        self.assert_page_has(test_link_type_name)
        test_type_link = find_test_type_link()
        self.assertIsNotNone(test_type_link)
        test_type_link.click()
        self.select_new_tab()
        self.click("submit-button-delete")
        alert = self.wait.until(expected_conditions.alert_is_present())
        alert.accept()
        sleep(1)
        self.assert_title("Manage Link Types")
        self.assert_page_has(
            f"Successfully deleted link type '{test_link_type_name}'."
        )
        test_type_link = find_test_type_link()
        self.assertIsNone(test_type_link)

    def test_lock_management(self):
        """Test the utilities for clearing processing locks."""

        # Test the interface for removing the file sweeper lock.
        self.navigate_to("clear-filesweeper-lockfile.py")
        self.assert_title("Clear File Sweeper Lockfile")
        self.assert_page_has("Instructions")
        self.assert_page_has("The processing performed by this")
        self.submit_form(new_tab=False)
        page_source = self.get_page_source()
        not_found = "Lock file not found."
        success = "Lock file successfully removed."
        self.assert_title("Clear File Sweeper Lockfile")
        self.assertTrue(success in page_source or not_found in page_source)

        # Test the interface for removing the media lock. There's a limit
        # to what we can test without risk of disrupting an ongoing sync
        # of the media documents. We don't want that risk, even on a non-
        # production server. So we just verify the expected error message
        # for an operation we know will fail.
        self.navigate_to("UnlockMedia.py")
        self.assert_title("Unlock Media")
        self.assert_page_has("Instructions")
        self.assert_page_has("Media documents published from the CDR are")
        self.set_field_value("to", "/some/bogus/path")
        self.submit_form(new_tab=False)
        regex = r"(Path .* not found)|(Directory rename failed)"
        self.assert_regex(regex)

    def test_log_viewers(self):
        """Test the interfaces for reviewing logs."""

        # Check the server log viewer's UI.
        script = "log-tail.py"
        self.navigate_to(script)
        self.assert_title("Log Viewer")
        options = self.find("#p option", all=True)
        option_value = None
        for option in options:
            if option.text == "testing.log":
                option_value = option.get_attribute("value")
                break
        self.assertIsNotNone(option_value)
        self.select_values("p", option_value)
        self.set_field_value("c", "1")
        self.get_test_board()
        expected = "[INFO] started get-boards API service"
        self.submit_form()
        self.assert_page_has(expected)

        # Fetch the raw log.
        params = dict(
            Request="Submit",
            Session=self.session,
            p=option_value,
            r="yes"
        )
        url = f"{self.cgi}/{script}?{urlencode(params)}"
        log = self.fetch_from_url(url)
        self.assertIsNotNone(log)
        self.assertIn(expected, str(log, encoding="utf-8"))

        # Test the client log viewer.
        self.navigate_to("ShowClientLogs.py")
        self.assert_title("Client Log Viewer")
        self.set_field_value("date_range-start", "1/1/2023")
        self.set_field_value("date_range-end", "12/31/2023")
        self.submit_form()
        self.assert_title("Client Log Viewer")
        self.assert_plain_report()
        tables = self.load_tables()
        self.assertEqual(len(tables), 10)
        expected = ["Session", "Saved", "User"]
        for table in tables:
            self.assertEqual(len(table.rows), 3)
            headers = table.node.find_elements(By.TAG_NAME, "th")
            actual = [header.text for header in headers]
            self.assertEqual(actual, expected)

    def test_messaging(self):
        """Test interface for sending an email message to logged-in users."""

        subject = "Test message for automated CDR tests"
        body = "This is sent by an automated test. It can be ignored."
        self.navigate_to("MessageLoggedInUsers.py")
        self.assert_title("Send a Message to Logged-In CDR Users")
        self.assert_page_has("All fields are required")
        self.set_field_value("subject", subject)
        self.set_field_value("body", body)
        self.submit_form()
        self.assert_title("Send a Message to Logged-In CDR Users")
        self.assert_single_table_report()
        table = self.load_table()
        self.assert_tables_in_grid_container()
        table.check_headers(["Message Recipients"])

    def test_permissions(self):
        """Test the interface for managing accounts and permissions."""

        # We'll need these more than once, but only locally.
        def find_test_link(target):
            for link in self.find("#primary-form ul li a", all=True):
                if link.text == target:
                    return link

        def can_do(account, action, doctype):
            params = dict(
                account=account,
                action=action,
                doctype=doctype,
                Session=self.session,
            )
            url = f"{self.cgi}/check-auth.py?{urlencode(params)}"
            response = self.fetch_from_url(url).strip()
            return response == b"Y"

        test_data_types = "group", "action", "user"

        def remove_test_data(final):
            for type in test_data_types:
                script = f"Edit{type.capitalize()}s.py"
                self.navigate_to(script)
                name = f"Automated Test {type.capitalize()}"
                if type == "action":
                    name = name.upper()
                link = find_test_link(name)
                if final:
                    self.assertIsNotNone(link)
                button_id = f"submit-button-delete-{type}"
                expected = f"Successfully deleted {type} {name!r}."
                if type == "user":
                    button_id = "submit-button-inactivate-account"
                while link is not None:
                    if not final:
                        self.logger.warning("clearing out leftover %s", type)
                    link.click()
                    self.select_new_tab()
                    if type == "user":
                        name_field = self.find("name", method=By.ID)
                        self.assertIsNotNone(name_field)
                        pattern = "Successfully retired account for user {}."
                        user_machine_name = name_field.get_attribute("value")
                        expected = pattern.format(user_machine_name)
                    self.logger.debug("clicking button %r", button_id)
                    self.click(button_id)
                    self.assert_page_has(expected)
                    link = find_test_link(name)

        # First clear out any dross left by a previous, failed test.
        remove_test_data(final=False)

        # Add the test action.
        action_name = "AUTOMATED TEST ACTION"
        self.navigate_to("EditActions.py")
        self.assert_title("Manage Actions")
        self.assert_page_has("Existing Actions (click to edit)")
        link = find_test_link(action_name)
        self.assertIsNone(link)
        self.click("submit-button-add-new-action")
        self.select_new_tab()
        self.assert_title("Add New Action")
        self.set_field_value("name", action_name)
        self.set_field_value("comment", "This is a test action.")
        self.click("submit-button-save-new-action")
        self.assert_page_has(f"Action {action_name!r} successfully added.")
        self.assert_title(f"Edit {action_name} Action")
        self.click("options-doctype-specific")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Action {action_name!r} successfully updated.")
        self.assert_title(f"Edit {action_name} Action")

        # Create the test user.
        user_fullname = "Automated Test User"
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        user_name = f"testuser_{timestamp}"
        self.navigate_to("EditUsers.py")
        self.assert_title("Manage Users")
        self.assert_page_has("Existing Users (click to edit)")
        link = find_test_link(user_fullname)
        self.assertIsNone(link)
        self.click("submit-button-add-new-user")
        self.select_new_tab()
        self.assert_title("Adding New User Account")
        self.set_field_value("name", user_name)
        self.set_field_value("full_name", user_fullname)
        self.set_field_value("office", "My basement office")
        self.set_field_value("email", f"{user_name}@example.com")
        self.set_field_value("phone", "Butterfield 8")
        self.set_field_value("comment", "This is a test user.")
        self.click("submit-button-save-new-user-account")
        self.assert_page_has(f"New user {user_name} saved successfully.")
        self.assert_title(f"Editing User Account {user_fullname}")

        # Create the test group.
        group_name = "Automated Test Group"
        self.navigate_to("EditGroups.py")
        self.assert_title("Manage Groups")
        self.assert_page_has("Existing Groups (click to edit)")
        link = find_test_link(group_name)
        self.assertIsNone(link)
        self.click("submit-button-add-new-group")
        self.select_new_tab()
        self.assert_title("Add New Group")
        self.set_field_value("name", group_name)
        self.set_field_value("description", "This is a test group.")
        self.click("submit-button-save-new-group")
        self.assert_page_has(f"Group {group_name!r} successfully added.")
        self.assert_title(group_name)

        # At this point the account should be denied the action, because
        # the user doesn't have a membership in a group with the appropriate
        # permission. So let's add the user to our new test group.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.navigate_to("EditUsers.py")
        self.assert_title("Manage Users")
        self.assert_page_has("Existing Users (click to edit)")
        link = find_test_link(user_fullname)
        self.assertIsNotNone(link)
        link.click()
        self.select_new_tab()
        self.assert_title(f"Editing User Account {user_fullname}")
        self.click("group-automated-test-group")
        self.click("submit-button-save-changes")
        self.assert_title(f"Editing User Account {user_fullname}")
        expected = f"Changes to account {user_name} saved successfully."
        self.assert_page_has(expected)

        # That shouldn't be sufficient, because the group of which the
        # user just became a member doesn't have permission for the
        # new action. Grant such permission to the group, and while
        # we're on the page, verify that the page reflects the membership
        # of the new test user in the group.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.navigate_to("EditGroups.py")
        self.assertTrue("Manage Groups")
        self.assert_page_has("Existing Groups (click to edit)")
        link = find_test_link(group_name)
        self.assertIsNotNone(link)
        link.click()
        self.select_new_tab()
        self.assert_title(group_name)
        checkbox = self.find(f"user-{user_name}", method=By.ID)
        self.assertIsNotNone(checkbox)
        self.assertTrue(checkbox.is_selected())
        self.click("automated_test_action-doctype-term")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")

        # The account is a member of the group and now the group has
        # permission to perform the action. But the permission has been
        # granted for a different document type than the type for which
        # the account is requesting permission, so permission should
        # still be denied. Grant the permission for the Summary document
        # type, and the permission should finally be granted.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.click("automated_test_action-doctype-summary")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")
        self.assertTrue(can_do(user_name, action_name, "Summary"))

        # Remove the user from the group and confirm that the permission
        # is no longer granted.
        self.click(f"user-{user_name}")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")
        self.assertFalse(can_do(user_name, action_name, "Summary"))

        # Finally, clean up behind ourselves, removing the test data.
        remove_test_data(final=True)

    def test_query_term_definitions(self):
        """Test the interface for managing query-term definitions."""

        # Test the Compare button. Using STAGE for the upper tier,
        # because while I was developing this test CDR PROD was down
        # all day, and CBIIT couldn't figure out how to bring it back up.
        self.navigate_to("EditQueryTermDefs.py")
        self.assert_title("Manage Query Term Definitions")
        self.select_values("upper", "STAGE")
        self.click("submit-button-compare")
        self.assert_title("Manage Query Term Definitions")

        # Make sure nothing is left behind from a previous run of the test.
        test_path = "/Country/Chicken"
        self.click("submit-button-return-to-form")
        self.assert_title("Manage Query Term Definitions")
        if test_path in self.get_page_source():
            self.click("path-SLASHcountrySLASHchicken")
            self.click("submit-button-remove")
            self.assert_page_has("/Country/Chicken successfully removed.")

        def find_chickens():
            return self.run_query(
                "SELECT doc_id, value FROM query_term "
                f"WHERE path = '{test_path}'"
            )
        for doc_id, _ in find_chickens():
            self.delete_doc(doc_id, reason="Cleaning up query term dross")
        self.assertEqual(len(find_chickens()), 0)

        # Add the path definition.
        self.navigate_to("EditQueryTermDefs.py")
        self.assert_title("Manage Query Term Definitions")
        self.set_field_value("new_path", test_path)
        self.click("submit-button-add")
        self.assert_title("Manage Query Term Definitions")
        self.assert_page_has("/Country/Chicken successfully added.")

        # Verify that the new query term definition gets applied.
        root = etree.Element("Country")
        etree.SubElement(root, "Chicken").text = "Nuggets"
        xml = etree.tostring(root, encoding="unicode")
        doc_id = self.save_doc(xml, "Country")
        self.assertEqual(find_chickens(), [[str(doc_id), "Nuggets"]])

        # Drop the definition.
        self.navigate_to("EditQueryTermDefs.py")
        self.assert_title("Manage Query Term Definitions")
        self.click("path-SLASHcountrySLASHchicken")
        self.click("submit-button-remove")
        self.assert_page_has("/Country/Chicken successfully removed.")
        self.save_doc(xml, "Country", id=f"CDR{doc_id:010d}", unlock=True)
        self.assertEqual(len(find_chickens()), 0)
        self.delete_doc(doc_id)
        self.assertEqual(len(find_chickens()), 0)

    def test_scheduler(self):
        """Test the interface for managing scheduled jobs."""

        # Create local variables we'll use throughout.
        job_name = "Automated Testing Job"
        job_class = "test.Stub"

        # Create a local function we'll use more than once.
        def find_test_job_link(table):
            for link in table.node.find_elements(By.TAG_NAME, "a"):
                if link.text == job_name:
                    return link
            return None

        # Start by clearing out any old jobs.
        for i in range(2):
            self.navigate_to("Scheduler.py")
            self.assert_title("Scheduled Jobs")
            tables = self.load_tables()
            self.assertEqual(len(tables), 2)
            link = find_test_job_link(tables[i])
            while link is not None:
                link.click()
                self.click("submit-button-delete-job")
                alert = self.wait.until(expected_conditions.alert_is_present())
                self.logger.debug("alert text: %s", alert.text)
                alert.accept()
                sleep(1)
                self.assert_title("Scheduled Jobs")
                self.assert_page_has(f"Job {job_name} successfully deleted.")
                tables = self.load_tables()
                link = find_test_job_link(tables[i])
                self.assertEqual(len(tables), 2)

        # Create the test job.
        self.click("submit-button-add-new-job")
        self.assert_title("Scheduled Jobs")
        self.click("submit-button-save")
        self.assert_page_has("Job name is required.")
        self.assert_page_has("Class name for job is required.")
        self.set_field_value("name", job_name)
        self.set_field_value("job_class", job_class)
        self.set_field_value("hour", "3")
        self.set_field_value("minute", "59")
        self.set_field_value("opt-name-1", "stooge")
        self.set_field_value("opt-value-1", "Larry")
        self.find("#primary-form img").click()
        self.set_field_value("opt-name-2", "pep-boy")
        self.set_field_value("opt-value-2", "Manny")
        self.click("submit-button-save")
        self.assert_title("Scheduled Jobs")
        self.assert_page_has(
            f"Enabled job {job_name!r} saved. "
            "Job will run every day at 03:59."
        )

        # Modify the job and save it again.
        self.click("opts-enabled")
        self.click("submit-button-save")
        self.assert_title("Scheduled Jobs")
        self.assert_page_has(f"Disabled job {job_name!r} saved.")

        # Verify that it shows up in the correct table on the Jobs page.
        self.click("submit-button-jobs")
        self.assert_title("Scheduled Jobs")
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        link = find_test_job_link(tables[0])
        self.assertIsNone(link)
        link = find_test_job_link(tables[1])
        self.assertIsNotNone(link)

        # Run the test job manually.
        link.click()
        self.assert_title("Scheduled Jobs")
        self.click("submit-button-run-job-now")
        alert = self.wait.until(expected_conditions.alert_is_present())
        self.logger.debug("alert text: %s", alert.text)
        alert.accept()
        sleep(1)
        self.assert_title("Scheduled Jobs")
        self.assert_page_has(f"Job {job_name!r} queued.")

        # Verify that the JSON report reflects the job's presence and values.
        # Make sure that it only shows up once.
        params = dict(
            Request="JSON",
            Session=self.session,
        )
        json_url = f"{self.cgi}/Scheduler.py?{urlencode(params)}"
        job = None
        for values in load_from_json(self.fetch_from_url(json_url)):
            if values[0] == job_name:
                self.assertIsNone(job)
                options = load_from_json(values[3]) if values[3] else {}
                schedule = load_from_json(values[4]) if values[4] else {}
                job = dict(
                    name=values[0],
                    enabled=values[1],
                    job_class=values[2],
                    options=options,
                    schedule=schedule,
                )
        expected_options = {"pep-boy": "Manny", "stooge": "Larry"}
        expected_schedule = {"hour": "3", "minute": "59"}
        self.assertIsNotNone(job)
        self.assertFalse(job["enabled"])
        self.assertEqual(job["job_class"], job_class)
        self.assertEqual(job["options"], expected_options)
        self.assertEqual(job["schedule"], expected_schedule)

        # Remove the test job and verify that it is gone.
        self.click("submit-button-delete-job")
        alert = self.wait.until(expected_conditions.alert_is_present())
        self.logger.debug("alert text: %s", alert.text)
        alert.accept()
        sleep(1)
        self.assert_title("Scheduled Jobs")
        self.assert_page_has(f"Job {job_name} successfully deleted.")
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        link = find_test_job_link(tables[0])
        self.assertIsNone(link)
        link = find_test_job_link(tables[1])
        self.assertIsNone(link)

        # An enabled job with no schedule should be run but not retained.
        self.click("submit-button-add-new-job")
        self.assert_title("Scheduled Jobs")
        self.set_field_value("name", job_name)
        self.set_field_value("job_class", job_class)
        self.set_field_value("opt-name-1", "stooge")
        self.set_field_value("opt-value-1", "Curly")
        self.find("#primary-form img").click()
        self.set_field_value("opt-name-2", "pep-boy")
        self.set_field_value("opt-value-2", "Jack")
        self.click("submit-button-save")
        self.assert_title("Scheduled Jobs")
        expected = f"Enabled job {job_name!r} will be run but not retained"
        self.assert_page_has(expected)
        self.click("submit-button-jobs")
        self.assert_title("Scheduled Jobs")
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        link = find_test_job_link(tables[0])
        self.assertIsNone(link)
        link = find_test_job_link(tables[1])
        self.assertIsNone(link)

        # Verify that there is no trace of the test remaining.
        for values in load_from_json(self.fetch_from_url(json_url)):
            self.assertNotEqual(values[0], job_name)

    def test_schemas(self):
        """Test the interface for managing schemas on the CDR server."""

        # We'll need this a couple of times.
        def fetch_schema(name):
            for link in self.find("form ul li a", all=True):
                if link.text == name:
                    url = link.get_attribute("href")
                    self.logger.debug("fetching from %s", url)
                    return self.fetch_from_url(url)
            return None

        # Get a copy of the test schema and save it to disk.
        self.navigate_to("GetSchema.py")
        self.assert_title("Show Schema")
        schema = fetch_schema("xxtest").strip().replace(b"\r", b"")
        self.assertIsNotNone(schema)
        path = Path("xxtest.xml").resolve()
        path.write_bytes(schema)

        # Post it to the server.
        self.navigate_to("post-schema.py")
        self.assert_title("Post CDR Schema")
        self.set_field_value("file", str(path))
        self.set_field_value("comment", "posted by automated test")
        self.submit_form(new_tab=False)
        self.assert_title("Post CDR Schema")
        self.assert_page_has("Schema posted successfully.")

        # Verify that it made the round trip intact.
        self.navigate_to("GetSchema.py")
        self.assert_title("Show Schema")
        fetched = fetch_schema("xxtest").strip().replace(b"\r", b"")
        self.assertEqual(fetched, schema)

        # Clean up after ourselves.
        path.unlink()

    def test_tier_settings(self):
        """Test the service to fetch the server settings on the tested tier."""

        script = "fetch-tier-settings.py"
        self.navigate_to(script, prompt="yes")
        self.assert_title("Tier Settings")
        self.assert_page_has("Instructions")
        self.assert_page_has("Click Submit to generate a JSON representation")
        url = f"{self.cgi}/{script}?Session={self.session}"
        self.logger.debug("fetching tier settings from %s", url)
        json = self.fetch_from_url(url)
        self.assertIsNotNone(json)
        values = load_from_json(json)
        self.assertIn("windows", values)
        self.assertIn("version", values["windows"])
        self.assertIn("major", values["windows"]["version"])

    def test_value_tables(self):
        """Test the interface for managing status valid value tables."""

        # We'll use this a bunch in this test, but nowhere else.
        def collect_values():
            class ValueLink:
                PATTERN = r"(.+) \(position (\d+)\)"

                def __init__(self, tester, node):
                    self.element = node
                    re_match = re_search(self.PATTERN, node.text)
                    tester.assertIsNotNone(re_match)
                    self.value = re_match.group(1)
                    self.position = int(re_match.group(2))
            values = {}
            for element in self.find("fieldset ul li a", all=True):
                value_link = ValueLink(self, element)
                values[value_link.value] = value_link
            return values

        # Verify the tables on the starting page.
        script = "edit-value-table.py"
        self.navigate_to(script)
        self.assert_title("Edit Value Tables")
        self.assert_page_has("Select Table")
        tables = [
            "glossary_translation_state",
            "media_translation_state",
            "summary_translation_state",
            "summary_change_type",
        ]
        selector = "fieldset input[type='radio']"
        for i, button in enumerate(self.find(selector, all=True)):
            id = button.get_attribute("id")
            self.assertEqual(id, f"table-{tables[i]}")

        # Verify the required table selection.
        self.submit_form(new_tab=False)
        self.assert_page_has("Please select a table to edit.")

        # Test editing each of the tables.
        test_value = "Regression Test Valid Value"
        for table_name in tables:
            self.navigate_to(script)
            self.click(f"table-{table_name}")
            self.submit_form(new_tab=False)
            self.assert_title("Edit Value Tables")
            self.assert_page_has("Values (click link to edit a value)")

            # Clear out any old test data.
            values = collect_values()
            if test_value in values:
                values[test_value].element.click()
                self.click("submit-button-drop")
                values = collect_values()
                self.assertNotIn(test_value, values)
            self.assert_page_not_has(test_value)

            # Test error conditions on the Add form.
            self.click("submit-button-add")
            self.assert_title("Edit Value Tables")
            self.click("submit-button-save")
            self.assert_page_has("Value name is required.")
            self.assert_page_has("Position field is required.")
            last_value = values[list(values)[-1]].value
            last_pos = values[list(values)[-1]].position
            self.set_field_value("value", last_value)
            self.set_field_value("position", last_pos)
            self.click("submit-button-save")
            self.assert_page_has(f"The name '{last_value}' is already in use.")
            self.assert_page_has(f"Position {last_pos} is already in use.")

            # Add the new value and confirm that it's in the list.
            new_pos = last_pos + 10
            self.set_field_value("value", test_value)
            self.set_field_value("position", new_pos)
            self.click("submit-button-save")
            self.assert_page_has(f"Value '{test_value}' saved.")
            old_count = len(values)
            values = collect_values()
            self.assertEqual(old_count, len(values)-1)
            self.assertIn(test_value, values)
            value_link = values[test_value]
            self.assertEqual(value_link.position, new_pos)
            self.assertEqual(value_link.value, test_value)

            # Drop the test value and confirm that it's gone.
            value_link.element.click()
            self.click("submit-button-drop")
            self.assert_page_has(f"Value {test_value!r} successfully dropped.")
            values = collect_values()
            self.assertNotIn(test_value, values)
            self.assertEqual(len(values), old_count)


class DrugTests(Tester):
    """Tests of report and utilities for DrugInformationSummary documents.

    Note that some of the tests assume certain conventions in the patterns for
    comments associated with saved versions of the documents. Those patterns
    are used consistently, and since the test docuemtns chosen are the ones
    created first, we can be confident that we will always find a matching
    version in the document we're testing, even if the currently used
    conventions are later changed.
    """

    def test_drug_advanced_search(self):
        "Test advanced search for DrugInformationSummary documents."

        self.navigate_to("DISSearch.py")
        self.assert_page_has("Drug Information Summary")
        labels = (
            "Title",
            "FDA Approved",
            "Last Modified",
            "Approved Indication",
            "Drug Reference Type",
        )
        for label in labels:
            self.assert_page_has(label)
        drug = self.get_test_drug_info()
        drug_id = drug["id"]
        drug_name = drug["name"]
        self.set_field_value("title", f"{drug_name[0]}%")
        selector = 'main form input[type="submit"]'
        self.driver.find_element(By.CSS_SELECTOR, selector).click()
        self.select_new_tab()
        self.assert_page_has(f"documents match '{drug_name[0]}%'")
        qc_report_link = None
        cdr_id = f"CDR{drug_id:010d}"
        for link in self.driver.find_elements(By.CSS_SELECTOR, "td a"):
            if link.text == cdr_id:
                qc_report_link = link
                break
        self.assertIsNotNone(qc_report_link)
        qc_report_link.click()
        expected = f"(?s)Drug Information Summary.+QC Report.+{date.today()}"
        self.assert_regex(expected)
        self.assert_page_has(f"CDR{drug_id}")
        self.assert_page_has(drug_name)

    def test_drug_comprehensive_review_dates_report(self):
        """Test the Drugs Comprehensive review Dates report.

        Note that there are two flavors of the report, one of which shows the
        dates for all of each drug's comprehensive reviews, and the other of
        which shows only the date of the most recent comprehensive review for
        each drug. There's no difference between the two versions of the
        report, since no DrugInformationSummary documents have more than one
        comprehensive review.
        """

        self.navigate_to("DrugCRD.py")
        title = "Drugs Comprehensive Review Dates"
        self.assert_title(title)
        self.submit_form()
        self.assert_title(f"{title} ({date.today()})")
        self.assert_page_has("Single Agent Drugs")
        self.assert_page_has("Combination Drugs")
        self.assert_regex("(?s)Doc Title.+Date")

    def test_drug_date_last_modified_report(self):
        """Test the Drug Date Last Modified report."""

        # There are two versions of the report, one of which ("user") selects
        # the documents based on when a user determined the last subtantive
        # changes were made to the documents. The other version ("system")
        # selects documents based on when they were last saved, including
        # saves for which the changes were insignificant (even if there were
        # no changes at all). The default is the first version ("user").
        form_page = self.navigate_to("DrugDateLastModified.py")
        self.assert_page_has("DIS Date Last Modified")
        self.set_field_value("start", "01/01/2020")
        self.set_field_value("end", "12/31/2020")
        self.submit_form()
        expected = "Drug Information Summary Date Last Modified (User) Report"
        self.assert_page_has(expected)
        self.assert_page_has("2020-01-01 — 2020-12-31")
        self.assert_page_has(f"Report date: {date.today()}")
        columns = (
            "DocId",
            "Summary Title",
            "Date Last Modified",
            "Last Modify Action Date (System)",
            "LastV Publishable?",
            "User"
        )
        escaped = [re_escape(column) for column in columns]
        self.assert_regex("(?s)" + ".+".join(escaped))
        self.switch_to(form_page)
        self.click("report-type-system")
        self.submit_form()
        self.assert_page_has(expected.replace("User", "System"))

    def test_drug_description_report(self):
        """Test the Drug Description report."""

        # There are four filtering methods for this report. The first method
        # is referred to as "By Drug Name," and uses a picklist from which the
        # user can select one or more drugs to be shown on the report. The
        # picklist also has entries for "All Drugs," "All Single-Agent Drugs,"
        # and "All Drug Combinations."
        form = self.navigate_to("DrugDescriptionReport.py")
        self.assert_title("Drug Description Report")
        drugs = self.get_test_data("drug-info")
        select = Select(self.driver.find_element(By.ID, "drugs"))
        for drug in drugs:
            select.select_by_value(str(drug["id"]))
        self.submit_form()
        self.assert_title("Drug Description Report")
        self.assert_page_has(str(date.today()))
        self.assert_page_has(f"{len(drugs)} documents found by name")
        for drug in drugs:
            self.assert_page_has(drug["name"])
            self.assert_page_has(str(drug["id"]))

        # The second method supports filtering by date of the last publishable
        # versions of the drug documents.
        self.switch_to(form)
        self.click("method-date")
        self.set_field_value("start", "01/01/2020")
        self.set_field_value("end", "12/31/2020")
        self.submit_form()
        expected = (
            r"\d+ documents found with last publishable versions created " +
            re_escape("2020-01-01--2020-12-31 (inclusive)")
        )
        self.assert_regex(expected)

        # Method #3 filters by drug reference type (NCI, FDA, or NLM).
        self.switch_to(form)
        self.click("method-type")
        self.click("reftype-nci")
        self.submit_form()
        self.assert_regex(r"\d+ documents found with reference type NCI")

        # Finally, the report can be filtered by FDA approval information.
        self.switch_to(form)
        self.click("method-fda")
        self.click("reftype-nci")
        self.submit_form()
        expected = (
            r"\d+ documents found with "
            "Accelerated approval and Approved in children"
        )
        self.assert_regex(expected)

    def test_drug_indications_report(self):
        """Test the Drug Indications report."""

        # There are three flavors for this report. The simplest is
        # "Indications Only," which produces a single table with a
        # single column, just listing the conditions for which any
        # of the drugs are indicated.
        form = self.navigate_to("DrugIndicationsReport.py")
        self.assert_title("Drug Indications")
        self.assert_page_has("Select Data To Be Displayed")
        self.click("type-plain")
        self.submit_form()
        self.assert_tables_in_grid_container()
        self.assert_single_table_report()
        cols = self.driver.find_elements(By.CSS_SELECTOR, "main colgroup col")
        self.assertEqual(len(cols), 1)
        self.assert_page_has("Full List of Drug Indications")

        # The default version of the report produces a description list (DL)
        # structure (instead of a table), with each condition presented as a
        # description term (DT) element, and the drugs indicated for that
        # condition following in description detail (DD) elements. The user
        # can optionally narrow the report by specifying one or more conditions
        # ("indications") to be represented in the report. For each drug
        # indicated for one ore more of the selected conditions, all of the
        # conditions for which that drug is indicated are picked up by the
        # report, accompanied by the drugs indicated for those conditions
        # which are selected for the report. In other words, there will
        # likely be more DT elements than there are conditions selected for
        # the report.
        self.switch_to(form)
        self.click("type-drug")
        select = Select(self.driver.find_element(By.ID, "indication"))
        select.select_by_value("bone cancer")
        self.submit_form()
        expected = "Approved Indications for Drug Information Summaries"
        self.assert_page_has(expected)
        self.assert_page_has("<dt>bone cancer</dt>")
        conditions = self.driver.find_elements(By.CSS_SELECTOR, "dl dt")
        self.assertGreater(len(conditions), 1)
        self.assert_non_tabular_report()

        # The drug name in each DD element is accompanied by the drug's
        # CDR ID in parentheses, marked up as a link to the drug's QC report.
        self.driver.find_element(By.CSS_SELECTOR, "dd a").click()
        self.assert_page_has("QC Report")

        # There is a variation on this second flavor of the report. Instead
        # of grouping drugs under indications (the default described above),
        # this variation groups the indications by drug. It does this in a
        # table instead of a description list, with three columns: one for
        # the CDR ID for the drug information document, one for the drug
        # name, and one for all the conditions for which the drug is approved.
        # The CDR ID in the first column is marked up as a link to the drug's
        # QC report.
        self.switch_to(form)
        self.click("grouping-drug")
        self.submit_form()
        self.assert_single_table_report()
        table = self.load_table()
        table.check_headers(("CDR ID", "Drug Name", "Approved Indication(s)"))
        table.rows[0][0].find_element(By.TAG_NAME, "a").click()
        self.assert_page_has("QC Report")

        # The third flavor for the report includes the drugs' brand names.
        self.switch_to(form)
        self.click("type-brand")
        self.click("grouping-indication")
        self.submit_form()
        self.assert_single_table_report()
        table = self.load_table()
        headers = "Approved Indication", "Drug Name", "Brand Name(s)"
        table.check_headers(headers)

        # As with the second flavor, the grouping can be flipped so that
        # conditions are grouped under drugs.
        self.switch_to(form)
        self.click("grouping-drug")
        self.submit_form()
        self.assert_single_table_report()
        table = self.load_table()
        headers = "CDR ID", "Drug Name (Brand Name)", "Approved Indication(s)"
        table.check_headers(headers)

    def test_drug_lists_report(self):
        """Test the Drug Info Summaries List report."""

        self.navigate_to("DISLists.py")
        self.assert_page_has("Drug Info Summaries List")
        self.submit_form()
        self.assert_page_has(f"Drug Info Summaries List -- {date.today()}")
        self.assert_regex(r"Single Agent Drug [(]\d+[)]")
        self.assert_regex(r"Combination Drug [(]\d+[)]")
        self.assert_regex("(?s)CDR ID.+Title.+Drug Type")

    def test_drug_markup_report(self):
        """Test Drug Summary with Markup report."""

        self.navigate_to("DISWithMarkup.py")
        self.assert_page_has("Drug Summaries with Markup")
        self.assert_page_has("Type of mark-up to Include")
        types = "Publish", "Approved", "Proposed", "Rejected", "Other"
        for type in types:
            checkbox = self.driver.find_element(By.ID, f"type-{type.lower()}")
            self.assertTrue(checkbox.is_selected())
        self.submit_form()
        expected = f"Count of Revision Level Markup - {date.today()}"
        self.assert_page_has(expected)
        for type in types:
            self.assert_page_has(f"{type}</th>")

    def test_drug_processing_status_report(self):
        """Test the DIS Processing Status Report."""

        self.navigate_to("DISProcessingStatusReport.py")
        self.assert_page_has("DIS Processing Status Report")
        self.assert_page_has("Select Status Value(s)")
        self.assert_page_has("Date Range")
        self.click("status-dis-ok-to-publish")
        self.click("status-dis-to-expert-reviewer")
        self.set_field_value("start", "1/1/2022")
        self.set_field_value("end", "12/31/2022")
        self.submit_form()
        self.assert_page_has("DIS OK to publish")
        self.assert_page_has("DIS to expert reviewer")
        self.assert_page_not_has("Needs DailyMed link")
        self.assert_page_not_has("Needs glossary draft review")
        columns = (
            "CDR ID",
            "DIS Title",
            "Processing Status Value",
            "Processing Status Date",
            "Entered By",
            "Comments",
            "Last Version Publishable?",
            "Date First Published",
            "Published Date",
        )
        for column in columns:
            self.assert_page_has(column)

    def test_drug_publish_preview(self):
        """Test publish preview of drug information summaries.

        Note that you will need to make sure the CDR web server is configured
        to connect to a Drupal server which is functioning correctly, which
        is not always guaranteed, as the lower-tier servers are frequently
        being rebuilt and tested.

        This test takes longer than most (roughly a minute), as the lower-tier
        Drupal servers are pretty sluggish.
        """

        params = dict(DocType="DrugInformationSummary", ReportType="pp")
        self.navigate_to("QcReport.py", **params)
        self.assert_title("Publish Preview Report")
        drug = self.get_test_drug_info()
        drug_id = drug["id"]
        drug_name = drug["name"]
        self.set_field_value("DocId", drug_id)
        self.submit_form()
        self.assert_page_has(f"Select Document Version For CDR{drug_id}")
        self.assertTrue(self.select_version("PP version"))
        self.submit_form()
        self.assertEqual(self.driver.title, f"{drug_name} - NCI")
        self.assert_title(drug_name)
        self.assert_page_has("A to Z List of Cancer Drugs")

    def test_drug_qc_report(self):
        """Test the DrugInformationSummary QC report."""

        # Error in user input should be reported and the form redisplayed.
        self.navigate_to("QcReport.py", DocType="DrugInformationSummary")
        self.assert_title("Drug Information Summary QC Report")
        self.set_field_value("DocId", "999999")
        self.submit_form()
        self.assert_page_has("Document 999999 not found.")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_page_has(f"CDR{summary_id} is a Summary document.")

        # A version selection form should be displayed after a document
        # has been selected.
        drug = self.get_test_drug_info()
        drug_id = drug["id"]
        drug_name = drug["name"]
        self.set_field_value("DocId", drug_id)
        self.submit_form()
        self.assert_page_has(f"Select Document Version For CDR{drug_id}")

        # If we select a version with redline/strikeout revision markup, we
        # should get the QC report for the drug with the markup highlighted.
        # Also, verify that the QC report does not use the USWDS framework.
        self.assertTrue(self.select_version("RL"))
        self.click("options-markup-publish")
        self.click("options-markup-proposed")
        self.submit_form()
        page = self.get_page_source()
        self.assertIn("Drug Information Summary", page)
        self.assertIn("QC Report", page)
        self.assertIn(str(date.today()), page)
        self.assertIn(drug_name, page)
        self.assertIn(f"CDR{drug_id}", page)
        insert_span = '<span class="insert'
        delete_span = '<span class="delete'
        self.assertTrue(insert_span in page or delete_span in page)
        self.assert_plain_report()

    def test_drug_type_of_change_report(self):
        """Test the DIS Type of Change report."""

        # The default version of the report has a pair of columns for
        # each type of change selected for display. By default, all
        # types are selected. The first column in each pair shows the
        # date of the change, with the second column for any associated
        # comment. There are two columns at the left of the report table
        # for the document ID and title. Verify that the report does not
        # use the USWDS framework.
        form = self.navigate_to("DISTypeChangeReport.py")
        self.assert_title("DIS Type of Change")
        self.assert_page_has("Instructions")
        self.assert_page_has("To run this report ")
        self.click("selection_method-id")
        drugs = self.get_test_data("drug-info")
        drug_ids = [drug["id"] for drug in drugs]
        self.set_field_value("cdr-id", " ".join([str(id) for id in drug_ids]))
        types = (
            "New Doc - New Summary",
            "Revised Doc - Editorial Change",
            "Revised Doc - External Links",
            "Revised Doc - New or Updated FDA Approval",
            "Revised Doc - Other Change",
            "Revised Doc - Special Project",
        )
        for type in types:
            id = "change-type-" + type.lower().replace(" ", "-")
            checkbox = self.driver.find_element(By.ID, id)
            self.assertTrue(checkbox.is_selected())
        self.submit_form()
        self.assert_plain_report()
        self.assert_single_table_report()
        self.assert_page_has("Type of Change Report (Most Recent Change)")
        columns = ["CDR-ID", "Title"]
        for type in types:
            columns += [type, "Comment"]
        for column in columns:
            self.assert_page_has(f"{column}</th>")

        # The other version of the report ("Historical") shows all the changes
        # for the specified date range (which by default covers all activity
        # since the release of the CDR in June 2002). By default this version
        # of the report has only five columns, with one row for each of a
        # document's changes. The first two columns (ID and title) span the
        # rows for each document.
        self.switch_to(form)
        self.click("type-historical-(all-changes-for-a-given-date-range)")
        self.submit_form()
        self.assert_regex("(?s)CDR-ID.+Title.+Date.+Type of Change.+Comment")
        self.assert_page_has(
            "Type of Change Report (All Changes by Drug Info Summary)"
        )
        self.assert_page_has("Complete History")
        self.assert_regex(r'<td rowspan="\d+">')
        self.assert_plain_report()
        self.assert_single_table_report()

        # In this version of the report the document id is marked up as a link
        # to the document's QC report (but not in the other version of the
        # report, for some reason).
        self.driver.find_element(By.CSS_SELECTOR, "td a").click()
        expected = f"(?s)Drug Information Summary.+QC Report.+{date.today()}"
        self.assert_regex(expected)

        # There is an option for the historical version of the report to show
        # separate tables for each type of change found in the selected docs.
        # The option is displayed even when the "Current" version of the report
        # is selected, but it's ignored.
        self.switch_to(form)
        self.click("organization-one-table-for-each-type-of-change")
        self.submit_form()
        self.assert_plain_report()
        self.assert_multi_table_report()
        for header in "CDR-ID", "Title", "Date", "Comment":
            self.assert_page_has(f"{header}</th>")
        self.assert_page_not_has("Type of Change</th>")

    def test_drug_type_report(self):
        """Test Drug By Drug Type report."""

        form = self.navigate_to("DISByDrugType.py")
        self.assert_title("Drug Report By Drug Type")
        self.click("format-html")
        self.submit_form()
        self.assert_page_has("Drug Information Summary by Drug Type Report")
        self.assert_page_has("Drug Type(s): All")
        for column in "CDR ID", "Title of DIS", "Drug Types", "Publishable?":
            self.assert_page_has(f"{column}</th>")

        # The report supports narrowing by drug types.
        self.switch_to(form)
        self.click("drug-type-hormone-therapy")
        self.click("drug-type-immunotherapy")
        self.submit_form()
        self.assert_page_not_has("Drug Type(s): All")
        self.assert_page_has("Drug Type(s): Hormone therapy, Immunotherapy")

    def test_permissions(self):
        """Test the interface for managing accounts and permissions."""

        # We'll need these more than once, but only locally.
        def find_test_link(target):
            for link in self.find("#primary-form ul li a", all=True):
                if link.text == target:
                    return link

        def can_do(account, action, doctype):
            params = dict(
                account=account,
                action=action,
                doctype=doctype,
                Session=self.session,
            )
            url = f"{self.cgi}/check-auth.py?{urlencode(params)}"
            response = self.fetch_from_url(url).strip()
            return response == b"Y"

        test_data_types = "group", "action", "user"

        def remove_test_data(final):
            for type in test_data_types:
                script = f"Edit{type.capitalize()}s.py"
                self.navigate_to(script)
                name = f"Automated Test {type.capitalize()}"
                if type == "action":
                    name = name.upper()
                link = find_test_link(name)
                if final:
                    self.assertIsNotNone(link)
                button_id = f"submit-button-delete-{type}"
                expected = f"Successfully deleted {type} {name!r}."
                if type == "user":
                    button_id = "submit-button-inactivate-account"
                while link is not None:
                    if not final:
                        self.logger.warning("clearing out leftover %s", type)
                    link.click()
                    self.select_new_tab()
                    if type == "user":
                        name_field = self.find("name", method=By.ID)
                        self.assertIsNotNone(name_field)
                        pattern = "Successfully retired account for user {}."
                        user_machine_name = name_field.get_attribute("value")
                        expected = pattern.format(user_machine_name)
                    self.logger.debug("clicking button %r", button_id)
                    self.click(button_id)
                    self.assert_page_has(expected)
                    link = find_test_link(name)

        # First clear out any dross left by a previous, failed test.
        remove_test_data(final=False)

        # Add the test action.
        action_name = "AUTOMATED TEST ACTION"
        self.navigate_to("EditActions.py")
        self.assert_title("Manage Actions")
        self.assert_page_has("Existing Actions (click to edit)")
        link = find_test_link(action_name)
        self.assertIsNone(link)
        self.click("submit-button-add-new-action")
        self.select_new_tab()
        self.assert_title("Add New Action")
        self.set_field_value("name", action_name)
        self.set_field_value("comment", "This is a test action.")
        self.click("submit-button-save-new-action")
        self.assert_page_has(f"Action {action_name!r} successfully added.")
        self.assert_title(f"Edit {action_name} Action")
        self.click("options-doctype-specific")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Action {action_name!r} successfully updated.")
        self.assert_title(f"Edit {action_name} Action")

        # Create the test user.
        user_fullname = "Automated Test User"
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        user_name = f"testuser_{timestamp}"
        self.navigate_to("EditUsers.py")
        self.assert_title("Manage Users")
        self.assert_page_has("Existing Users (click to edit)")
        link = find_test_link(user_fullname)
        self.assertIsNone(link)
        self.click("submit-button-add-new-user")
        self.select_new_tab()
        self.assert_title("Adding New User Account")
        self.set_field_value("name", user_name)
        self.set_field_value("full_name", user_fullname)
        self.set_field_value("office", "My basement office")
        self.set_field_value("email", f"{user_name}@example.com")
        self.set_field_value("phone", "Butterfield 8")
        self.set_field_value("comment", "This is a test user.")
        self.click("submit-button-save-new-user-account")
        self.assert_page_has(f"New user {user_name} saved successfully.")
        self.assert_title(f"Editing User Account {user_fullname}")

        # Create the test group.
        group_name = "Automated Test Group"
        self.navigate_to("EditGroups.py")
        self.assert_title("Manage Groups")
        self.assert_page_has("Existing Groups (click to edit)")
        link = find_test_link(group_name)
        self.assertIsNone(link)
        self.click("submit-button-add-new-group")
        self.select_new_tab()
        self.assert_title("Add New Group")
        self.set_field_value("name", group_name)
        self.set_field_value("description", "This is a test group.")
        self.click("submit-button-save-new-group")
        self.assert_page_has(f"Group {group_name!r} successfully added.")
        self.assert_title(group_name)

        # At this point the account should be denied the action, because
        # the user doesn't have a membership in a group with the appropriate
        # permission. So let's add the user to our new test group.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.navigate_to("EditUsers.py")
        self.assert_title("Manage Users")
        self.assert_page_has("Existing Users (click to edit)")
        link = find_test_link(user_fullname)
        self.assertIsNotNone(link)
        link.click()
        self.select_new_tab()
        self.assert_title(f"Editing User Account {user_fullname}")
        self.click("group-automated-test-group")
        self.click("submit-button-save-changes")
        self.assert_title(f"Editing User Account {user_fullname}")
        expected = f"Changes to account {user_name} saved successfully."
        self.assert_page_has(expected)

        # That shouldn't be sufficient, because the group of which the
        # user just became a member doesn't have permission for the
        # new action. Grant such permission to the group, and while
        # we're on the page, verify that the page reflects the membership
        # of the new test user in the group.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.navigate_to("EditGroups.py")
        self.assertTrue("Manage Groups")
        self.assert_page_has("Existing Groups (click to edit)")
        link = find_test_link(group_name)
        self.assertIsNotNone(link)
        link.click()
        self.select_new_tab()
        self.assert_title(group_name)
        checkbox = self.find(f"user-{user_name}", method=By.ID)
        self.assertIsNotNone(checkbox)
        self.assertTrue(checkbox.is_selected())
        self.click("automated_test_action-doctype-term")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")

        # The account is a member of the group and now the group has
        # permission to perform the action. But the permission has been
        # granted for a different document type than the type for which
        # the account is requesting permission, so permission should
        # still be denied. Grant the permission for the Summary document
        # type, and the permission should finally be granted.
        self.assertFalse(can_do(user_name, action_name, "Summary"))
        self.click("automated_test_action-doctype-summary")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")
        self.assertTrue(can_do(user_name, action_name, "Summary"))

        # Remove the user from the group and confirm that the permission
        # is no longer granted.
        self.click(f"user-{user_name}")
        self.click("submit-button-save-changes")
        self.assert_page_has(f"Group {group_name!r} successfully updated.")
        self.assertFalse(can_do(user_name, action_name, "Summary"))

        # Finally, clean up behind ourselves, removing the test data.
        remove_test_data(final=True)


class GeneralTests(Tester):
    """Tests not belonging to a specific category."""

    def test_activity_report(self):
        """Test the Document Activity report."""

        self.navigate_to("ActivityReport.py")
        self.assert_title("Document Activity Report")
        self.select_values("doctype", "DrugInformationSummary")
        self.set_field_value("start_date", "1/1/2020")
        self.set_field_value("end_date", "1/31/2020")
        self.submit_form()
        self.assert_title("Document Activity Report")
        caption = f"DrugInformationSummary Documents -- {date.today()}"
        self.assert_table_caption(caption)
        self.assert_table_caption("From 2020-01-01 to 2020-01-31")
        self.assert_single_table_report()
        self.assert_wide_report()
        columns = (
            "Who",
            "When",
            "Action",
            "DocType",
            "DocID",
            "DocTitle",
            "Comment",
        )
        table = self.load_table()
        table.check_headers(columns)
        self.driver.find_element(By.CSS_SELECTOR, "tr a").click()
        self.assert_page_has("Drug Information Summary")
        self.assert_page_has("QC Report")

    def test_country_qc_report(self):
        """Test the Country QC reort."""

        self.navigate_to("CountrySearch.py")
        self.assert_title("Country")
        self.set_field_value("name", "canada")
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Country")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        self.assertEqual(table.caption.text, "1 document matches 'canada'")
        self.assertEqual(len(table.rows), 1)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][2].text, "Canada")
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_page_has("<center>Country<br>QC Report</center>")
        self.assert_page_has("Canada")

    def test_current_sessions_report(self):
        """Test the Current Sessions report."""

        self.navigate_to("ActiveLogins.py", prompt="yes")
        self.assert_title("Current Sessions")
        self.assert_page_has("Instructions")
        self.assert_page_has("Click Submit to generate an HTML table report")
        self.submit_form()
        self.assert_title("Current Sessions")
        self.assert_single_table_report()
        columns = (
            "Started",
            "User",
            "Name",
            "Office",
            "Email",
            "Phone",
            "Last Activity",
        )
        table = self.load_table()
        table.check_headers(columns)

    def test_cwd_replacement(self):
        """Test the CWD Replacement (utility and report)."""

        # First test the tool to make the replacements.
        root = etree.Element("Country")
        etree.SubElement(root, "CountryFullName").text = "Freedonia"
        etree.SubElement(root, "Continent").text = "Europe"
        etree.SubElement(root, "PostalCodePosition").text = "after City"
        xml = etree.tostring(root, encoding="unicode")
        opts = dict(publishable=True, validate=True)
        doc_id = self.save_doc(xml, "Country", **opts)
        opts["id"] = cdr_id = f"CDR{doc_id:010d}"
        comment = "Because we like to test."
        xml = xml.replace("Freedonia", "Freedonya")
        self.save_doc(xml, "Country", **opts)
        self.navigate_to("ReplaceCWDwithVersion.py")
        self.assert_title("Replace CWD With Older Version")
        self.assert_page_has("Instructions")
        self.assert_page_has("This program will replace the current")
        self.set_field_value("id", doc_id)
        self.set_field_value("version", -2)
        self.set_field_value("comment", comment)
        self.submit_form()
        self.assert_title("Replace CWD With Older Version")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        self.assert_page_has("Click Confirm to perform action(s).")
        self.assert_page_has("Click Cancel to start over.")
        table = self.load_table()
        self.assertEqual(len(table.rows), 8)
        expected = (
            ("Replace CWD for", cdr_id),
            ("Document type", "Country"),
            ("Document title", "Freedonya"),
            ("Current total versions", "2"),
            ("Make this version the CWD", "1"),
            ("Also create a new version", "True"),
            ("Make version publishable", "True"),
            ("Reason to be logged", comment)
        )
        for i, (label, value) in enumerate(expected):
            self.assertEqual(table.rows[i][0].text, label)
            self.assertEqual(table.rows[i][1].text, value)
        self.click("submit-button-confirm")
        self.assert_title("Replace CWD With Older Version")
        self.assert_page_has(f"Successfully updated CDR{doc_id}.")
        self.navigate_to("ReplaceCWDReport.py")
        self.assert_title("Report CWD Replacements")
        self.set_field_value("start", date.today().strftime("%m/%d/%Y"))
        self.submit_form()
        self.assert_title("Report CWD Replacements")
        self.assert_page_has("Document Replacements")
        self.assert_single_table_report()
        self.assert_plain_report()
        columns = {
            "Date/time": "When did the replacement occur?",
            "DocID": "CDR ID of the affected document",
            "Doc type": "Document type for the affected document",
            "User": "ID of the user promoting the version",
            "LV": "Version number of last version after promotion",
            "PV": (
                "Version number of last publishable version at that time"
                ", -1 = None"
            ),
            "Chg": "'Y' = CWD was different from last version, else 'N'",
            "V#": "Version number promoted to become CWD",
            "V": "Was new CWD also versioned? (Y/N)",
            "P": "Was new CWD also versioned as publishable? (Y/N)",
            "Comment": "System-generated comment ':' user-entered comment",
        }
        table = self.load_table()
        caption = f"Document Replacements ({len(table.rows)})"
        self.assertEqual(table.caption.text, caption)
        headers = table.node.find_elements(By.CSS_SELECTOR, "thead th")
        self.assertEqual(len(columns), len(headers))
        for header in headers:
            tooltip = header.get_attribute("title")
            self.assertEqual(tooltip, columns[header.text])
        row = table.rows[-1]
        self.assertTrue(row[0].text.startswith(str(date.today())))
        self.assertEqual(row[1].text, str(doc_id))
        self.assertEqual(row[2].text, "Country")
        self.assertEqual(row[4].text, "3")
        self.assertEqual(row[5].text, "3")
        self.assertEqual(row[6].text, "N")
        self.assertEqual(row[7].text, "1")
        self.assertEqual(row[8].text, "Y")
        self.assertEqual(row[9].text, "Y")
        self.assertEqual(row[10].text, comment)

        # Clean up after ourselves.
        self.delete_doc(cdr_id, reason="Don't leave clutter behind.")

    def test_date_last_modified_report(self):
        """Test the Date Last Modified report."""

        self.navigate_to("DateLastModified.py")
        self.assert_title("Date Last Modified")
        doctype_id = self.doctypes["DrugInformationSummary"]
        self.click(f"doctype-{doctype_id}")
        self.set_field_value("start", "1/1/2023")
        self.set_field_value("end", "")
        self.submit_form()
        self.assert_title("Date Last Modified (since 2023-01-01)")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        self.assert_table_caption("DrugInformationSummary")
        columns = "Date Last Modified", "CDR ID", "Document Title"
        table = self.load_table()
        table.check_headers(columns)

    def test_documentation_search(self):
        """Test Documentation advanced search page."""

        self.navigate_to("HelpSearch.py")
        self.assert_title("Documentation")
        self.set_field_value("keyword", "global changes")
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Documentation")
        self.assert_tables_in_grid_container()
        self.assert_single_table_report()
        table = self.load_table()
        expected = "1 document matches 'global changes'"
        self.assertEqual(table.caption.text, expected)
        self.assertEqual(len(table.rows), 1)
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_page_has("<h1>Global Changes</h1>")

    def test_doc_viewer(self):
        """Test the document viewer."""

        # Entering a missing version number should trigger an alert.
        self.navigate_to("ShowCdrDocument.py")
        self.assert_title("CDR Document Viewer")
        doc = self.get_test_summary()
        doc_id = doc["id"]
        self.set_field_value("doc-id", doc_id)
        self.click("vtype-num")
        self.set_field_value("version", "9999999")
        self.submit_form()
        alert = self.driver.find_element(By.CSS_SELECTOR, "#alerts-block p")
        self.assertIsNotNone(alert)
        self.assert_regex(rf"CDR{doc_id} has only \d+ versions.")

        # The alert should provide what we need to pick a valid version.
        latest = alert.text.split()[-2]
        self.set_field_value("version", latest)
        self.submit_form(new_tab=False)
        self.assert_page_has(doc["title"])
        self.assert_page_has("<SummaryType>Treatment</SummaryType>")

    def test_external_map_failures_report(self):
        """Test the External Map Failures report."""

        self.navigate_to("ExternMapFailures.py")
        self.assert_title("External Map Failures Report")
        self.click("usage-glossaryterm-phrases")
        self.click("usage-spanish-glossaryterm-phrases")
        self.set_field_value("age", "10000")
        self.click("options-non-mappable")
        self.submit_form()
        self.assert_title("External Map Failures Report")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        self.assert_table_caption("GlossaryTerm Phrases")
        self.assert_table_caption("Spanish GlossaryTerm Phrases")
        for table in self.load_tables():
            table.check_headers(("Value", "Recorded"))

    def test_global_change_test_results(self):
        """Test the Global Change Test Results pages."""

        self.navigate_to("ShowGlobalChangeTestResults.py")
        self.assert_title("Global Change Test Results")
        self.assert_page_has("Choose Test Run Job")
        self.assert_not_found("submit-button-submit", method=By.ID)
        self.find("fieldset li a").click()
        results_page = self.select_new_tab()
        self.assert_title("Global Change Test Results")
        self.assert_single_table_report()
        patterns = (
            r"Test job run at \d{4}-\d\d-\d\d \d\d:\d\d:\d\d",
            r"Total number of documents = \d+",
            r"Total number of versions = \d+",
        )
        for expected in patterns:
            self.assert_regex(expected)
        table = self.load_table()
        headers = "CDR ID", "Ver.", "Files", "New Size", "Diff Size"
        table.check_headers(headers)

        # The table should start out in order by descending diff size.
        first_size = int(table.rows[0][-1].text)
        last_size = int(table.rows[-1][-1].text)
        self.assertGreaterEqual(first_size, last_size)

        # Verify that sort by CDR works.
        self.assert_page_has("Sort By CDR ID")
        self.find("#primary-form input").click()
        self.assert_page_not_has("Sort By CDR ID")
        table = self.load_table()
        first_id = table.rows[0][0].text
        last_id = None
        index = len(table.rows) - 1
        while index >= 0:
            if len(table.rows[index]) == 5:
                last_id = table.rows[index][0].text
                break
            index -= 1
        self.assertIsNotNone(last_id)
        self.assertLessEqual(first_id, last_id)

        # Same check for sorting by document size.
        self.assert_page_has("Sort By Document Size")
        self.find("#primary-form input").click()
        self.assert_page_not_has("Sort By Document Size")
        table = self.load_table()
        first_size = int(table.rows[0][-2].text)
        last_size = int(table.rows[-1][-2].text)
        self.assertGreaterEqual(first_size, last_size)

        # Verify that the links are present and work.
        links = table.rows[0][2].find_elements(By.TAG_NAME, "a")
        self.assertGreaterEqual(len(links), 3)
        for link in links:
            link.click()
            self.select_new_tab()
            self.assertIsNotNone(self.find("body pre"))
            self.switch_to(results_page)

    def test_help_pages(self):
        """Test the CDR documentation pages."""

        self.navigate_to("Help.py")
        self.assert_title("CDR Help")
        self.assert_page_has("Primarily For Users")
        self.assert_page_has("Primarily For Developers")
        self.find("form h4").click()
        self.assert_page_has("Creating/Editing Documents")
        self.find("main ul li a").click()
        self.select_new_tab()
        self.assert_page_has("<h1>Create/Edit Documents</h1>")

    def test_home_page(self):
        """Test the landing page for the site."""

        # The page should have cards on it showing dasehboard information.
        # For this to pass, the session used for the tests must be for an
        # account which can see more than one menu group.
        self.navigate_to("Admin.py")
        self.assertEqual(self.driver.title, "CDR Administration")
        self.assert_page_has("usa-card")
        self.assert_page_has("Last full publishing job")

        # Confirm that the search form at the top of the page works.
        summary = self.get_test_summary()
        search_box = self.driver.find_element(By.ID, "search-field")
        search_box.send_keys(str(summary["id"]))
        search_box.submit()
        self.select_new_tab()
        self.assert_page_has(summary["title"])
        self.assert_page_has("<SummaryType>Treatment</SummaryType>")

    def test_inactivity_report(self):
        """Test Checked-Out Documents With No Activity report."""

        self.navigate_to("InactivityReport.py")
        self.assert_title("Checked Out Documents With No Activity")
        self.assert_page_has("Inactivity Threshold (up to 99 days)")
        self.submit_form()
        self.assert_title(
            "Checked Out Documents With No Activity for Longer Than 10 Days"
        )
        self.assert_table_caption("Inactive Documents")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        columns = (
            "Document ID",
            "Type",
            "User",
            "Checked Out",
            "Last Action",
            "Action Date",
        )
        table = self.load_table()
        table.check_headers(columns)

    def test_invalid_docs_report(self):
        """Test the Invalid Documents report."""

        self.navigate_to("InvalidDocs.py")
        self.assert_title("Invalid Documents")
        self.assert_page_has("Select Document Type")
        doctype_id = self.doctypes["DrugInformationSummary"]
        self.click(f"doctype-{doctype_id}")
        self.submit_form()
        self.assert_title("Invalid Documents")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        tables = self.load_tables()
        captions = (
            "Invalid DrugInformationSummary Documents",
            "Blocked DrugInformationSummary Documents",
        )
        self.assertEqual(len(tables), len(captions))
        for i, table in enumerate(tables):
            self.assertEqual(table.caption.text, captions[i])
            table.check_headers(("ID", "Title"))

    def test_job_status_report(self):
        """Test the Batch Job Status report."""

        self.navigate_to("getBatchStatus.py")
        self.assert_title("Batch Job Status")
        self.set_field_value("jobName", "Broken URLs")
        self.set_field_value("jobAge", 365)
        self.select_values("jobStatus", "Completed")
        self.submit_form()
        self.assert_title("Batch Job Status")
        self.assert_single_table_report()
        table = self.load_table()
        headers = (
            "ID",
            "Job Name",
            "Started",
            "Status",
            "Last Info",
            "Last Message",
        )
        last_year = date.today() - timedelta(365)
        table.check_headers(headers)
        self.assertRegex(table.rows[0][0].text, r"\d+")
        self.assertEqual(table.rows[0][1].text, "Broken URLs")
        self.assertGreaterEqual(table.rows[0][2].text, str(last_year))
        self.assertEqual(table.rows[0][3].text, "Completed")
        self.assertGreaterEqual(table.rows[0][4].text, str(last_year))
        self.assertRegex(table.rows[0][5].text, r"Checked \d+ urls for \d+")

    def test_linked_docs_report(self):
        """Test the Linked Documents report."""

        doc = self.get_test_summary()
        doc_id = doc["id"]
        doc_title = doc["title"]
        self.navigate_to("LinkedDocs.py")
        self.assert_title("Linked Documents Report")
        self.set_field_value("doc_id", doc_id)
        self.submit_form()
        self.assert_title("Linked Documents Report")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        self.assert_table_caption("Target Document")
        self.assert_table_caption("Links From Summary (English HP) Documents")
        self.assert_page_has(f"{doc_title};Treatment;Health professionals")
        self.assert_page_has(doc_id)
        columns = "Doc ID", "Doc Title", "Linking Element", "Fragment ID"
        for column in columns:
            self.assert_regex(f"<th[^>]*>{column}</th>")

    def test_locked_docs_report(self):
        """Test Checked-Out Documents report."""

        self.navigate_to("CheckedOutDocs.py")
        self.assert_title("Checked Out Documents")
        option = self.driver.find_element(By.CSS_SELECTOR, "#User option")
        self.assertIsNotNone(option)
        re_match = re_search(r"(.*) \((\d+) locks\)", option.text)
        self.assertIsNotNone(re_match)
        user = re_match.group(1)
        expected_count = int(re_match.group(2))
        self.submit_form()
        self.assert_title("Checked Out Documents")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        self.assert_table_caption(f"Checked out by {user}")
        rows = self.driver.find_elements(By.CSS_SELECTOR, "main tbody tr")
        self.assertEqual(len(rows), expected_count)
        columns = "Checked Out", "Type", "CDR ID", "Document Title"
        for column in columns:
            self.assert_regex(rf"<th[^>]*>{column}</th>")

    def test_miscellaneous_doc_qc_report(self):
        """Test the Miscellanous Document QC report."""

        self.navigate_to("MiscSearch.py")
        self.assert_title("Miscellaneous Documents")
        self.select_values("type", "Cover letter")
        selector = 'main form input[type="submit"]'
        self.driver.find_element(By.CSS_SELECTOR, selector).click()
        self.select_new_tab()
        self.assert_regex("documents? matche?s? 'Cover letter'")
        self.driver.find_element(By.CSS_SELECTOR, "main td a").click()
        self.assert_page_has("Miscellaneous Document")
        self.assert_page_has("QC Report")
        self.assert_multi_table_report()
        self.assert_plain_report()

    def test_modified_docs_report(self):
        """Test Modified Since Last Publishable Version report."""

        self.navigate_to("ModWithoutPubVersion.py")
        self.assert_title("Documents Modified Since Last Publishable Version")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        title = "Documents Modified Between 2020-01-01 And 2020-01-31"
        self.assert_title(title)
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        columns = (
            "Doc ID",
            "Latest Publishable Version Date",
            "Modified By",
            "Modified Date",
            "Latest Non-publishable Version Date",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

    def test_new_docs_reports(self):
        """Test the reports on new CDR documents."""

        # The first report shows new documents with publication status.
        self.navigate_to("NewDocsWithPubStatus.py")
        self.assert_title("New Documents With Publication Status")
        self.click("type-druginformationsummary")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title(
            "New DrugInformationSummary Documents "
            "Created 2020-01-01 – 2020-01-31"
        )
        self.assert_table_caption("DrugInformationSummary")
        self.assert_single_table_report()
        self.assert_wide_report()
        self.driver.set_window_size(1400, 1024)
        columns = (
            "CDR ID",
            "Document Title",
            "Created By",
            "Creation Date",
            "Latest Version Date",
            "Latest Version By",
            "Pub?",
            "Earlier Pub Ver?",
        )
        table = self.load_table()
        table.check_headers(columns)

        # The other report shows statistical information about new docs.
        self.navigate_to("NewDocReport.py")
        self.assert_title("New Documents Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title(
            "New Documents Created Between 2020-01-01 and 2020-01-31"
        )
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        for table in self.load_tables():
            table.check_headers(("Status", "Count"))

    def test_organization_qc_report(self):
        """Test the Organization QC report."""

        self.navigate_to("OrgSearch2.py")
        self.assert_title("Organization")
        self.set_field_value("city", "Bethesda")
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Organization")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        pattern = r"(\d+) documents match 'Bethesda'"
        re_match = re_search(pattern, table.caption.text)
        self.assertIsNotNone(re_match)
        target = "National Cancer Institute; NCI; Bethesda; Maryland"
        link = None
        for row in table.rows:
            if row[2].text == target:
                link = row[1].find_element(By.TAG_NAME, "a")
                break
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Organization<br>QC Report")
        self.assert_page_has(date.today())
        self.assert_page_has("National Cancer Institute")

    def test_person_qc_report(self):
        """Test the Person search page."""

        doc = self.get_test_board_member()
        id = doc["person"]["id"]
        surname = doc["person"]["surname"]
        forename = doc["person"]["forename"]
        self.navigate_to("PersonSearch.py")
        self.assert_title("Person")
        self.set_field_value("surname", surname)
        self.set_field_value("forename", forename)
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Person")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        self.assertEqual(table.caption.text, f"1 document matches '{surname}'")
        self.assertEqual(len(table.rows), 1)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][1].text, f"CDR{id:010d}")
        self.assertTrue(table.rows[0][2].text.startswith(surname))
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_plain_report()
        self.assert_page_has("<center>Person<br>QC Report</center>")
        self.assert_page_has(surname)
        self.assert_page_has(forename)
        self.assert_page_has(f"CDR{id}")

    def test_persons_with_locations_search(self):
        """Test the Persons With Locations search page."""

        doc = self.get_test_board_member()
        id = doc["person"]["id"]
        surname = doc["person"]["surname"]
        forename = doc["person"]["forename"]
        self.navigate_to("PersonLocSearch.py")
        self.assert_title("Persons With Locations")
        self.set_field_value("surname", surname)
        self.set_field_value("forename", forename)
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Persons With Locations")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected = f"1 document matches '{surname}' and '{forename}'"
        self.assertEqual(table.caption.text, expected)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][1].text, f"CDR{id:010d}")
        self.assertTrue(table.rows[0][2].text.startswith(surname))
        row_nodes = table.node.find_elements(By.CSS_SELECTOR, "tbody tr")
        if row_nodes[0].get_attribute("class") == "has-addresses":
            self.assertEqual(len(table.rows), 2)
            expected = "person-addresses"
            self.assertEqual(row_nodes[1].get_attribute("class"), expected)
            self.assertEqual(table.rows[1][0].get_attribute("colspan"), "3")
            address = table.rows[1][0].find_element(By.CSS_SELECTOR, "ul li")
            self.assertIsNotNone(address)
        else:
            self.assertEqual(len(table.rows), 1)
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_plain_report()
        self.assert_page_has("<center>Person<br>QC Report</center>")
        self.assert_page_has(surname)
        self.assert_page_has(forename)
        self.assert_page_has(f"CDR{id}")

    def test_political_subunit_search(self):
        """Test the Political Subunit search page."""

        self.navigate_to("PoliticalSubUnitSearch.py")
        self.assert_title("Political SubUnit")
        self.set_field_value("state", "North Dak%")
        self.find('form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Political SubUnit")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        self.assertEqual(table.caption.text, "1 document matches 'North Dak%'")
        self.assertEqual(len(table.rows), 1)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][2].text, "North Dakota")
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_page_has("<center>Political SubUnit<br>QC Report</center>")
        self.assert_page_has("North Dakota")
        self.assert_page_has("ND")

    def test_spanish_spellcheck_file_creation(self):
        """Test the utility to generate spellcheck files."""

        script = "SpanishSpellcheckerFiles.py"
        self.navigate_to(script)
        self.assert_title("Create Spanish Spellcheck Files from Dictionary")
        self.assert_page_has("Instructions")
        self.assert_page_has("Select the audience for which to create")
        params = dict(audience="hp", Request="Submit")
        url = f"{self.cgi}/{script}?{urlencode(params)}"
        data = str(self.fetch_from_url(url), "utf-8")
        self.assertGreater(len(data.split("\r\n")), 300)

    def test_statuses(self):
        """Test pages for managing document statuses."""

        # Test unblocking of an inactive document.
        title = "Test summary for automated test of statuses"
        xml = f"<Summary><SummaryTitle>{title}</SummaryTitle></Summary>"
        doc_id = self.save_doc(xml, "Summary", block=True)
        cdr_id = f"CDR{doc_id:010d}"
        self.navigate_to("UnblockDoc.py")
        self.assert_title("Unblock CDR Document")
        self.assert_page_has("Document To Be Unblocked")
        self.set_field_value("id", doc_id)
        self.submit_form(new_tab=False)
        self.assert_title("Unblock CDR Document")
        self.assert_page_has(f"Successfully unblocked {cdr_id}.")

        # Confirm that attempting to unblock a document which is not
        # blocked results in a warning message.
        self.set_field_value("id", doc_id)
        self.submit_form(new_tab=False)
        self.assert_title("Unblock CDR Document")
        self.assert_page_has(f"{cdr_id} is not blocked.")

        # Test deletion of the document.
        self.navigate_to("del-some-docs.py")
        self.assert_title("CDR Document Deletion")
        self.assert_page_has("Instructions")
        self.set_field_value("ids", cdr_id)
        self.click("options-validate")
        self.submit_form(new_tab=False)
        self.assert_title("CDR Document Deletion")
        self.assert_page_has(f"{cdr_id} has been deleted successfully.")

        # Test resurrection of the document.
        self.navigate_to("RestoreDeletedDocs.py")
        self.assert_title("CDR Document Restoration")
        self.assert_page_has("Instructions")
        self.set_field_value("ids", cdr_id)
        self.submit_form(new_tab=False)
        self.assert_title("CDR Document Restoration")
        self.assert_page_has(f"{cdr_id} restored successfully.")

        # Clean up behind ourselves.
        sql = (
            "SELECT doc_id"
            "  FROM query_term"
            " WHERE path = '/Summary/SummaryTitle'"
            f"   AND value = '{title}'"
        )
        ids = [row[0] for row in self.run_query(sql)]
        self.navigate_to("del-some-docs.py")
        self.assert_title("CDR Document Deletion")
        self.set_field_value("ids", " ".join(ids))
        self.click("options-validate")
        self.submit_form(new_tab=False)
        self.assert_title("CDR Document Deletion")
        self.assert_page_has("deleted successfully.")

    def test_unchanged_docs_report(self):
        """Test the Unchanged Documents report."""

        self.navigate_to("UnchangedDocs.py")
        self.assert_title("Unchanged Documents")
        self.submit_form()
        self.assert_title("Unchanged Documents")
        self.assert_single_table_report()
        caption = f"Documents Unchanged for 365 Days as of {date.today()}"
        table = self.load_table()
        self.assertEqual(table.caption.text, caption)
        self.assert_table_caption(caption)
        columns = "Doc ID", "Doc Title", "Last Change"
        table.check_headers(columns)

    def test_url_check_report(self):
        """Test the URL Check report."""

        self.navigate_to("CheckUrls.py")
        self.assert_title("URL Check")
        self.click("opt-quick")
        self.set_field_value("max-seconds", "5")
        self.submit_form()
        self.assert_title("CDR Report on Inactive Hyperlinks")
        self.assert_page_has("<h2>Citation</h2>")
        self.assert_plain_report()
        self.assert_single_table_report()
        columns = "CDR ID", "Title", "Stored URL", "Problem", "Element"
        for i, th in enumerate(self.find("table tr:first-child th", all=True)):
            self.assertEqual(th.text, columns[i])

    def test_url_list_report(self):
        """Test the URL List report."""

        self.navigate_to("UrlListReport.py")
        self.assert_title("URL List Report")
        self.assert_page_has("Select Document Type For Report")
        self.click("doctype-druginformationsummary")
        self.submit_form()
        self.assert_title("URL List Report")
        self.assert_single_table_report()
        self.assert_plain_report()
        self.assert_page_has("Document Type: DrugInformationSummary")
        columns = "Doc ID", "Doc Title", "URL", "Display Text", "Source Title"
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

    def test_version_history_report(self):
        """Test the Document Version History report."""

        self.navigate_to("DocVersionHistory.py")
        self.assert_title("Document Version History Report")
        self.assert_page_has("Specify Document ID or Title")
        doc = self.get_test_summary()
        doc_id = doc["id"]
        doc_title = doc["title"]
        self.set_field_value("DocId", doc_id)
        self.submit_form()
        self.assert_title("Document Version History Report")
        self.assert_page_has(f"CDR{doc_id:010d} (Summary)")
        self.assert_page_has(f"{doc_title};Treatment;Health professionals")
        self.assert_page_has("Created")
        self.assert_page_has("Updated")
        columns = (
            "Ver",
            "Comment",
            "Date",
            "User",
            "Val",
            "Pub?",
            "Publication Date(s)",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")
        self.assert_tables_in_grid_container()
        self.assert_multi_table_report()

    def test_xmetal_cdr_icons_report(self):
        """Test the XMetaL CDR Icons report."""

        self.navigate_to("xmetal-icons.py")
        self.assert_title("XMetaL CDR Icons")
        self.assert_page_has("Instructions")
        self.assert_page_has("Press Submit to bring up a page showing")
        self.submit_form()
        self.assert_title("XMetaL CDR Icons")
        self.assert_page_has("Annotations (Custom)")
        images = self.driver.find_elements(By.TAG_NAME, "img")
        self.assertEqual(len(images), 13)


class GlossaryTests(Tester):
    """Tests for the GlossaryTermConcept and GlossaryTermName documents."""

    def test_documents_linked_to_gtn_report(self):
        """Test the Glossary Term Links QC report."""

        term = self.get_test_glossary_term()
        term_id = term["id"]
        self.navigate_to("GlossaryTermLinks.py")
        self.assert_title("Glossary Term Links QC Report")
        self.assert_page_has("Select a Glossary Term by Name or ID")
        self.set_field_value("id", term_id)
        self.submit_form()
        self.assert_tables_in_grid_container()
        self.assert_multi_table_report()
        self.assert_regex("(?s)Doc ID.+Doc Title.+Element Name.+Fragment ID")

    def test_external_map(self):
        """Test the external map (now only used for glossary terms)."""

        # First make sure we have no residue from a previous, failed test.
        test_phrase = "test phrase for mapping"
        test_name = "test term for mapping"
        usage = "GlossaryTerm Phrases"
        deleted_message = (
            f"Change(s) to {usage} mapping of {test_phrase}: "
            "Mapping deleted."
        )
        mapping_page = self.navigate_to("EditExternalMap.py")
        self.assert_title("External Map Editor")
        usages = {}
        for option in self.find("#usage option", all=True):
            usages[option.text] = option.get_attribute("value")
        self.select_values("usage", usages[usage])
        self.set_field_value("value_pattern", test_phrase)
        self.click("options-also-include-unmappable-values")
        self.click("submit-button-get-values")
        no_mappings = "No mappings found matching the filtering criteria."
        if no_mappings not in self.get_page_source():
            self.logger.debug("Cleaning up leftover test external map entry.")
            table = self.load_table()
            self.assertIsNotNone(table)
            row = table.rows[0]
            self.assertEqual(row[0].text, test_phrase)
            row[1].find_element(By.TAG_NAME, "input").clear()
            delete_checkbox = row[-3].find_element(By.TAG_NAME, "input")
            self.click(delete_checkbox.get_attribute("id"))
            self.click("submit-button-save-changes")
            self.assert_page_has(deleted_message)
            self.assert_page_has(no_mappings)

        # Create a test glossary document and a test mapping.
        doc_id = self.create_test_gtn(name=test_name)
        mapping_id = self.create_external_mapping(test_phrase, usage=usage)
        delete_id = f"delete-{mapping_id}"
        bogus_id = f"bogus-{mapping_id}"
        mappable_id = f"mappable-{mapping_id}"
        cdrid_id = f"cdrid-{mapping_id}"

        # Refresh the mapping editor and verify that the new mapping shows up.
        self.click("submit-button-get-values")
        table = self.load_table()
        self.assertIsNotNone(table)
        self.assertEqual(len(table.rows), 1)
        row = table.rows[0]
        self.assertEqual(row[0].text, test_phrase)
        self.assertEqual(row[1].find_element(By.TAG_NAME, "input").text, "")
        self.assertEqual(row[2].text, "")
        self.assertFalse(self.find(delete_id, By.ID).is_selected())
        self.assertFalse(self.find(bogus_id, By.ID).is_selected())
        self.assertTrue(self.find(mappable_id, By.ID).is_selected())

        # Bogus and mappable cannot both be selected at the same time.
        self.click(bogus_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Error(s) in GlossaryTerm Phrases mapping of test phrase for "
            "mapping: Mapping cannot be both bogus and unmappable."
        )
        self.assertFalse(self.find(bogus_id, By.ID).is_selected())
        self.assertTrue(self.find(mappable_id, By.ID).is_selected())

        # The software should detect a missing or inappropriate mapped doc.
        self.set_field_value(cdrid_id, "999999")
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Error(s) in GlossaryTerm Phrases mapping of test phrase "
            "for mapping: CDR document 999999 not found."
        )
        value = self.find(cdrid_id, method=By.ID).get_attribute("value")
        self.assertEqual(value, "")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        self.set_field_value(cdrid_id, summary_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Error(s) in GlossaryTerm Phrases mapping of test phrase for "
            "mapping: GlossaryTerm Phrases mappings can't link to Summary "
            f"document CDR{summary_id}."
        )
        value = self.find(cdrid_id, method=By.ID).get_attribute("value")
        self.assertEqual(value, "")

        # An attempt to map the value with the "unmappable" flag set
        # should be rejected.
        self.set_field_value(cdrid_id, doc_id)
        self.click(mappable_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Error(s) in GlossaryTerm Phrases mapping of test phrase for "
            "mapping: Mapped value must be flagged 'mappable'."
        )
        self.assertTrue(self.find(mappable_id, By.ID).is_selected())
        value = self.find(cdrid_id, method=By.ID).get_attribute("value")
        self.assertEqual(value, "")

        # Mapping to an appropriate document should succeed, and a View
        # button should appear which we can use to visit the glossary
        # document's QC report.
        self.set_field_value(cdrid_id, doc_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Change(s) to GlossaryTerm Phrases mapping of test phrase "
            f"for mapping: Mapping to CDR{doc_id} added."
        )
        value = self.find(cdrid_id, method=By.ID).get_attribute("value")
        self.assertEqual(value, str(doc_id))
        table = self.load_table()
        self.assertIsNotNone(table)
        self.assertEqual(len(table.rows), 1)
        row = table.rows[0]
        self.assertEqual(row[0].text, test_phrase)
        id_field = row[1].find_element(By.TAG_NAME, "input")
        self.assertEqual(id_field.get_attribute("value"), str(doc_id))
        self.assertEqual(row[2].text, "View")
        row[2].find_element(By.TAG_NAME, "a").click()
        self.select_new_tab()
        expected = "<center>Glossary Term Name<br>QC Report</center>"
        self.assert_page_has(expected)
        self.assert_page_has(f"CDR{doc_id}")
        self.assert_page_has(test_name)

        # It should not be possible to delete the mapping at this stage.
        self.switch_to(mapping_page)
        self.click(delete_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(
            "Error(s) in GlossaryTerm Phrases mapping of test phrase for "
            "mapping: Deletion request blocked because the value is mapped."
        )
        self.assertFalse(self.find(delete_id, By.ID).is_selected())
        value = self.find(cdrid_id, method=By.ID).get_attribute("value")
        self.assertEqual(value, str(doc_id))

        # But if we unlink the phrase from the document, deletion should work.
        self.set_field_value(cdrid_id, "")
        self.click(delete_id)
        self.click("submit-button-save-changes")
        self.assert_page_has(deleted_message)
        self.assert_page_has(no_mappings)

        # Finally, remove the test glossary document.
        self.delete_doc(doc_id)

    def test_glossary_servers(self):
        """Test the interface for managing where to send glossary data."""

        # We'll need these multiple times locally.
        test_alias = "Bogus Server For Automated Testing"
        test_url = "https://example.com"
        bogus_url = "not a real URL"

        def get_server_count():
            field = self.find("input[name='num-servers']")
            self.assertIsNotNone(field)
            return int(field.get_attribute("value"))

        def find_test_server():
            blocks = self.find("fieldset.server-block", all=True)
            self.assertGreaterEqual(len(blocks), 1)
            for block in blocks:
                alias = block.find_element(By.CSS_SELECTOR, "input.alias")
                self.assertIsNotNone(alias)
                value = alias.get_attribute("value")
                if value == test_alias:
                    return block
            return None

        def add_server_block():
            img = self.find("img[title='Add another server']")
            self.assertIsNotNone(img)
            img.click()
            sleep(1)

        # Before we do anything else, clear out any leftover test data.
        self.navigate_to("glossary-servers.py")
        self.assert_title("Manage Glossary Servers")
        self.assert_page_has("Instructions")
        self.assert_page_has("Use this form to manage which servers will")
        test_server_block = find_test_server()
        while test_server_block is not None:
            self.logger.warning("Clearing out old test server.")
            server_count = get_server_count() - 1
            selector = "img[title='Remove server']"
            img = test_server_block.find_element(By.CSS_SELECTOR, selector)
            self.assertIsNotNone(img)
            img.click()
            self.submit_form(new_tab=False)
            s = "" if server_count == 1 else "s"
            message = f"Successfully stored {server_count} glossary server{s}."
            self.assert_page_has(message)
            test_server_block = find_test_server()

        # Confirm that there is at least one server on the form.
        original_server_count = get_server_count()
        self.assertGreaterEqual(original_server_count, 1)

        # Verify proper handling of error conditions.
        i = original_server_count + 1
        add_server_block()
        self.set_field_value(f"alias-{i}", test_alias)
        self.set_field_value(f"url-{i}", test_url)
        i += 1
        add_server_block()
        self.set_field_value(f"alias-{i}", test_alias)
        self.set_field_value(f"url-{i}", test_url)
        i += 1
        add_server_block()
        self.set_field_value(f"alias-{i}", test_alias)
        self.set_field_value(f"url-{i}", bogus_url)
        self.submit_form(new_tab=False)
        self.assert_title("Manage Glossary Servers")
        expected = (
            f"Duplicate alias {test_alias!r}.",
            f"{test_url!r} appears more than once.",
            f"{bogus_url!r} is not an HTTP URL."
        )
        for message in expected:
            self.assert_page_has(message)
        self.assertEqual(get_server_count(), original_server_count)

        # Add the test server and verify that it was stored.
        i = original_server_count + 1
        add_server_block()
        self.set_field_value(f"alias-{i}", test_alias)
        self.set_field_value(f"url-{i}", test_url)
        self.submit_form(new_tab=False)
        new_server_count = original_server_count + 1
        self.assertEqual(get_server_count(), new_server_count)
        server_block = find_test_server()
        self.assertIsNotNone(server_block)
        self.assert_title("Manage Glossary Servers")
        s = "" if new_server_count == 1 else "s"
        message = f"Successfully stored {new_server_count} glossary server{s}."
        self.assert_page_has(message)

        # Remove the test server and verify that it's gone.
        selector = "img[title='Remove server']"
        img = server_block.find_element(By.CSS_SELECTOR, selector)
        self.assertIsNotNone(img)
        img.click()
        self.submit_form(new_tab=False)
        n = original_server_count
        s = "" if n == 1 else "s"
        message = f"Successfully stored {n} glossary server{s}."
        self.assert_page_has(message)
        test_server_block = find_test_server()
        self.assertIsNone(test_server_block)
        self.assertEqual(get_server_count(), original_server_count)

    def test_gtc_advanced_search(self):
        """Test the Glossary Term Concept advanced search page."""

        self.navigate_to("GlossaryTermConceptSearch.py")
        self.assert_title("Glossary Term Concept")
        self.set_field_value("concept", "%sequence%")
        self.select_values("audience", "Patient")
        self.select_values("dictionary", "Genetics")
        self.find('main form input[value="Search"]').click()
        self.select_new_tab()
        self.assert_title("Glossary Term Concept")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        pattern = (
            r"(\d+) documents match '%sequence%' and 'Patient' "
            "and 'Genetics'"
        )
        re_match = re_search(pattern, table.caption.text)
        self.assertIsNotNone(re_match)
        count = int(re_match.group(1))
        self.assertEqual(len(table.rows), count)
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_plain_report()
        self.assert_page_has("Glossary Term Concept<br>QC Report")

    def test_gtc_by_english_definition_status_report(self):
        """Test the GTC by English Definition Status report."""

        params = dict(report="English")
        self.navigate_to("GlossaryConceptByDefinitionStatus.py", **params)
        self.assert_title("GTC by English Definition Status")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_page_has("Patient Concepts With Approved Status")
        self.assert_page_has("2020-01-01 to 2020-01-31")
        columns = (
            "CDR ID of GTC",
            "Term Name (Pronunciation)",
            "Definition",
            "Definition Resource"
        )
        for column in columns:
            self.assert_page_has(column)
        self.assert_plain_report()
        self.assert_single_table_report()

    def test_gtc_by_spanish_definition_status_report(self):
        """Test the GTC by Spanish Definition Status report."""

        params = dict(report="Spanish")
        self.navigate_to("GlossaryConceptByDefinitionStatus.py", **params)
        self.assert_title("GTC by Spanish Definition Status")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.click("opts-english")
        self.click("opts-resources")
        self.click("opts-notes")
        self.submit_form()
        self.assert_page_has("Patient Concepts With Approved Status")
        self.assert_page_has("2020-01-01 to 2020-01-31")
        columns = (
            "CDR ID of GTC",
            "Term Name (EN)",
            "Term Name (ES)",
            "Definition (EN)",
            "Definition (ES)",
            "Comment",
            "Translation Resource",
            "QC Notes",
        )
        for column in columns:
            self.assert_page_has(column)
        self.assert_plain_report()
        self.assert_single_table_report()

    def test_gtc_by_type_report(self):
        """Test the Glossary Term Concept By Type report."""

        self.navigate_to("GlossaryConceptByTypeReport.py")
        self.assert_title("Glossary Term Concept By Type Report")
        self.select_values("type", "Diagnosis")
        self.select_values("statuses", "Approved")
        self.select_values("audience", "Patient")
        self.set_field_value("name", "cancer")
        self.click("spanish-y")
        self.submit_form()
        self.assert_title("Glossary Term Concept By Type Report")
        expected = "Glossary Term Concept By Type Report - English & Spanish"
        self.assert_page_has(html_escape(expected))
        self.assert_page_has("Diagnosis")
        self.assert_page_has(date.today())
        self.assert_page_has("<th>CDR ID of GTC</th>")
        self.assert_page_has("<th>Term Names (English)</th>")
        self.assert_page_has("<th>Term Names (Spanish)</th>")
        self.assert_page_has("<th>Definition (English)</th>")
        self.assert_page_has("<th>Definition (Spanish)</th>")
        self.assert_wide_report()
        self.assert_single_table_report()

    def test_gtc_documents_modified_report(self):
        """Test GTC Documents Modified report."""

        script = "GlossaryConceptDocsModified.py"
        self.navigate_to(script)
        self.assert_title("GTC Documents Modified Report")
        self.assert_page_has("Date Range")
        self.assert_page_has("Language")
        self.assert_page_has("Audience")
        self.assert_page_has("Instructions")
        self.assert_page_has("Specify the date range ")
        params = dict(
            start_date="2020-01-01",
            end_date="2020-01-31",
            language="en",
            audience="Patient",
            Request="Submit",
        )
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "Glossary Term Concepts")
        self.assertEqual(sheet["A1"].value, "CDR ID")
        self.assertEqual(sheet["B1"].value, "Date Last Modified")
        self.assertEqual(sheet["C1"].value, "Publishable?")
        self.assertEqual(sheet["D1"].value, "Date First Published (*)")
        self.assertEqual(sheet["E1"].value, "Last PT Comment")
        footnote = (
            "(*) Date any GlossaryTermName document linked to the concept "
            "document was first published."
        )
        self.assertEqual(sheet[f"A{sheet.max_row}"].value, footnote)

    def test_gtc_qc_report(self):
        """Test the GlossaryTermConcept QC report."""

        self.navigate_to("QcReport.py", DocType="GlossaryTermConcept")
        self.assert_page_has("Glossary Term Concept QC Report")
        term = self.get_test_glossary_term()
        concept_id = term["concept_id"]
        self.set_field_value("DocId", concept_id)
        self.submit_form()
        self.assert_page_has("QC Report")
        self.assert_page_has(f"CDR{concept_id}")
        self.assert_page_has("Definition [en]")
        self.assert_page_has("Definition [es]")
        self.assert_page_has("Processing Statuses")

    def test_gtc_qc_report_full(self):
        """Test the Glossary Full QC report."""

        # This report shows the term names along with the concept.
        term = self.get_test_glossary_term()
        concept_id = term["concept_id"]
        params = {"initial-menu-page": "true"}
        form = self.navigate_to("GlossaryConceptFull.py", **params)
        self.assert_page_has("Glossary QC Report - Full")
        self.assert_page_has("Enter Document ID or Term Name")
        self.set_field_value("name", term["english_name"])
        self.submit_form()
        self.assert_page_has("QC Report")
        self.assert_page_has(date.today().ctime()[:10])
        self.assert_page_has(f"CDR{concept_id}")
        self.assert_page_has("<h2>English - Patient</h2>")
        self.assert_page_has("<h2>Spanish - Patient</h2>")
        self.assert_page_has('<div class="lang-wrapper">')

        # There is ab option to change from the default side-by-side layout
        # to a stacked layout with English on top of the Spanish. The
        # "lang-wrapper" element is what lays the languages out side-by-side.
        # This option removes that wrapper.
        self.switch_to(form)
        self.click("layout-english-and-spanish-stacked")
        self.assert_page_not_has('<div class="lang-wrapper">')

    def test_gtn_advanced_search(self):
        """Test Glossary Term Name advanced search."""

        self.navigate_to("GlossaryTermNameSearch.py")
        self.assert_title("Glossary Term Name")
        term = self.get_test_glossary_term()
        term_id = term["id"]
        english_name = term["english_name"]
        spanish_name = None
        for candidate in term["spanish_names"]:
            if not candidate["alternate"]:
                spanish_name = candidate["name"]
                break
        self.assertIsNotNone(spanish_name)
        self.set_field_value("name_en", english_name)
        self.find("main form input[value='Search']").click()
        self.select_new_tab()
        self.assert_title("Glossary Term Name")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        self.assertEqual(table.caption.text, "1 document matches 'cancer'")
        self.assertEqual(len(table.rows), 1)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][1].text, f"CDR{term_id:010d}")
        expected = f"{english_name}; {spanish_name} [es]"
        self.assertEqual(table.rows[0][2].text, expected)

    def test_gtn_documents_modified_report(self):
        """Test GTN Documents Modified report."""

        script = "GlossaryNameDocsModified.py"
        self.navigate_to(script)
        self.assert_title("GTN Documents Modified Report")
        self.assert_page_has("Date Range")
        self.assert_page_has("Language")
        self.assert_page_has("Audience")
        self.assert_page_has("Term Status(es)")
        selected = (
            "language-en",
            "audience-health-professional",
            "audience-patient",
            "status-approved",
            "status-new-pending",
            "status-rejected",
            "status-revision-pending",
        )
        for id in selected:
            self.assertTrue(self.driver.find_element(By.ID, id).is_selected)
        params = [
            ("start_date", "2020-01-01"),
            ("end_date", "2020-01-31"),
            ("language", "en"),
            ("audience", "Patient"),
            ("audience", "Health Professional"),
            ("status", "Approved"),
            ("status", "New pending"),
            ("status", "Rejected"),
            ("status", "Revision pending"),
            ("Request", "Submit"),
        ]
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "GlossaryTerm")
        self.assertEqual(sheet["A1"].value, "CDR ID")
        self.assertEqual(sheet["B1"].value, "Term Name")
        self.assertEqual(sheet["C1"].value, "Date Last Modified")
        self.assertEqual(sheet["D1"].value, "Publishable?")
        self.assertEqual(sheet["E1"].value, "Date First Published")
        self.assertEqual(sheet["F1"].value, "Last Comment")
        self.assertEqual(sheet["G1"].value, "Date Last Publishable")

    def test_gtn_qc_report(self):
        """Test the Glossary Term Name QC report."""

        term = self.get_test_glossary_term()
        term_id = term["id"]
        concept_id = term["concept_id"]
        self.navigate_to("QcReport.py", DocType="GlossaryTermName")
        self.assert_page_has("Glossary Term Name QC Report")
        self.assert_page_has("Title or Document ID")
        self.set_field_value("DocId", term_id)
        self.submit_form()
        self.assert_page_has("QC Report")
        self.assert_page_has(f"CDR{term_id}")
        self.assert_page_has(term["english_name"])
        for spanish_name in term["spanish_names"]:
            self.assert_page_has(spanish_name["name"])
        self.assert_page_has(f"CDR{concept_id}")

    def test_gtn_qc_with_concept_report(self):
        """Test the Glossary Term Name With Concept report.

        There's a scandalous amount of overlap with the "Glossary Full QC"
        report. Do we really need both?
        """

        term = self.get_test_glossary_term()
        term_id = term["id"]
        params = dict(DocType="GlossaryTermName", ReportType="gtnwc")
        self.navigate_to("QcReport.py", **params)
        self.assert_title("Glossary Term Name With Concept Report")
        self.assert_page_has("Title or Document ID")
        self.set_field_value("DocId", term_id)
        self.submit_form()
        self.assert_page_has("QC Report")
        self.assert_page_has(str(date.today()))
        self.assert_page_has(f"CDR{term_id}")
        self.assert_page_has("English - Patient")
        self.assert_page_has("Spanish - Patient")
        self.assert_page_has(term["english_name"])
        for spanish_name in term["spanish_names"]:
            self.assert_page_has(spanish_name["name"])
        self.assert_page_has("Definition")

    def test_hp_glossary_terms_report(self):
        """Test the Health Professional Glossary Terms report."""

        # The report has two versions; the default is for term names.
        form = self.navigate_to("HPGlossaryTermsReport.py")
        self.assert_title("Health Professional Glossary Terms Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "12/31/2020")
        self.submit_form()
        self.assert_title("Health Professional Glossary Terms Report")
        self.assert_page_has("English and Spanish HP Glossary Terms")
        self.assert_page_has("Genetics Dictionary")
        self.assert_page_has(date.today())
        self.assert_page_has("From 2020-01-01 to 2020-12-31")
        self.assert_regex(r"<th>Terms \(\d+\)</th>")

        # The other version is for term concepts.
        self.switch_to(form)
        self.select_values("type", "Concepts")
        self.click("opts-include-pronunciations")
        self.submit_form()
        self.assert_page_not_has("English and Spanish HP Glossary Terms")
        self.assert_page_has("English and Spanish HP Glossary Concepts")
        self.assert_page_has("<th>CDR ID of GTC</th>")
        self.assert_page_has("<th>Term Names (Pronunciations)</th>")
        self.assert_page_has("<th>Definition</th>")

    def test_keyword_search_report(self):
        """Test the Glossary Keyword Search report."""

        self.navigate_to("GlossaryKeywordSearchReport.py")
        self.assert_title("Glossary Keyword Search Report")
        self.select_values("language", "English")
        self.select_values("audience", "Patient")
        self.set_field_value("term", "healthy tissue")
        button = self.driver.find_element(By.CSS_SELECTOR, "#search-terms img")
        button.click()
        self.set_field_value("term-2", "doses of radiation")
        self.submit_form()
        self.assert_page_has(">GTN ID</th>")
        self.assert_page_has(">GTC ID</th>")
        self.assert_page_has(">Term Names</th>")
        self.assert_page_has(">Definitions</th>")
        self.assert_page_has("▶healthy tissue◀")
        self.assert_page_has("▶doses of radiation◀")

    def test_new_published_glossary_terms_report(self):
        """Test the New Publish Glossary Terms report."""

        script = "NewlyPublishedGlossaryTerms.py"
        self.navigate_to(script)
        self.assert_title("New Published Glossary Terms")
        params = dict(start="2020-01-01", end="2020-01-31", Request="Submit")
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "Sheet1")
        expected = r"\d+ Newly Published Glossary Term Documents"
        self.assertRegex(sheet["A1"].value, expected)
        self.assertEqual(sheet.max_column, 4)
        self.assertEqual(sheet["A3"].value, "CDR ID")
        self.assertEqual(sheet["B3"].value, "Term Name (English)")
        self.assertEqual(sheet["C3"].value, "Term Name (Spanish)")
        self.assertEqual(sheet["D3"].value, "Date First Published")

    def test_processing_status_report(self):
        """Test Glossary Processing Status report."""

        self.navigate_to("GlossaryProcessingStatusReport.py")
        self.assert_title("Glossary Processing Status Report")
        self.select_values("status", "Ready for translation")
        self.submit_form()
        self.assert_single_table_report()
        self.assert_wide_report()
        self.assert_page_has('<th colspan="3">Glossary Term Concept</th>')
        self.assert_page_has('<th colspan="3">Glossary Term Name</th>')
        self.assert_page_has("<th>CDR ID</th>")
        self.assert_page_has("<th>Processing Status</th>")
        self.assert_page_has("<th>Last Comment</th>")
        self.assert_page_has("<th>Term Names</th>")

    def test_pronunciation_by_term_stem_report(self):
        """Test the Pronunciation by Term Stem report."""

        self.navigate_to("PronunciationByWordStem.py")
        self.assert_title("Pronunciation by Term Stem Report")
        self.assert_page_has("Enter a term or pronunciation word stem")
        self.set_field_value("pron_stem", "ser-juh-ree")
        self.submit_form()
        self.assert_title("Pronunciation by Term Stem Report")
        self.assert_page_has("Pronunciation Stem: ser-juh-ree")
        self.assert_page_has("SER-juh-ree")
        self.assert_plain_report()
        self.assert_single_table_report()
        columns = (
            "Doc ID",
            "Term Name",
            "Pronunciation",
            "Pronunciation Resource",
            "Comments",
        )
        for column in columns:
            self.assert_regex(rf'<th style="min-width: \d+px">{column}</th>')

    def test_publish_preview_report(self):
        """Test publish preview of glossary terms."""

        params = dict(DocType="GlossaryTermName", ReportType="pp")
        form = self.navigate_to("QcReport.py", **params)
        self.assert_title("Publish Preview Report")
        term = self.get_test_glossary_term()
        self.set_field_value("DocId", term["id"])
        self.submit_form()
        self.assert_page_has(term["english_name"])
        for spanish_name in term["spanish_names"]:
            if not spanish_name["alternate"]:
                self.assert_page_has(spanish_name["name"])
        self.assert_page_has("LEFT NAV GOES HERE")

        # Confirm that the concept ID won't work here.
        self.switch_to(form)
        concept_id = term["concept_id"]
        self.set_field_value("DocId", concept_id)
        self.submit_form()
        expected = f"CDR{concept_id} is a GlossaryTermConcept document."
        self.assert_page_has(expected)

    def test_term_phrases_report(self):
        """Test the Glossary Term Phrases report.

        This is another report for which you'll need to check your
        email inbox to see the results.
        """

        self.navigate_to("GlossaryTermPhrases.py")
        self.assert_title("Glossary Term Phrases Report")
        warning = (
            "The report can take several minutes to prepare; "
            "please be patient."
        )
        self.assert_page_has(warning)
        term = self.get_test_glossary_term()
        term_id = term["id"]
        self.set_field_value("id", term_id)
        self.click("types-patientsummaries")
        self.click("language-spanish")
        self.submit_form()
        self.assert_title("Glossary Term Search")
        self.assert_page_has("Report Queued for Background Processing")
        self.driver.find_element(By.CSS_SELECTOR, "main p a").click()
        self.assert_title("Batch Job Status")
        self.assert_page_has("<th>ID</th>")
        self.assert_page_has("<th>Job Name</th>")
        self.assert_page_has("<th>Started</th>")
        self.assert_page_has("<th>Status</th>")
        self.assert_page_has("<th>Last Info</th>")
        self.assert_page_has("<th>Last Message</th>")
        self.assert_page_has("Glossary Term Search")
        self.assert_page_has(str(date.today()))
        self.assert_page_has("Queued")

    def test_translation_job_workflow_report(self):
        """Test the Glossary Translation Job Workflow report."""

        self.navigate_to("glossary-translation-job-report.py")
        self.assert_title("Translation Job Workflow Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.click("type-history")
        self.submit_form()
        self.assert_wide_report()
        self.assert_single_table_report()
        columns = (
            "CDR ID",
            "Title",
            "Status",
            "Status Date",
            "Assigned To",
            "Comment",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

    def test_translation_queue(self):
        """Test the Glossary translation queue pages.

        This test assumes that two test accounts exist, one named tester
        (full name "Regression Tester") and one named translation_tester
        (full name "Translation Tester"). Give both accounts email addresses
        to which you have access.
        """

        # Take care of some preparatory steps first.
        test_name = "Test GTN for translation queue"
        doc_id = self.create_test_gtn(name=test_name)
        for name in "tester", "translation_tester":
            self.add_user_to_group(name, "Spanish Glossary Translators")
        sql = "SELECT id, name FROM usr WHERE name LIKE '%tester%'"
        user_ids = {}
        for id, name in self.run_query(sql):
            user_ids[name] = id

        # Clear out any dross left from earlier failed tests.
        self.navigate_to("glossary-translation-jobs.py")
        self.assert_title("Glossary Translation Job Queue")
        self.assert_single_table_report()

        def find_job_row():
            table = self.load_table()
            for row in table.rows:
                if row[3].text.startswith(test_name):
                    return row
            return None
        obsolete_docs = []
        while (job_row := find_job_row()) is not None:
            link = job_row[1].find_element(By.TAG_NAME, "a")
            obsolete_docs.append(link.text)
            self.logger.debug("removing translation job for %s", link.text)
            link.click()
            self.click("submit-button-delete")
            alert = self.wait.until(expected_conditions.alert_is_present())
            self.logger.debug("alert text: %s", alert.text)
            alert.accept()
            sleep(1)
        for obsolete_doc in obsolete_docs:
            self.delete_doc(obsolete_doc)

        # Open the glossary translation job queue and create a new job.
        self.navigate_to("glossary-translation-jobs.py")
        comment = "job for automated test"
        self.click("submit-button-add")
        self.assert_title("Glossary Translation Job")
        self.assert_page_has("Create Translation Job")
        self.set_field_value("doc_id", doc_id)
        self.select_values("assigned_to", user_ids["tester"])
        states = {}
        for option in self.find("select#state option", all=True):
            states[option.text] = option.get_attribute("value")
        self.select_values("state", states["Ready For Translation"])
        self.set_field_value("comments", comment)
        self.submit_form(new_tab=False)

        # Check the queue page, make sure the job is there.
        self.assert_title("Glossary Translation Job Queue")
        table = self.load_table()
        columns = (
            "✓",
            "Doc ID",
            "Doc Type",
            "Doc Title",
            "Status",
            "Status Date",
            "Assigned To",
            "Comment",
        )
        table.check_headers(columns)
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[1].text, f"CDR{doc_id}")
        self.assertEqual(job_row[2].text, "GTN")
        self.assertEqual(job_row[3].text, test_name)
        self.assertEqual(job_row[4].text, "Ready For Translation")
        self.assertEqual(job_row[5].text, str(date.today()))
        self.assertEqual(job_row[6].text, "Regression Tester")
        self.assertEqual(job_row[7].text, comment)

        # Test reassignment of the job.
        user_id = user_ids["translation_tester"]
        self.click(f"assign_to-{user_id}")
        self.click(doc_id)
        self.click("submit-button-assign")
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[6].text, "Translation Tester")

        # Test editing the job.
        job_row[1].find_element(By.TAG_NAME, "a").click()
        self.select_values("state", states["Translation Made Publishable"])
        self.submit_form(new_tab=False)
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[4].text, "Translation Made Publishable")

        # Our job has reached the final state, so it should be purgeable.
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.click("submit-button-purge")
        job_row = find_job_row()
        self.assertIsNone(job_row)

        # Clean up after ourselves.
        for name in "tester", "translation_tester":
            self.remove_user_from_group(name, "Spanish Glossary Translators")
        self.delete_doc(doc_id)


class ManagementTests(Tester):
    """Tests of reports run by the Board Managers."""

    def test_board_invitation_history_report(self):
        """Test the PDQ Board Invitation History report."""

        # By default, the report has three columns, for the
        # board member's ID, the board member's name, and the
        # name of the board. With these condistions, the report
        # should fit inside the USWDS container grid.
        form_page = self.navigate_to("BoardInvitationHistory.py")
        self.submit_form()
        self.assert_tables_in_grid_container()
        self.assert_title("PDQ Board Invitation History Report")
        self.assert_page_has("PDQ Board Invitation History for All Boards")
        cols = self.driver.find_elements(By.CSS_SELECTOR, "colgroup col")
        self.assertEqual(len(cols), 3)
        rows = self.driver.find_elements(By.CSS_SELECTOR, "tbody tr")
        default_row_count = len(rows)
        self.assert_regex("(?s)ID.+Name.+Board Name")

        # Narrowing the report to a single board should significantly
        # reduce the number of rows in the report.
        board = self.get_test_board()
        board_id = board["id"]
        short_name = board["name"].replace("PDQ ", "")
        self.switch_to(form_page)
        self.click(f"board-{board_id}")
        self.submit_form()
        sleep(2)
        self.assert_title("PDQ Board Invitation History Report")
        self.assert_page_has(f"<span>{short_name}</span>")
        self.assert_page_not_has("PDQ Board Invitation History for All Boards")
        rows = self.driver.find_elements(By.CSS_SELECTOR, "tbody tr")
        single_board_row_count = len(rows)
        args = default_row_count, single_board_row_count
        self.logger.debug("all-boards count=%d, single-board count=%d", *args)
        self.assertLess(single_board_row_count * 4, default_row_count)

        # Adding one or more columns should to a non-USWDS report.
        self.switch_to(form_page)
        self.click("optional-invitation-date")
        self.submit_form()
        self.assert_plain_report()

    def test_board_meeting_dates_report(self):
        """Test report listing the board meetings for a given date range."""

        # The report should show meetings for a given date range.
        form_tab = self.navigate_to("BoardMeetingDates.py")
        self.assert_page_has("PDQ Editorial Board Meetings")
        self.set_field_value("start", "01/01/2010")
        self.set_field_value("end", "12/31/2010")
        self.submit_form()
        self.assert_page_has(f"Report generated {date.today()}")
        self.assert_page_has("(between 2010-01-01 and 2010-12-31)")
        self.assert_page_has("2010-01-19 11:00AM-1:00PM")
        self.assert_not_regex("(?s)2010.01.19.*Tuesday")
        self.assert_tables_in_grid_container()

        # The "By Date" version of the report should be wider. Note that the
        # date and time for each meeting has a different format than the "By
        # Board" version of the report.
        self.switch_to(form_tab, wait=10)
        self.click("report_type-display_by_date")
        self.submit_form()
        self.assert_wide_report()
        self.assert_page_has(f"Report generated {date.today()}")
        self.assert_page_not_has("2010-01-19 11:00AM-1:00PM")
        self.assert_regex("(?s)2010.01.19.*Tuesday")

    def test_board_member_qc_report(self):
        """Test the PDQ Board Member Information QC report."""

        member = self.get_test_board_member()
        member_id = member["id"]
        person = member["person"]
        surname = person["surname"]
        forename = person["forename"]
        initials = person["initials"]
        parts = [part for part in (forename, initials, surname) if part]
        member_name = " ".join(parts)
        self.navigate_to("QcReport.py", DocType="PDQBoardMemberInfo")
        self.assert_page_has("PDQ Board Member Information QC Report")
        self.set_field_value("DocTitle", surname[0])
        self.submit_form()
        self.assert_page_has(f"Multiple matches found for '{surname[0]}'.")
        self.click(f"docid-{member_id}")
        self.submit_form()
        self.assert_page_has(f"CDR{member_id}")
        self.assert_page_has(member_name)
        self.assert_plain_report()

    def test_board_members_and_topics_report(self):
        """Test the Board Members and Topics report."""

        # This report is non-tabular. The default version shows the summaries
        # (referred to as "topics" in the context of this report) in headings
        # with the board member reviewers in itemized lists. By default, only
        # summaries which can only be used as modules are excluded.
        form_page = self.navigate_to("PdqBoards.py")
        summary = self.get_test_summary(type="Genetics")
        summary_title = summary["title"]
        board_id = board_name = short_name = None
        for board in summary["boards"]:
            if "Editorial Board" in board["name"]:
                board_id = board["id"]
                board_name = board["name"]
                short_name = board_name.replace("PDQ ", "")
        if not board_id:
            summary_id = summary["id"]
            raise Exception(f"CDR{summary_id} has no editorial board")
        select = Select(self.driver.find_element(By.ID, "board"))
        select.select_by_visible_text(short_name)
        self.submit_form()
        h1 = f"h1>PDQ Board Report by Topic — {date.today()}</h1>"
        h3 = f"<h3>Topics for {board_name} (Health professionals)</h3>"
        self.assert_page_has(h1)
        self.assert_page_has(h3)
        tables = self.driver.find_elements(By.CSS_SELECTOR, "table")
        self.assertEqual(len(tables), 0)
        self.assert_page_has(f"<h4>{summary_title}</h4>")
        summaries = self.driver.find_elements(By.CSS_SELECTOR, "h4")
        self.logger.debug("PDQBoards.py found %d summaries", len(summaries))

        # Switching to "Modules Only" results in fewer "topics."
        self.switch_to(form_page)
        self.click("included-m")
        self.submit_form()
        module = self.get_test_summary(type="Genetics", module=True)
        module_title = module["title"]
        self.assert_page_not_has(f"<h4>{summary_title}</h4>")
        self.assert_page_has(f"<h4>{module_title} (module)</h4>")
        modules = self.driver.find_elements(By.CSS_SELECTOR, "h4")
        self.assertLess(len(modules), len(summaries))
        self.logger.debug("PDQBoards.py found %d modules", len(modules))

        # Switching to "Summaries and modules" should match the number
        # of topics for the previous two runs combined.
        self.switch_to(form_page)
        self.click("included-a")
        self.submit_form()
        self.assert_page_has(f"<h4>{summary_title}</h4>")
        self.assert_page_has(f"<h4>{module_title} (module)</h4>")
        topics = self.driver.find_elements(By.CSS_SELECTOR, "h4")
        self.logger.debug("PDQBoards.py found %d topics", len(topics))
        self.assertEqual(len(topics), len(modules) + len(summaries))

        # Switching to grouping by board member moves the "topics"
        # from the headers to the lists.
        self.switch_to(form_page)
        self.click("grouping-member")
        self.submit_form()
        self.assert_page_not_has(f"<h4>{summary_title}</h4>")
        self.assert_page_not_has(f"<h4>{module_title} (module)</h4>")
        self.assert_page_has(f"<li>{summary_title}</li>")
        self.assert_page_has(f"<li>{module_title} (module)</li>")

    def test_board_roster_full_reports(self):
        """Test the roster reports for complete sets of boards."""

        # The default version of the report uses custom markup created
        # by XSL/T filters, but inside the USWDS grid container.
        form_page = self.navigate_to("BoardRosterFull.py")
        self.submit_form()
        self.assert_tables_in_grid_container()
        uswds_tables = self.driver.find_elements(By.CSS_SELECTOR, ".usa-table")
        self.assertEqual(len(uswds_tables), 0)
        self.assert_title("PDQ Board Roster")
        boards = self.get_test_data("boards")
        board_names = defaultdict(list)
        long_name = "Integrative, Alternative, and Complementary Therapies"
        for board_type in "advisory", "editorial":
            for board in boards[board_type]:
                name = board["name"]
                if long_name in name:
                    name = name.replace(long_name, "IACT")
                board_names[board_type].append(name)
        for board_name in board_names["editorial"]:
            self.assert_page_has(f"<p>{board_name}</p>")

        # Try the same report for the editorial advisory boards.
        self.switch_to(form_page)
        self.click("board_type-advisory")
        self.submit_form()
        for board_name in board_names["advisory"]:
            self.assert_page_has(f"<p>{board_name}</p>")

        # Choosing "Group by PDQ Board" is very similar, but grouped
        # under headings identifying the board for that group.
        self.switch_to(form_page)
        self.click("board_type-editorial")
        self.click("grouping-by_board")
        self.submit_form()
        self.assert_tables_in_grid_container()
        for board_name in board_names["editorial"]:
            self.assert_page_not_has(f"<p>{board_name}</p>")
            self.assert_page_has(f"<h3>{board_name}</h3>")

        # The "Summary" version of the report uses standard tables, and
        # stays withing the USWDS container grid as long as the number of
        # columns is three or fewer (the default is two columns, one for
        # the board member's name and one for the board name). We're adding
        # one additional column here, for the board member's start date.
        # Grouping by member puts the entire report in a single table.
        self.switch_to(form_page)
        self.click("format-summary")
        self.click("grouping-by_member")
        self.click("column-start_date")
        self.submit_form()
        self.assert_tables_in_grid_container()
        uswds_tables = self.driver.find_elements(By.CSS_SELECTOR, ".usa-table")
        self.assertGreater(len(uswds_tables), 0)
        self.assert_page_has("PDQ Editorial Board Member Roster")
        self.assert_regex("(?s)Board Member.+Board Name.+Start Date")
        self.assert_single_table_report()

        # When we go over that threshold, even by just one more column, the
        # report switches back to a non-USWDS page, still with one table.
        self.switch_to(form_page)
        self.click("column-email")
        self.submit_form()
        self.assert_plain_report()
        self.assert_single_table_report()
        self.assert_regex("(?s)Board Member.+Board Name.+Email.+Start Date")

        # Group the summary version of the report results in multiple tables.
        self.switch_to(form_page)
        self.click("grouping-by_board")
        self.submit_form()
        self.assert_plain_report()
        self.assert_multi_table_report()
        self.assert_regex("(?s)Board Member.+Board Name.+Email.+Start Date")
        self.assert_page_not_has("PDQ Editorial Board Member Roster")
        for board_name in board_names["editorial"]:
            self.assert_page_has(f"{board_name} Roster")

    def test_board_roster_reports(self):
        """Test PDQ Board Roster reports."""

        # The default version of the report should not use the USWDS framework.
        form_page = self.navigate_to("BoardRoster.py")
        board = self.get_test_board()
        board_name = board["name"]
        select = Select(self.driver.find_element(By.ID, "board"))
        select.select_by_visible_text(board_name)
        self.click("option-show-all-contact-information")
        self.click("option-show-subgroup-information")
        self.click("option-show-assistant-information")
        self.submit_form()
        self.assert_plain_report()
        self.assert_page_has(f"<h1>{board_name}")
        self.assert_page_has("Editor-in-Chief")
        self.assert_page_has("<u>Assistant</u>")

        # The Summary version of the report should use the framework and fit
        # within the framework's grid container, as long as we stick with the
        # default columns.
        self.switch_to(form_page)
        self.click("type-summary")
        self.submit_form()
        self.assert_tables_in_grid_container()
        self.assert_regex("(?s)Name.+Phone.+Email")
        self.assert_not_regex("(?s)Name.+Phone.+Email.+Start Date")

        # But if we add more columns, the report should abandon the framework.
        self.switch_to(form_page)
        self.click("column-start-date")
        self.submit_form()
        self.assert_plain_report()
        self.assert_regex("(?s)Name.+Phone.+Email.+Start Date")

    def test_correspondence_mailer(self):
        """Test the PDQ Board Member Correspondence Mailer interface.

        The CDR prevents the creation of a publishing job if there is
        already another job of that type pending in the queue. This
        means that you cannot run this test multiple times in close
        succession. Check the publishing queue to make sure that any
        PDQ Board Member Correspondence Mailer job from a previous
        test run has already finished.
        """

        # As with the PCIB statistics report, we can't check in this test
        # what goes to our email box, but we can verify the confirmation
        # page (and you can look at the email inbox for further checks).
        self.navigate_to("BoardMemberMailerReqForm.py")
        self.assert_title("PDQ Board Member Correspondence Mailers")
        board_member = self.get_test_board_member()
        person = board_member["person"]
        surname = person["surname"]
        forename = person["forename"]
        member_name = f"{surname}, {forename}"
        member_id = board_member["id"]
        board = None
        for values in board_member["boards"]:
            if values["current"]:
                board = values
                break
        if not board:
            # This shouldn't happen, as we've asked the service to give
            # us board members with at least one active board membership.
            raise Exception("board member has no active board membership")
        board_id = board["id"]
        if "Advisory" in board["name"]:
            letter = "adv-thankyou"
        else:
            letter = "ed-welcome"
        self.click(f"board-{board_id}")

        # Wait for the AJAX to modify the available radio buttons on the rest
        # of the form. The Submit button should not show up until a board
        # member and a letter have both been selected.
        # self.driver.implicitly_wait(2)
        button = self.driver.find_element(By.ID, self.SUBMIT)
        self.assertFalse(button.is_displayed())
        self.click(letter)
        self.click(f"member-{member_id}")
        self.assertTrue(button.is_displayed())
        self.logger.debug("submitting mailer publishing job")
        self.submit_form()
        self.assert_title("PDQ Board Member Correspondence Mailers")
        self.assert_page_has("Queued 1 Mailer(s)")
        self.assert_page_has(f"<li>{member_name}</li>")
        self.click("submit-button-status-page")
        self.select_new_tab()
        self.assert_title("Publishing Status")
        self.assert_page_has("<td>PDQ Board Member Correspondence Mailer</td>")

    def test_pcib_stat_report(self):
        """Test the monthly PCIB statistics report.

        We can verify that the form is OK, and that the confirmation
        screen looks right, but we can't confirm what lands in the
        recipient's email box. You can always look at the output,
        though, to confirm that it's OK.
        """

        self.navigate_to("RunPCIBStatReport.py")
        self.assert_page_has("PCIB Statistics Report")
        self.set_field_value("start", "01/01/2023")
        self.set_field_value("end", "01/31/2023")
        self.submit_form()
        self.assert_page_has(f"Report generated {date.today()}")
        self.assert_page_has("The report has been sent to you by email.")

    def test_pdq_content_counts(self):
        """Test the PDQ Content Counts report."""

        self.navigate_to("PDQContentCounts.py", prompt="yes")
        self.assert_page_has("This report pulls together")
        self.submit_form()
        self.assert_page_has("PDQ Content Counts")
        self.assert_page_has("PDQ English HP summaries")


class MediaTests(Tester):
    """Tests of Media reports."""

    def test_advanced_search(self):
        """Test Media Advanced Search."""

        self.navigate_to("MediaSearch.py")
        self.assert_title("Media")
        doc = self.get_test_media_doc()
        doc_id = doc["id"]
        doc_title = doc["title"]
        self.set_field_value("title", f"{doc_title[0]}%")
        selector = 'main form input[type="submit"]'
        self.driver.find_element(By.CSS_SELECTOR, selector).click()
        self.select_new_tab()
        self.assert_page_has(f"documents match '{doc_title[0]}%'")
        qc_report_link = None
        cdr_id = f"CDR{doc_id:010d}"
        for link in self.driver.find_elements(By.CSS_SELECTOR, "td a"):
            if link.text == cdr_id:
                qc_report_link = link
                break
        self.assertIsNotNone(qc_report_link)
        qc_report_link.click()
        self.assert_regex(f"(?s)Media.+QC Report.+{date.today()}")
        self.assert_page_has(f"CDR{doc_id}")
        self.assert_page_has(doc_title.split(";")[0])

    def test_audio_request_report(self):
        """Test the Audio Request report."""

        self.navigate_to("AudioRequestSpreadsheet.py")
        self.assert_title("Audio Spreadsheet Creation")
        self.assert_page_has("Instructions")
        self.assert_page_has("Click Submit to request an Excel workbook ")
        self.submit_form(new_tab=False)
        self.assert_title("Audio Spreadsheet Creation")
        self.assert_page_has("Glossary Term Names Without Pronunciation")
        self.assert_page_has("Pronunciation files are needed")
        self.assert_page_has("Download the workbook")
        para = self.driver.find_element(By.CSS_SELECTOR, "main p")
        self.assertIsNotNone(para)
        pattern = r"Pronunciation files are needed for (\d+) glossary term "
        re_match = re_search(pattern, para.text)
        self.assertIsNotNone(re_match)
        count = int(re_match.group(1))
        link = para.find_element(By.CSS_SELECTOR, "a")
        self.assertIsNotNone(link)
        book = self.fetch_workbook(link.get_attribute("href"))
        sheet = book.active
        self.assertEqual(sheet.title, "Term Names")
        self.assertEqual(sheet.max_column, 8)
        self.assertEqual(sheet["A1"].value, "CDR ID")
        self.assertEqual(sheet["B1"].value, "Term Name")
        self.assertEqual(sheet["C1"].value, "Language")
        self.assertEqual(sheet["D1"].value, "Pronunciation")
        self.assertEqual(sheet["E1"].value, "Filename")
        self.assertEqual(sheet["F1"].value, "Notes (Vanessa)")
        self.assertEqual(sheet["G1"].value, "Notes (NCI)")
        self.assertEqual(sheet["H1"].value, "Reuse Media ID")
        self.assertEqual(sheet.max_row, count+1)

    def test_audio_review_scripts(self):
        """Test the scripts used for reviewing MP3 pronunciation files."""

        # Start with a clean slate.
        self.navigate_to("remove-audio-test-data.py")
        self.assert_title("Remove Audio Pronunciation Test Data")
        self.submit_form(new_tab=False)
        self.assert_page_not_has("usa-alert--error")

        # Generate the test audio files we'll need.
        self.navigate_to("make-audio-test-data.py")
        self.assert_title("Create Audio Pronunciation Test Data")
        self.submit_form(new_tab=False)
        self.assert_title("Create Audio Pronunciation Test Data")
        tier = "qa" if "qa" in self.HOST else "dev"
        audio_directory = f"/sftp/sftphome/cdrstaging/ciat/{tier}/Audio"
        source_directory = f"{audio_directory}/Term_Audio"
        path = f"{source_directory}/Week_2099_01.zip"
        self.assert_page_has(f"Successfully stored {path}.")
        path = f"{source_directory}/Week_2099_01_Rev1.zip"
        self.assert_page_has(f"Successfully stored {path}.")
        pattern = r"Created test GlossaryTermName document CDR0*(\d+)\."
        re_match = re_search(pattern, self.get_page_source())
        self.assertIsNotNone(re_match)
        doc_id = int(re_match.group(1))

        # Download the test set from the s/FTP server to the CDR server.
        # Use test mode first.
        form = self.navigate_to("FtpAudio.py")
        self.assert_title("Retrieve Audio Files From Secure FTP Server")
        self.click("options-test")
        self.submit_form()
        self.assert_title("Retrieve Audio Files From Secure FTP Server")
        self.assert_page_has("Processing mode: test")

        # Next do the download for real.
        destination_directory = "D:/cdr/Audio_from_CIPSFTP"
        transferred_directory = f"{audio_directory}/Audio_Transferred"
        self.switch_to(form)
        self.click("options-test")
        self.submit_form()
        self.assert_title("Retrieve Audio Files From Secure FTP Server")
        self.assert_page_has("Processing mode: live")
        self.assert_page_has(f"Source directory: {source_directory}")
        self.assert_page_has(f"Destination directory: {destination_directory}")
        self.assert_page_has(f"Transferred directory: {transferred_directory}")
        self.assert_page_has("Retrieved Week_2099_01.zip")
        self.assert_page_has("Retrieved Week_2099_01_Rev1.zip")
        self.assert_page_has("Moved Week_2099_01.zip to Transferred directory")
        self.assert_page_has("Moved Week_2099_01_Rev1.zip to Transferred")

        # Bring up the review landing page.
        self.navigate_to("GlossaryTermAudioReview.py")
        self.assert_title("Glossary Term Audio Review")
        self.assert_page_has("Click a link to a zip file to review")
        self.assert_page_has("Audio Zip Files")
        table = self.load_table()
        columns = "File name", "Review status", "Date modified"
        table.check_headers(columns)

        # Find one of the archives we just created.
        def find_test_archive(driver, name):
            logger = self.logger

            class Archive:

                def __init__(self, row):
                    self.row = row

                @cached_property
                def cells(self):
                    return list(self.row.find_elements(By.TAG_NAME, "td"))

                @cached_property
                def link(self):
                    driver.implicitly_wait(0)
                    nodes = self.cells[0].find_elements(By.TAG_NAME, "a")
                    logger.debug("found %d nodes", len(nodes))
                    driver.implicitly_wait(Tester.DEFAULT_WAIT)
                    return nodes[0] if nodes else None

                @cached_property
                def modified(self):
                    return self.cells[2].text

                @cached_property
                def name(self):
                    return self.cells[0].text

                @cached_property
                def status(self):
                    return self.cells[1].text

            for row in driver.find_elements(By.CSS_SELECTOR, "tbody tr"):
                archive = Archive(row)
                if archive.name == name:
                    return archive
            return None

        archive = find_test_archive(self.driver, "Week_2099_01.zip")
        self.assertIsNotNone(archive)
        self.assertIsNotNone(archive.link)
        self.assertEqual(archive.status, "Unreviewed")
        self.assertTrue(archive.modified.startswith(str(date.today())))

        # Open the archive for review.
        archive.link.click()
        self.select_new_tab()
        self.assert_title("Glossary Term Audio Review")
        self.assert_page_has("Click 'Save' to save any changes")
        columns = (
            "Disposition",
            "CDR ID",
            "Term name",
            "Lang",
            "Pronunciation",
            "MP3 file",
            "Reader note",
            "Reviewer note",
        )
        table = self.load_table()
        table.check_headers(columns)

        # Collect the information for each MP3 row.
        class MP3Row:

            def __init__(self, row):
                self.row = row

            @cached_property
            def buttons(self):
                nodes = self.cells[0].find_elements(By.CSS_SELECTOR, "input")
                keys = "approved", "rejected", "unreviewed"
                buttons = {}
                for i, node in enumerate(nodes):
                    buttons[keys[i]] = node
                return buttons

            @cached_property
            def cells(self):
                return list(self.row.find_elements(By.TAG_NAME, "td"))

            @cached_property
            def doc_id(self):
                return self.cells[1].text

            @cached_property
            def id(self):
                name = self.buttons["approved"].get_attribute("name")
                return int(name.split("-")[1])

            @cached_property
            def language(self):
                return self.cells[3].text

            @cached_property
            def name(self):
                return self.cells[2].text

            @cached_property
            def path(self):
                return self.cells[5].find_element(By.TAG_NAME, "a").text

            @cached_property
            def pronunciation(self):
                return self.cells[4].text

            @cached_property
            def reader_note(self):
                return self.cells[6].text

            @cached_property
            def reviewer_note(self):
                return self.cells[7].find_element(By.TAG_NAME, "textarea").text
        rows = []
        for row in self.driver.find_elements(By.CSS_SELECTOR, "tbody tr"):
            rows.append(MP3Row(row))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].name, "sample term for unit tests")
        self.assertEqual(rows[0].language, "English")
        self.assertEqual(rows[0].pronunciation, "test SAM-pul term")
        self.assertEqual(rows[0].path, f"Week_2099_01/{doc_id}_en.mp3")
        self.assertEqual(rows[0].reader_note, "")
        self.assertEqual(rows[1].name, "término de prueba de muestra")
        self.assertEqual(rows[1].language, "Spanish")
        self.assertEqual(rows[1].pronunciation, "")
        self.assertEqual(rows[1].path, f"Week_2099_01/{doc_id}_es.mp3")
        self.assertEqual(rows[1].reader_note, "I think this is right")

        # Perform and save the reviews.
        self.click(f"status-{rows[0].id}-a")
        self.click(f"status-{rows[1].id}-r")
        self.driver.find_element(By.ID, "submit-button-save").click()
        self.assert_title("Glossary Term Audio Review")
        self.assert_page_has("Audio Set Week_2099_01.zip Review Complete")
        archive = find_test_archive(self.driver, "Week_2099_01.zip")
        self.assertIsNotNone(archive)
        self.assertIsNone(archive.link)
        self.assertEqual(archive.status, "Completed")
        self.assertTrue(archive.modified.startswith(str(date.today())))
        self.assert_page_has("You can retrieve")

        # Open up the followup set for review.
        archive = find_test_archive(self.driver, "Week_2099_01_Rev1.zip")
        self.assertIsNotNone(archive)
        self.assertIsNotNone(archive.link)
        self.assertEqual(archive.status, "Unreviewed")
        self.assertTrue(archive.modified.startswith(str(date.today())))
        archive.link.click()
        self.select_new_tab()
        table = self.load_table()
        rows = []
        for row in self.driver.find_elements(By.CSS_SELECTOR, "tbody tr"):
            rows.append(MP3Row(row))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].name, "término de prueba de muestra")
        self.assertEqual(rows[0].language, "Spanish")
        self.assertEqual(rows[0].pronunciation, "")
        self.assertEqual(rows[0].path, f"Week_2099_01_Rev1/{doc_id}_es.mp3")
        self.assertEqual(rows[0].reader_note, "Second time's the charm!")

        # Perform and save the review.
        self.click(f"status-{rows[0].id}-a")
        self.driver.find_element(By.ID, "submit-button-save").click()
        self.assert_title("Glossary Term Audio Review")
        self.assert_page_has("Audio Set Week_2099_01_Rev1.zip Review Complete")
        archive = find_test_archive(self.driver, "Week_2099_01_Rev1.zip")
        self.assertIsNotNone(archive)
        self.assertIsNone(archive.link)
        self.assertEqual(archive.status, "Completed")
        self.assertTrue(archive.modified.startswith(str(date.today())))
        self.assert_page_not_has("You can retrieve")
        self.assert_page_has("None of the files in the set were rejected")
        self.assert_page_has(", so there is no new workbook for a subsequent")

        # Import the media documents and link them to the glossary term doc.
        self.navigate_to("LoadGlossaryAudioFiles.py")
        self.assert_title("Load Glossary Audio Files")
        self.assert_page_has("Compressed Archives containing Audio files")
        self.assert_page_has("Week_2099_01.zip")
        self.assert_page_has("Week_2099_01_Rev1.zip")
        self.submit_form()
        self.assert_title("Load Glossary Audio Files")
        self.assert_tables_in_grid_container()
        self.assert_single_table_report()
        table = self.load_table()
        table.check_headers(["CDR ID", "Processing"])
        media_en = table.rows[0][0].text
        media_es = table.rows[1][0].text
        template = "Adding link from this document to Media document {}"
        expected = (
            f"created Media doc for CDR{doc_id} ('sample term for unit "
            "tests' [en]) from Week_2099_01.zip"
        )
        self.assertEqual(table.rows[0][1].text, expected)
        expected = (
            f"created Media doc for CDR{doc_id} ('término de prueba de "
            "muestra' [es]) from Week_2099_01_Rev1.zip"
        )
        self.assertEqual(table.rows[1][1].text, expected)
        self.assertEqual(table.rows[2][0].text, f"CDR{doc_id}")
        self.assertEqual(table.rows[2][1].text, template.format(media_en))
        self.assertEqual(table.rows[3][0].text, f"CDR{doc_id}")
        self.assertEqual(table.rows[3][1].text, template.format(media_es))

        # Clean up after ourselves.
        self.navigate_to("remove-audio-test-data.py")
        self.assert_title("Remove Audio Pronunciation Test Data")
        self.submit_form(new_tab=False)
        alerts = (
           f"Removed remote /sftp/sftphome/cdrstaging/ciat/{tier}/Audio"
           "/Audio_Transferred/Week_2099_01.zip.",
           f"Removed remote /sftp/sftphome/cdrstaging/ciat/{tier}/Audio"
           "/Audio_Transferred/Week_2099_01_Rev1.zip.",
           r"Removed local D:\cdr\Audio_from_CIPSFTP\Week_2099_01.zip.",
           r"Removed local D:\cdr\Audio_from_CIPSFTP\Week_2099_01_Rev1.zip.",
           "Removed 3 rows from the term_audio_mp3 table.",
           "Removed 2 rows from the term_audio_zipfile table.",
           f"Deleted GlossaryTermName document CDR{doc_id:010d}.",
           f"Deleted Media document CDR{int(media_en[3:]):010d}.",
           f"Deleted Media document CDR{int(media_es[3:]):010d}.",
        )
        for alert in alerts:
            self.assert_page_has(alert)
        self.assert_page_not_has("usa-alert--error")

    def test_caption_and_content_report(self):
        """Test the Media and Content report."""

        self.navigate_to("MediaCaptionContent.py")
        self.assert_title("Media Caption and Content Report")
        params = dict(
            start_date="2020-01-01",
            end_date="2020-01-31",
            Request="Submit",
        )
        book = self.fetch_workbook("MediaCaptionContent.py", params)
        sheet = book.active
        self.assertEqual(sheet.title, "Media Caption-Content")
        self.assertEqual(sheet["A1"].value, "Media Caption and Content Report")
        self.assertEqual(sheet["A2"].value, "2020-01-01 -- 2020-01-31")
        self.assertEqual(sheet["A3"].value, "CDR ID")
        self.assertEqual(sheet["B3"].value, "Title")
        self.assertEqual(sheet["C3"].value, "Diagnosis")
        self.assertEqual(sheet["D3"].value, "Proposed Summaries")
        self.assertEqual(sheet["E3"].value, "Proposed Glossary Terms")
        self.assertEqual(sheet["F3"].value, "Label Names")
        self.assertEqual(sheet["G3"].value, "Content Description")
        self.assertEqual(sheet["H3"].value, "Caption")

    def test_image_demographic_information_report(self):
        """Test the Image Demographic Information report."""

        # Test the image version of the report.
        form = self.navigate_to("ImageDemographicInformationReport.py")
        self.assert_title("Image Demographic Information")
        self.click("image_method-id")
        image = self.get_test_media_doc()
        image_id = image["id"]
        self.set_field_value("image-id", image_id)
        self.submit_form()
        self.assert_title("Image Demographic Information")
        self.assert_page_has("Report Type: Images")
        self.assert_page_has(f"Image CDR ID: {image_id}")
        self.assert_page_has("Number of rows: 1")
        for thing in "ID", "Title":
            for language in "English", "Spanish":
                self.assert_page_has(f"<th>{language} Image {thing}</th>")
        for c in "Age", "Sex", "Race", "Skin Tone", "Ethnicity", "Image Link":
            for language in "en", "es":
                self.assert_page_has(f"<th>{c} ({language})</th>")
        link = self.driver.find_element(By.CSS_SELECTOR, "td a")
        if link.location_once_scrolled_into_view:
            link.click()
        self.select_new_tab()
        self.assert_regex(rf"(?s)Media.+QC Report.+{date.today()}")
        self.assert_page_has(f"CDR{image_id}")
        self.assert_page_has(image["title"].split(";")[0])

        # Test the summary version of the report.
        self.switch_to(form)
        self.click("type-summaries")
        self.click("summary_method-id")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        self.set_field_value("summary-id", summary_id)
        self.submit_form()
        self.assert_title("Image Demographic Information")
        self.assert_page_has("Report Type: Summaries")
        self.assert_page_has("Summary Options: exclude summary modules")
        for thing in "Summary ID", "Summary Title", "Image ID":
            for language in "English", "Spanish":
                self.assert_page_has(f"<th>{language} {thing}</th>")
        for c in "Age", "Sex", "Race", "Skin Tone", "Ethnicity", "Image Link":
            for language in "en", "es":
                self.assert_page_has(f"<th>{c} ({language})</th>")
        link = self.driver.find_element(By.CSS_SELECTOR, "td a")
        if link.location_once_scrolled_into_view:
            link.click()
        self.select_new_tab()
        self.assert_regex(rf"(?s)Media.+QC Report.+{date.today()}")

    def test_images_processing_status_report(self):
        """Test the Media (Images) Processing Status report."""

        self.navigate_to("ImageMediaProcessingStatusReport.py")
        self.assert_title("Media (Images) Processing Status Report")
        self.select_values("status", "Processing Complete")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "12/31/2020")
        select = Select(self.driver.find_element(By.ID, "diagnosis"))
        select.select_by_visible_text("skin cancer")
        self.submit_form()
        self.assert_title("Media (Images) Processing Status Report")
        self.assert_plain_report()
        self.assert_page_has("From 2020-01-01 - 2020-12-31")
        self.assert_page_has("Status: Processing Complete")
        self.assert_page_has(">CDR ID</th>")
        self.assert_page_has(">Media Title</th>")
        self.assert_page_has(">Diagnosis</th>")
        self.assert_page_has(">Processing Status</th>")
        self.assert_page_has(">Processing Status Date</th>")
        self.assert_page_has(">Proposed Summaries</th>")
        self.assert_page_has(">Proposed Glossary Terms</th>")
        self.assert_page_has(">Comments</th>")
        self.assert_page_has(">Last Version Publishable</th>")
        self.assert_page_has(">Published</th>")

    def test_images_report(self):
        """Test the Media Images report."""

        doc = self.get_test_media_doc()
        doc_id = doc["id"]
        doc_title = doc["title"].split(";")[0].strip()
        self.navigate_to("MediaLanguageCompare.py")
        self.assert_title("Media Images Report")
        self.click("selection_method-id")
        self.set_field_value("cdr-id", doc_id)
        self.submit_form()
        self.assert_page_has('<body class="double">')
        self.assert_page_has("Media Images Report")
        self.assert_page_has("Language Comparison")
        self.assert_page_has(date.today())
        self.assert_page_has(f"Media selected: CDR{doc_id}")
        self.assert_page_has(f"CDR{doc_id} - {doc_title}")

    def test_keyword_search_report(self):
        """Test the Media Keyword Search report."""

        self.navigate_to("MediaKeywordSearchReport.py")
        self.assert_title("Media Keyword Search Report")
        self.click("language-english")
        self.set_field_value("term", "cardiac")
        button = self.driver.find_element(By.CSS_SELECTOR, "#search-terms img")
        button.click()
        self.set_field_value("term-2", "monitor")
        self.submit_form()
        self.assert_title("Media Keyword Search Report")
        self.assert_page_has(">CDR ID</th>")
        self.assert_page_has(">Title</th>")
        self.assert_page_has(">Terms</th>")
        self.assert_page_has("▶cardiac◀")
        self.assert_page_has("monitor◀")

    def test_linked_media_documents_report(self):
        """Test the Linked Media Documents report."""

        self.navigate_to("MediaLinks.py")
        self.assert_title("Documents that Link to Media Documents")
        self.submit_form()
        self.assert_title("Documents that Link to Media Documents")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        self.assert_regex(
            r"<caption>[^<]*<span>Glossary Terms \(\d+\)</span>[^<]*</caption>"
        )
        self.assert_regex(
            r"<tr>[^<]*<th>CDR ID</th>[^<]*<th>Document Title</th>[^<]*</tr>"
        )

    def test_lists_report(self):
        """Test the Media Lists report."""

        self.navigate_to("MediaLists.py")
        self.assert_title("Media Lists")
        self.submit_form()
        self.assert_title("Media Lists")
        self.assert_regex(r"\d+ Media Documents")
        self.assert_page_has("Report Filtering")
        self.assert_page_has("Diagnosis: All Diagnoses")
        self.assert_page_has("Condition: All Categories")
        self.assert_page_has("Media Type: Any")
        self.assert_regex("<th[^>]*>Doc Title</th>")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()

    def test_media_in_summary_report(self):
        """Test the Media in Summary report."""

        self.navigate_to("MediaInSummary.py")
        self.assert_title("Media in Summary Report")
        self.click("selection_method-id")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.set_field_value("cdr-id", summary_id)
        self.submit_form()
        self.assert_page_has("<div>Media in Summary Report</div>")
        self.assert_page_has("<div>Side-by-Side</div>")
        self.assert_page_has(f"<div>{date.today()}</div>")
        self.assert_page_has(f"Summary selected: CDR{summary_id}")
        self.assert_page_has(summary_title)
        self.assert_plain_report()

    def test_meeting_recording_tracking_report(self):
        """Test the Board Meeting Recordings Tracking report."""

        script = "RecordingTrackingReport.py"
        title = "Board Meeting Recordings Tracking Report"
        start = "2015-01-01"
        end = "2015-12-31"
        self.navigate_to(script)
        self.assert_title(title)
        self.assert_page_has("Select Date Range")
        params = dict(start=start, end=end, Request="Submit")
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "Board Meeting Recordings")
        self.assertEqual(sheet["A1"].value, title)
        self.assertEqual(sheet["A2"].value, f"Between {start} and {end}")
        self.assertEqual(sheet["A4"].value, "CDRID")
        self.assertEqual(sheet["B4"].value, "Media Title")
        self.assertEqual(sheet["C4"].value, "Encoding")
        self.assertEqual(sheet["D4"].value, "Date Created")
        self.assertEqual(sheet["E4"].value, "Last Version Publishable")
        self.assertEqual(sheet["F4"].value, "Version Date")
        self.assertEqual(sheet["G4"].value, "Comments")

    def test_permissions_report(self):
        """Test the Media Permissions report."""

        self.navigate_to("MediaPermissionsReport.py")
        self.assert_title("Media Permissions Report")
        self.set_field_value("req_start", "1/1/2015")
        self.set_field_value("req_end", "12/31/2015")
        self.submit_form()
        self.assert_title("Media Permissions Report")
        self.assert_plain_report()
        self.assert_page_has("English and Spanish Permission Requests")
        columns = (
            "Media DocTitle",
            "Permission Request Date",
            r"Permission Response \(Response Date\)",
            "Expiration",
            r"Spanish Permission Requested \(Permission Response\)",
            "Approved Use",
            "Comment",
        )
        for column in columns:
            self.assert_regex(f"<th[^>]*>{column}</th>")

    def test_pronunciation_recording_tracking_report(self):
        """Test the Pronunciation Recordings Tracking report.

        Another queued batch report for which a notification with a link
        will be emailed to the mailbox of the account for the session
        used to run the tests.
        """

        self.navigate_to("PronunciationRecordings.py")
        self.assert_title("Pronunciation Recordings Tracking Report")
        self.submit_form()
        self.assert_title("Audio Pronunciation Recordings Tracking Report")
        self.assert_page_has("Report Queued for Background Processing")
        self.driver.find_element(By.CSS_SELECTOR, "main p a").click()
        self.assert_title("Batch Job Status")
        self.assert_page_has("Batch Jobs")
        self.assert_page_has("<th>ID</th>")
        self.assert_page_has("<th>Job Name</th>")
        self.assert_page_has("<th>Started</th>")
        self.assert_page_has("<th>Status</th>")
        self.assert_page_has("<th>Last Info</th>")
        self.assert_page_has("<th>Last Message</th>")
        self.assert_page_has("Audio Pronunciation Recordings Tracking Report")
        self.assert_page_has(str(date.today()))
        self.assert_regex("Queued|In process")

    def test_pronunciation_review_statistics_report(self):
        """Test the Glossary Term Audio Review statistical report."""

        self.navigate_to("GlossaryTermAudioReviewReport.py")
        self.assert_title("Glossary Term Audio Review Statistical Report")
        self.set_field_value("start_date", "1/1/2020")
        self.set_field_value("end_date", "12/31/2020")
        self.submit_form()
        self.assert_title("Approved English Audio Pronunciations")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        self.assert_page_has("Status Totals")
        self.assert_page_has("<th>Approved</th>")
        self.assert_page_has("<th>Rejected</th>")
        self.assert_page_has("<th>Unreviewed</th>")
        self.assert_page_has("<th>Term</th>")
        self.assert_page_has("<th>Review Date</th>")
        self.assert_page_has("<th>User</th>")

    def test_publishing_report(self):
        """Test the Media Document Publishing Report"""

        self.navigate_to("PublishedMediaDocuments.py")
        self.assert_title("Media Doc Publishing Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "12/31/2020")
        self.submit_form()
        self.assert_title("Media Doc Publishing Report")
        self.assert_wide_report()
        self.assert_page_has(
            "Media Documents Published Between 2020-01-01 and 2020-12-31"
        )
        self.assert_page_has("<th>CDR ID</th>")
        self.assert_page_has("<th>Media Title</th>")
        self.assert_page_has("<th>First Pub Date</th>")
        self.assert_page_has("<th>Version Date</th>")
        self.assert_page_has("<th>Last Version Publishable</th>")
        self.assert_page_has("<th>Blocked from VOL</th>")

    def test_qc_report(self):
        """Test the Media QC report."""

        self.navigate_to("QcReport.py", DocType="Media", ReportType="img")
        self.assert_title("Media QC Report")
        doc = self.get_test_media_doc()
        doc_id = doc["id"]
        self.set_field_value("DocId", doc_id)
        self.submit_form()
        self.assert_regex(rf"(?s)Media.+QC Report.+{date.today()}")
        self.assert_page_has(f"CDR{doc_id}")
        self.assert_page_has(doc["title"].split(";")[0])
        self.assertIsNotNone(self.driver.find_element(By.CSS_SELECTOR, "img"))

    def test_translation_job_workflow_report(self):
        """Test the Media Translation Job Workflow report."""

        self.navigate_to("media-translation-job-report.py")
        self.assert_title("Translation Job Workflow Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.click("type-history")
        self.submit_form()
        self.assert_title("Translation Job History Report")
        self.assert_wide_report()
        self.assert_single_table_report()
        columns = (
            "CDR ID EN",
            "Title EN",
            "CDR ID ES",
            "Status",
            "Status Date",
            "Assigned To",
            "Comment",
        )
        for column in columns:
            self.assert_regex(f"<th[^>]*>{column}</th>")

    def test_translation_queue(self):
        """Test the Media translation queue pages.

        This test assumes that two test accounts exist, one named tester
        (full name "Regression Tester") and one named translation_tester
        (full name "Translation Tester"). Give both accounts email addresses
        to which you have access.
        """

        # Take care of some preparatory steps first.
        test_title = "Test Media doc for translation queue"
        doc_id = self.create_test_media_doc(title=test_title)
        for name in "tester", "translation_tester":
            self.add_user_to_group(name, "Spanish Media Translators")
        sql = "SELECT id, name FROM usr WHERE name LIKE '%tester%'"
        user_ids = {}
        for id, name in self.run_query(sql):
            user_ids[name] = id

        # Clear out any dross left from earlier failed tests.
        self.navigate_to("media-translation-jobs.py")
        self.assert_title("Media Translation Job Queue")
        self.assert_single_table_report()

        def find_job_row():
            table = self.load_table()
            for row in table.rows:
                if row[2].text == test_title:
                    return row
            return None
        obsolete_docs = []
        while (job_row := find_job_row()) is not None:
            link = job_row[1].find_element(By.TAG_NAME, "a")
            obsolete_docs.append(link.text)
            self.logger.debug("removing translation job for %s", link.text)
            link.click()
            self.click("submit-button-delete")
            alert = self.wait.until(expected_conditions.alert_is_present())
            self.logger.debug("alert text: %s", alert.text)
            alert.accept()
            sleep(1)
        for obsolete_doc in obsolete_docs:
            self.delete_doc(obsolete_doc)

        # Open the glossary translation job queue and create a new job.
        self.navigate_to("media-translation-jobs.py")
        comment = "job for automated test"
        self.click("submit-button-add")
        self.assert_title("Media Translation Job")
        self.assert_page_has("Create Translation Job")
        self.set_field_value("english_id", doc_id)
        self.select_values("assigned_to", user_ids["tester"])
        states = {}
        for option in self.find("select#state option", all=True):
            states[option.text] = option.get_attribute("value")
        self.select_values("state", states["Ready For Translation"])
        self.set_field_value("comments", comment)
        self.submit_form(new_tab=False)

        # Check the queue page, make sure the job is there.
        self.assert_title("Media Translation Job Queue")
        table = self.load_table()
        self.assertEqual(table.caption.text, "Jobs")
        columns = (
            "✓",
            "Doc ID EN",
            "Title EN",
            "Doc ID ES",
            "Status",
            "Status Date",
            "Assigned To",
            "Comment",
        )
        table.check_headers(columns)
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[1].text, f"CDR{doc_id}")
        self.assertEqual(job_row[2].text, test_title)
        self.assertEqual(job_row[3].text, "")
        self.assertEqual(job_row[4].text, "Ready For Translation")
        self.assertEqual(job_row[5].text, str(date.today()))
        self.assertEqual(job_row[6].text, "Regression Tester")
        self.assertEqual(job_row[7].text, comment)

        # Test reassignment of the job.
        user_id = user_ids["translation_tester"]
        self.click(f"assign_to-{user_id}")
        self.click(doc_id)
        self.click("submit-button-assign")
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[6].text, "Translation Tester")

        # Test editing the job.
        job_row[1].find_element(By.TAG_NAME, "a").click()
        self.select_values("state", states["Translation Made Publishable"])
        self.submit_form(new_tab=False)
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[4].text, "Translation Made Publishable")

        # Our job has reached the final state, so it should be purgeable.
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.click("submit-button-purge")
        job_row = find_job_row()
        self.assertIsNone(job_row)

        # Clean up after ourselves.
        for name in "tester", "translation_tester":
            self.remove_user_from_group(name, "Spanish Glossary Translators")
        self.delete_doc(doc_id)


class PublishingTests(Tester):
    """Tests of the CDR publishing system."""

    def test_count_by_doctype_report(self):
        """Test the Publishing Count By Doctype report."""

        self.navigate_to("CountByDoctype.py")
        self.assert_title("Publishing Count By Doctype")
        self.assert_single_table_report()
        table = self.load_table()
        table.check_headers(("Document Type", "Count"))
        self.assertEqual(len(table.rows), 6)
        labels = (
            "DrugInformationSummary",
            "GlossaryTermName",
            "Media",
            "Summary",
            "Term",
            "TOTAL",
        )
        regex = r"^Exported By The Latest Weekly Export Job \(#\d+\)$"
        self.assertRegex(table.caption.text, regex)
        for i, label in enumerate(labels):
            self.assertEqual(table.rows[i][0].text, label)
            self.assertRegex(table.rows[i][1].text, r"^\d+$")

    def test_filtering(self):
        """Test the filtering interface."""

        filter_set = "QC Summary Set (Bold/Underline)"
        doc = self.get_test_summary()
        id = doc["id"]
        title = doc["title"]
        form = self.navigate_to("Filter.py")
        self.assert_title("CDR Document Filtering")
        self.set_field_value("DocId", id)
        self.set_field_value("filter-1", f"set:{filter_set}")
        self.click("rsmarkup")
        self.click("submit-button-qc-filter-sets")
        self.select_new_tab()
        self.assert_title(f"Filter Sets for Filtering of Summary Doc CDR{id}")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        table.check_headers(("Set Name", "Action", "Set Detail"))
        self.assertEqual(table.rows[0][0].text, filter_set)
        self.switch_to(form)
        self.click("submit-button-submit-filter-request")
        self.select_new_tab()
        self.assert_page_has(f"<h2>{title}</h2>")
        self.assert_page_has("<h3>Treatment - Health professionals</h3>")

    def test_filters(self):
        """Test the interface for reviewing the publishing filters."""

        # The page has two tables, one for each sort option.
        # Only one is visible at a time, starting with the title sort.
        filters_page = self.navigate_to("EditFilters.py")
        self.assert_title("Manage Filters")
        tables = self.load_tables()
        self.logger.debug("%d tables", len(tables))
        tables[0].check_headers(("CDR ID", "Filter Title"))
        tables[1].check_headers(("", ""))
        self.assertIn("Sorted By Title", tables[0].caption.text)
        self.assertEqual("", tables[1].caption.text)

        # Switch to the sort by document ID.
        tables[0].headers[0].click()
        tables[0].check_headers(("", ""))
        tables[1].check_headers(("CDR ID", "Filter Title"))
        self.assertEqual("", tables[0].caption.text)
        self.assertIn("Sorted By CDR ID", tables[1].caption.text)

        # Test the tier comparison button.
        self.click("submit-button-compare-with-prod")
        self.select_new_tab()
        identical = "Filters on (DEV)|(QA)|(STAGE) and PROD are identical"
        no_diffs = re_search(identical, self.get_page_source()) is not None
        only_on = "Only on" in self.get_page_source()
        gone_from_prod = "--- PROD" in self.get_page_source()
        self.assertTrue(no_diffs or only_on or gone_from_prod)

        # Bring up the catalog of global parameters used in the filters.
        # The table has three columns, but the cells in the first column
        # are header elements (th), not data elements (td).
        self.switch_to(filters_page)
        self.click("submit-button-filter-params")
        self.select_new_tab()
        self.assert_single_table_report()
        self.assert_plain_report()
        table = self.load_table()
        self.assertEqual(table.caption.text, "Global Filter Parameters")
        row = table.node.find_element(By.TAG_NAME, "tr")
        cells = row.find_elements(By.XPATH, "*")
        self.assertEqual(len(cells), 3)
        self.assertEqual(cells[0].tag_name, "th")
        self.assertEqual(cells[1].tag_name, "td")
        self.assertEqual(cells[2].tag_name, "td")

        # Bring up a filter.
        self.switch_to(filters_page)
        tables[1].headers[1].click()
        tables[0].check_headers(("CDR ID", "Filter Title"))
        tables[1].check_headers(("", ""))
        self.assertIn("Sorted By Title", tables[0].caption.text)
        self.assertEqual("", tables[1].caption.text)
        link = tables[0].rows[0][0].find_element(By.TAG_NAME, "a")
        cdr_id = link.text
        title = tables[0].rows[0][1].text
        link.click()
        self.select_new_tab()
        self.assert_title(f"{title} ({cdr_id})")
        self.assert_page_has(f"Filter title: {title}")
        self.assert_page_has("xsl:transform")
        self.click("submit-button-compare")
        self.assert_title(f"{title} ({cdr_id})")
        self.assert_page_has("Filter Comparison")
        self.assert_page_has(f"- {title} on PROD")
        self.assert_regex(rf"\+ {title} on (DEV)|(QA)|(STAGE)")

    def test_filter_sets(self):
        """Test the interface for managing filter sets."""

        # This value is used multiple times locally.
        set_name = "Automated Test Filter Set"

        # We'll need to use this more than once.
        def find_test_set_link():
            for link in self.find("#primary-form ul li a", all=True):
                if link.text == set_name:
                    return link
            return None

        # First is step is to clear out any test data left by earlier tests.
        self.navigate_to("EditFilterSets.py")
        self.assert_title("Manage Filter Sets")
        self.assert_page_has("Filter Sets (click to edit)")
        link = find_test_set_link()
        if link is not None:
            link.click()
            self.select_new_tab()
            self.assert_title("Edit Filter Set")
            self.click("submit-button-delete-set")
            self.assert_title("Manage Filter Sets")
            self.assert_page_has(f"Successfully deleted {set_name}")
            link = find_test_set_link()
            self.assertIsNone(link)

        # Create the test filter set.
        self.click("submit-button-new-filter-set")
        self.select_new_tab()
        self.assert_title("Edit Filter Set")
        self.assert_page_has("Instructions")
        self.set_field_value("name", set_name)
        self.set_field_value("description", "This set is just for testing.")
        self.set_field_value("notes", "Yada, yada.")
        members = self.find("#members li", all=True)
        self.assertEqual(len(members), 1)
        self.assertEqual(members[0].text, "Add some members, please!")
        filters = self.find("#filters li", all=True)
        sets = self.find("#sets li", all=True)
        choices = sets[2], filters[3], sets[1], filters[2], sets[3]
        for choice in choices:
            webdriver.ActionChains(self.driver).double_click(choice).perform()
        expected = [choice.text for choice in choices]
        members = self.find("#members li", all=True)
        actual = [member.text for member in members]
        self.assertEqual(actual, expected)
        args = sets[0], members[2]
        webdriver.ActionChains(self.driver).drag_and_drop(*args).perform()
        expected.insert(2, sets[0].text)
        actual = [m.text for m in self.find("#members li", all=True)]
        self.assertEqual(actual, expected)
        self.click("submit-button-save-set")
        self.assert_title("Edit Filter Set")
        self.assert_regex(rf"Saved set \d+ \({set_name}\) with 6 members")
        self.click("submit-button-manage-filter-sets")
        self.assertIsNotNone(find_test_set_link())

        # Visit the report page for the filter sets.
        self.click("submit-button-report")
        self.select_new_tab()
        self.assert_title("CDR Filter Sets -- Shallow Report")
        self.assert_page_has(f'<dt class="set">{set_name}</dt>')
        types = "set", "filter", "set", "set", "filter", "set"
        for i, type in enumerate(types):
            label = f"[{type[0].upper()}] {actual[i]}"
            self.assert_page_has(f'<dd class="{type}">{label}</dd>')

        # Bring up the deep report.
        self.click("submit-button-deep-report")
        self.assert_title("CDR Filter Sets -- Deep Report")
        test_set_item = None
        for item in self.find("fieldset > ul > li", all=True):
            if item.text.startswith(set_name):
                test_set_item = item
                break
        self.assertIsNotNone(test_set_item)
        link = test_set_item.find_element(By.TAG_NAME, "a")
        self.assertIsNotNone(link)
        filter_id = link.text
        link.click()
        self.select_new_tab()
        self.assert_page_has(f"CDR Document {filter_id}")
        self.assert_page_has("Filter title: ")

        # Clear out the test set.
        self.navigate_to("EditFilterSets.py")
        self.assert_title("Manage Filter Sets")
        link = find_test_set_link()
        self.assertIsNotNone(link)
        link.click()
        self.select_new_tab()
        self.assert_title("Edit Filter Set")
        self.click("submit-button-delete-set")
        self.assert_title("Manage Filter Sets")
        self.assert_page_has(f"Successfully deleted {set_name}")
        link = find_test_set_link()
        self.assertIsNone(link)

    def test_statistics_report(self):
        """Test the Publishing Job Statistics by Date report."""

        self.navigate_to("PubStatsByDate.py")
        self.assert_title("Publishing Job Statistics by Date")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title("Publishing Job Statistics by Date")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected = "Documents Published Between 2020-01-01 and 2020-01-31"
        self.assertEqual(table.caption.text, expected)
        columns = (
            "Doc Type",
            "Re-Added",
            "New",
            "Updated",
            "Updated*",
            "Removed",
            "Total",
        )
        table.check_headers(columns)

    def test_status_report(self):
        """Test the Publishing Status report.

        Because we've selected a time slice in the past, and the
        historical data is static, we are able to make assumptions
        about values in ways which won't work in other tests.

        Additional coverage in test_utilities() below.
        """

        self.navigate_to("PubStatus.py")
        self.assert_title("Publishing Status")
        self.assert_page_has("Report Parameters")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title("Publishing Status")
        self.assert_single_table_report()
        self.assert_wide_report()
        columns = (
            "ID",
            "Type",
            "Status",
            "Started",
            "Finished",
            "Errors",
            "Warnings",
            "Total",
        )
        push = "Push_Documents_To_Cancer.Gov_Export"
        table = self.load_table()
        table.check_headers(columns)
        self.assertEqual(table.rows[0][1].text, push)
        table.rows[0][0].find_element(By.TAG_NAME, "a").click()
        self.assert_title("Publishing Status")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        tables = self.load_tables()
        self.assertEqual(len(tables), 3)
        self.assertEqual(tables[0].caption.text, "Push Job 17808")
        self.assertEqual(tables[1].caption.text, "Parameters")
        self.assertEqual(tables[2].caption.text, "Documents")
        self.click("submit-button-show-document-details")
        self.assert_title("Publishing Status")
        tables = self.load_tables()
        self.assertEqual(len(tables), 3)
        self.assertEqual(tables[0].rows[1][1].text, push)
        self.assertEqual(tables[0].rows[-1][0].text, "Export Job")
        tables[0].rows[-1][1].find_element(By.TAG_NAME, "a").click()
        self.assert_title("Publishing Status")
        tables = self.load_tables()
        self.assertEqual(len(tables), 4)
        self.assertEqual(tables[0].caption.text, "Publishing Job 17807")
        self.assertEqual(tables[1].caption.text, "Parameters")
        self.assertEqual(tables[2].caption.text, "Documents")
        self.assertEqual(tables[3].caption.text, "Warnings")
        self.assertEqual(tables[0].rows[1][1].text, "Export")
        self.assertEqual(tables[0].rows[-1][0].text, "Push Job")
        tables[3].check_headers(("ID", "Type", "Details"))

    def test_utilities(self):
        """Test creating a publishing job, failing it, then re-running it.

        This test can fail if someone else has a hotfix job running
        on the same tier. Nothing to be done about that except try
        the test again after that job has finished. You can also cause
        the test to fail yourself by running it twice in succession so
        close together that the job started by the first test has not
        had enough time to finish when the second test is started. Of
        course, you can avoid that problem by using the Fail Job script
        to kill the first test publishing job. Or we could rewrite this
        test to do that for you.
        """

        # Create the job.
        self.navigate_to("Publishing.py")
        self.assert_title("Publishing")
        self.assert_page_has("Select a Publishing System")
        self.submit_form(new_tab=False)
        self.assert_title("Publishing")
        self.assert_page_has("Select Primary Publication System Subset")
        self.click("subset-hotfix-export")
        self.submit_form(new_tab=False)
        self.assert_title("Hotfix-Export")
        self.assert_page_has("Documents to Publish")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        title = f"{summary_title};Treatment;Health professionals"
        self.set_field_value("docs", summary_id)
        self.select_values("notify", "No")
        self.click("submit-button-publish")
        self.select_new_tab()
        self.assert_title("Publishing Status")
        self.assert_wide_report()
        re_match = re_search(r"Job (\d+) started.", self.get_page_source())
        self.assertIsNotNone(re_match)
        job_id = int(re_match.group(1))
        tables = self.load_tables()
        self.assertEqual(len(tables), 3)
        self.assertEqual(tables[0].caption.text, f"Publishing Job {job_id}")
        self.assertEqual(tables[1].caption.text, "Parameters")
        self.assertEqual(tables[2].caption.text, "Documents")
        self.assertEqual(tables[0].rows[0][0].text, "Publishing System")
        self.assertEqual(tables[0].rows[0][1].text, "Primary")
        self.assertEqual(tables[0].rows[1][0].text, "System Subset")
        self.assertEqual(tables[0].rows[1][1].text, "Hotfix-Export")
        self.assertEqual(tables[0].rows[2][0].text, "User Name")
        self.assertEqual(tables[0].rows[3][0].text, "Output Location")
        self.assertEqual(tables[0].rows[4][0].text, "Started")
        self.assertEqual(tables[0].rows[5][0].text, "Completed")
        self.assertEqual(tables[0].rows[6][0].text, "Status")
        self.assertEqual(tables[0].rows[7][0].text, "Messages")
        self.assertEqual(tables[0].rows[8][0].text, "Total Documents")
        self.assertEqual(tables[0].rows[8][1].text, "1")
        self.assertEqual(tables[2].rows[0][0].text, str(summary_id))
        self.assertEqual(tables[2].rows[0][2].text, "Summary")
        self.assertEqual(tables[2].rows[0][3].text, title)
        self.assertEqual(tables[2].rows[0][4].text, "No")
        self.assertEqual(tables[2].rows[0][5].text, "No")
        self.assertEqual(tables[2].rows[0][6].text, "No")
        tables[1].check_headers(("Name", "Value"))
        columns = (
            "ID",
            "Version",
            "Type",
            "Title",
            "Failed",
            "Warnings",
            "Removed",
        )
        tables[2].check_headers(columns)

        # Kill the job.
        self.navigate_to("FailBatchJob.py")
        self.assert_title("Mark Stuck Job(s) as Failed")
        self.assert_page_has("Instructions")
        self.assert_page_has("This tool marks stalled publishing and batch")
        self.assert_single_table_report()
        table = self.load_table()
        pattern = r"\d+ Stalled or Active Jobs?"
        self.assert_regex(pattern)
        columns = "✓", "Job", "Type", "Started", "Status", "Name"
        table.check_headers(columns)
        job_row = None
        for row in table.rows:
            if row[1].text == str(job_id):
                job_row = row
                break
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[2].text, "publishing")
        self.assertEqual(job_row[3].text, str(date.today()))
        self.assertEqual(job_row[5].text, "Hotfix-Export")
        self.click(f"job-p{job_id}")
        self.submit_form(new_tab=False)
        self.assert_page_has(f"Marked publishing job {job_id} as failed.")

        # Republish the job.
        self.navigate_to("Republish.py")
        self.assert_title("Re-Publish CDR Documents to Cancer.gov")
        self.assert_page_has("Instructions")
        self.assert_page_has("This page can be used to request re-publishing")
        self.set_field_value("jobs", job_id)
        self.submit_form()
        self.assert_title("Re-Publish CDR Documents to Cancer.gov")
        self.assert_regex(r"Export job .+ created successfully.")


class SummaryTests(Tester):
    """Tests of reports for CDR Summary documents."""

    def test_advanced_search(self):
        """Test the Summary Advanced Search page."""

        self.navigate_to("SummarySearch.py")
        self.assert_title("Summary")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.set_field_value("title", summary_title)
        self.select_values("audience", "Health professionals")
        self.select_values("status", "A")
        self.find("main form input[value='Search']").click()
        self.select_new_tab()
        self.assert_title("Summary")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected = (
            f"1 document matches '{summary_title}' and 'Health professionals' "
            "and 'A'"
        )
        self.assertEqual(table.caption.text, expected)
        self.assertEqual(len(table.rows), 1)
        self.assertEqual(table.rows[0][0].text, "1.")
        self.assertEqual(table.rows[0][1].text, f"CDR{summary_id:010d}")
        expected = f"{summary_title}; Treatment; Health professionals"
        self.assertEqual(table.rows[0][2].text, expected)
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.assert_plain_report()
        self.assert_page_has(f"<h2>{summary_title}</h2>")
        self.assert_page_has("<h3>Treatment - Health professionals</h3>")

    def test_changes_to_summaries_reports(self):
        """Test the Changes to Summaries report."""

        # There are two versions of this report. This is the standard version.
        board = self.get_test_board()
        self.navigate_to("ChangesToSummaries.py")
        self.assert_title("Changes To Summaries Report")
        self.assert_page_has("Select PDQ Board For Report")
        self.select_values("board", board["id"])
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_plain_report()
        self.assert_page_has("Changes to Summaries Report (Excluding Modules)")
        self.assert_page_has("From 2020-01-01 to 2020-01-31")
        self.assert_page_has("<th>Only usable as standalone summaries:</th>")
        self.assert_page_has("<th>Only usable as summary modules:</th>")
        self.assert_page_has("<th>Can be used standalone or as modules:</th>")
        self.assert_page_has("<th>Total summaries:</th>")
        self.assert_page_has(board["name"])
        self.assert_page_has("Health Professionals")
        self.assert_page_has("Publishable Version Date:")
        self.assert_page_has("Changes to This Summary")

        # And this is the historical version (single summary).
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        form = self.navigate_to("SummaryChanges.py")
        self.assert_title("History of Changes to Summary")
        self.assert_page_has("Term ID or Title for Summary")
        self.set_field_value("DocId", summary_id)
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "12/31/2020")
        self.submit_form()
        self.assert_title("History of Changes to Summary")
        self.assert_page_has(summary_title)
        self.assert_page_has("Changes made 2020-01-01 through 2020-12-31")
        self.assert_page_has("Publishable Version Date:")
        self.assert_page_has("Changes to This Summary")

        # Test the "Complete History" option.
        self.switch_to(form)
        self.click("scope-all")
        self.submit_form()
        self.assert_title("History of Changes to Summary")
        self.assert_page_has(summary_title)
        self.assert_page_has("Complete history of changes in the CDR")
        self.assert_page_has("Publishable Version Date:")
        self.assert_page_has("Changes to This Summary")

    def test_citations_report(self):
        """Test Summary Citations report."""

        self.navigate_to("SummaryCitations.py")
        self.assert_title("Summary Citations Report")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("Summary Citations Report")
        self.assert_page_has(f"Select version for CDR{summary_id}")
        self.select_values("DocVersion", "0")
        self.submit_form(new_tab=False)
        self.assert_title("Summary Citations Report")
        self.assert_title(summary_title)
        self.assert_page_has("<h2>References</h2>")
        ordered_list = self.driver.find_element(By.CSS_SELECTOR, "ol li")
        self.assertIsNotNone(ordered_list)
        self.assert_non_tabular_report()

    def test_comments_report(self):
        """Test the Summary Comments report."""

        self.navigate_to("SummaryComments.py")
        self.assert_title("Summary Comments Report")
        board = self.get_test_board()
        board_id = board["id"]
        self.click(f"board-{board_id}")
        self.submit_form()
        short_board_name = board["name"].replace("PDQ ", "")
        shorter_board_name = short_board_name.replace(" Editorial Board", "")
        self.assert_title(
            f"Comments for English Health Professional {shorter_board_name} "
            f"Summaries – {date.today()}"
        )
        self.assert_plain_report()
        self.assert_multi_table_report()
        summary = self.get_test_summary()
        self.assert_page_has(summary["title"])
        columns = "Summary Section Title", "Comments", "Blank"
        for column in columns:
            self.assert_regex(rf'<th style="width: \d+px">{column}</th>')
        self.assert_page_has("color: brown")
        self.assert_page_has("color: green")

    def test_comprehensive_review_dates_report(self):
        """Test the Summaries Comprehensive Review Dates report."""

        self.navigate_to("SummaryCRD.py")
        self.assert_title("Summaries Comprehensive Review Dates")
        self.submit_form()
        title = f"English Health Professional Summaries {date.today()}"
        self.assert_title(title)
        board = self.get_test_board()
        short_name = board["name"].replace("PDQ ", "")
        self.assert_page_has(f"{short_name} (summaries)")
        columns = "Summary Title", "Date", "Status", "Comment"
        for column in columns:
            self.assert_regex(f"<th style=[^>]+>{column}</th>")
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()

    def test_date_last_modified_report(self):
        """Test the Summary Date Last modified report."""

        # First version of the report: by user dates.
        script = "SummaryDateLastModified.py"
        self.navigate_to(script)
        self.assert_title("Summary Date Last Modified")
        end = date.today()
        start = end - timedelta(365)
        params = [
            ("est", "all"),
            ("sst", "all"),
            ("opt", "modules"),
            ("opt", "blocked"),
            ("opt", "unpub"),
            ("u-start", str(start)),
            ("u-end", str(end)),
            ("Request", "Submit"),
        ]
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "DLM Report")
        title = "Summary Date Last Modified (User) Report"
        self.assertEqual(sheet["A1"].value, title)
        self.assertEqual(sheet["A2"].value, f"{start} - {end}")
        self.assertEqual(sheet["A3"].value, f"Report Date: {date.today()}")
        self.assertEqual(sheet.max_column, 6)
        self.assertEqual(sheet["A7"].value, "DocID")
        self.assertEqual(sheet["B7"].value, "Summary Title")
        self.assertEqual(sheet["C7"].value, "Date Last Modified")
        self.assertEqual(sheet["D7"].value, "Last Modify Action Date (System)")
        self.assertEqual(sheet["E7"].value, "LastV Publish?")
        self.assertEqual(sheet["F7"].value, "User")

        # The other version: by system date.
        params = [
            ("est", "all"),
            ("sst", "all"),
            ("opt", "modules"),
            ("opt", "blocked"),
            ("opt", "unpub"),
            ("s-start", str(start)),
            ("s-end", str(end)),
            ("Request", "Submit"),
        ]
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "DLM Report")
        title = "Summary Last Modified Date (System) Report"
        self.assertEqual(sheet["A1"].value, title)
        self.assertEqual(sheet["A2"].value, f"{start} - {end}")
        self.assertEqual(sheet["A3"].value, f"Report Date: {date.today()}")
        self.assertEqual(sheet.max_column, 9)
        self.assertEqual(sheet["A7"].value, "DocID")
        self.assertEqual(sheet["B7"].value, "Summary Title")
        self.assertEqual(sheet["C7"].value, "Type")
        self.assertEqual(sheet["D7"].value, "Aud")
        self.assertEqual(sheet["E7"].value, "Last Comment")
        self.assertEqual(sheet["F7"].value, "Date Last Modified")
        self.assertEqual(sheet["G7"].value, "Last Modify Action Date (System)")
        self.assertEqual(sheet["H7"].value, "LastV Publish?")
        self.assertEqual(sheet["I7"].value, "User")

    def test_internal_links_report(self):
        """Test the Summary Internal Links report."""

        script = "ocecdr-3650.py"
        self.navigate_to(script)
        self.assert_title("Summary Internal Links")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        params = dict(id=summary_id, Request="Submit")
        book = self.fetch_workbook(script, params)
        sheet = book.active
        self.assertEqual(sheet.title, "Sheet1")
        suffix = "Treatment;Health professionals"
        expected = f"Links for CDR{summary_id} ({summary_title};{suffix})"
        self.assertEqual(sheet.max_column, 6)
        self.assertEqual(sheet["A1"].value, expected)
        self.assertEqual(sheet["A3"].value, "FragID")
        self.assertEqual(sheet["B3"].value, "Target Section/Subsection")
        self.assertEqual(sheet["C3"].value, "Linking Section/Subsection")
        self.assertEqual(sheet["D3"].value, "Text in Linking Node")
        self.assertEqual(sheet["E3"].value, "In Table?")
        self.assertEqual(sheet["F3"].value, "In List?")

    def test_lists_report(self):
        """Test the Summaries Lists report."""

        self.navigate_to("SummariesLists.py")
        self.assert_title("Summaries Lists")
        self.click("show_id-y")
        self.submit_form()
        title = f"PDQ English Health Professional Summaries {date.today()}"
        self.assert_title(title)
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        board = self.get_test_board()
        short_name = board["name"].replace("PDQ ", "")
        self.assert_regex(rf"{short_name} \(\d+ summaries\)")
        self.assert_page_has("<th>CDR ID</th>")
        self.assert_page_has("<th>Title</th>")
        summary = self.get_test_summary()
        self.assert_page_has(summary["id"])
        self.assert_page_has(summary["title"])

    def test_mailer_reports(self):
        """Test the Summary Mailer report."""

        # Verify that the Excel workbook is successfully created,
        # and that it has the correct sheet name, title cell value,
        # and column header names. We also verify that the board
        # member column is sorted and the summary column is not.
        self.navigate_to("SummaryMailerReport.py", flavor="standard")
        self.assert_title("Summary Mailer Report")
        script = "SummaryMailerReport.py"
        params = dict(flavor="standard", sort="member", show="last")
        board = self.get_test_board()
        params["board"] = board["id"]
        params["Request"] = "Submit"
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "Summary Mailer Report")
        board_name = board["name"]
        report_type = "Summary Mailer Report (Last)"
        expected = f"{report_type} - {board_name} - {date.today()}"
        self.assertEqual(sheet["A1"].value, expected)
        self.assertEqual(sheet["A3"].value, "Mailer ID")
        self.assertEqual(sheet["B3"].value, "Board Member")
        self.assertEqual(sheet["C3"].value, "Summary")
        self.assertEqual(sheet["D3"].value, "Sent")
        self.assertEqual(sheet["E3"].value, "Response")
        self.assertEqual(sheet["F3"].value, "Changes")
        self.assertEqual(sheet["G3"].value, "Comments")
        rows = [row for row in sheet][3:]
        first_row_count = len(rows)
        members = [row[1].value for row in rows]
        summaries = [row[2].value for row in rows]
        self.assertEqual(members, sorted(members, key=str.lower))
        self.assertNotEqual(summaries, sorted(summaries, key=str.lower))

        # Change the sort to summaries and verify the change. Also verify
        # that we get the same number of rows we got the first time.
        params["sort"] = "summary"
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "Summary Mailer Report")
        rows = [row for row in sheet][3:]
        self.assertEqual(len(rows), first_row_count)
        members = [row[1].value for row in rows]
        summaries = [row[2].value for row in rows]
        self.assertNotEqual(members, sorted(members, key=str.lower))
        self.assertEqual(summaries, sorted(summaries, key=str.lower))

        # With the option to use the date the mailer response was received,
        # we should have fewer rows in the report, because it never happens
        # that all the requests for reviews get a response (that would be
        # a miracle well worth having the test fail). 😉
        params["show"] = "last-checked-in"
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "Summary Mailer Report")
        rows = [row for row in sheet][3:]
        row_count = len(rows)
        self.assertLess(row_count, first_row_count)

        # The history version of the report allows for narrowing by date range.
        self.navigate_to("SummaryMailerReport.py", flavor="history")
        self.assert_title("Summary Mailer History Report")
        params = dict(
            flavor="history",
            sort="member",
            board=board["id"],
            Request="Submit",
            start="2010-01-01",
            end="2010-12-31",
        )
        book = self.fetch_workbook(script, params)
        self.assertIsNotNone(book)
        sheet = book.active
        self.assertEqual(sheet.title, "Summary Mailer History Report")
        report_type = "Summary Mailer History Report (2010-01-01 - 2010-12-31)"
        expected = f"{report_type} - {board_name} - {date.today()}"
        rows = [row for row in sheet][3:]
        for row in rows:
            sent = row[3].value
            self.assertGreaterEqual(sent, "2010-01-01")
            self.assertLessEqual(sent, "2010-12-31")

    def test_mailer_requests(self):
        """Test the Summary Mailer request form."""

        self.navigate_to("SummaryMailerReqForm.py")
        self.assert_title("PDQ Advisory Board Members Tracking Request Form")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        board = None
        for values in summary["boards"]:
            if "Advisory" in values["name"]:
                board = values
        self.assertIsNotNone(board)
        self.select_values("board", board["id"])
        self.click("selection_method-summary")
        self.select_values("summaries", summary_id)
        self.submit_form()
        self.assert_title("PDQ Advisory Board Members Tracking Request Form")
        self.assert_page_has(board["name"])
        checkbox = self.driver.find_element(By.ID, f"outer-{summary_id}")
        self.assertTrue(checkbox.is_selected)
        report = self.submit_form()
        self.assert_title("Tracker Documents Generated")
        table = self.load_table()
        table.check_headers(("Tracker", "Summary", "Reviewer"))
        table.rows[0][0].find_element(By.TAG_NAME, "a").click()
        self.select_new_tab()
        self.assert_page_has("</Mailer>")
        self.assert_regex(rf"<Sent>{date.today()}T\d\d:\d\d:\d\d</Sent>")
        self.switch_to(report)
        table.rows[0][1].find_element(By.TAG_NAME, "a").click()
        self.select_new_tab()
        self.assert_page_has(f"<h2>{summary_title}</h2>")
        self.assert_page_has("<h3>Treatment - Health professionals</h3>")
        self.switch_to(report)
        table.rows[0][2].find_element(By.TAG_NAME, "a").click()
        self.select_new_tab()
        self.assert_page_has("<center>Person<br>QC Report</center>")

    def test_markup_report(self):
        """Test Summaries With Markup report."""

        self.navigate_to("SummariesWithMarkup.py")
        self.assert_title("Summaries With Markup")
        self.submit_form()
        self.assert_title(
            "English Health Professional Summaries "
            f"With Markup - {date.today()}"
        )
        board_name = self.get_test_board()["name"]
        short_name = board_name.replace("PDQ ", "")
        really_short_name = short_name.replace(" Editorial Board", "")
        self.assert_page_has(really_short_name)
        self.assert_multi_table_report()
        self.assert_tables_in_grid_container()
        columns = (
            "Doc ID",
            "Summary Title",
            "Publish",
            "Approved",
            "Proposed",
            "Rejected",
            "Advisory",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

    def test_metadata_report(self):
        """Test the Summary Metadata report."""

        # Test the single-summary version of the report.
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        form = self.navigate_to("SummaryMetaData.py")
        self.assert_title("Summary Metadata Report")
        self.set_field_value("doc-id", summary_id)
        self.submit_form()
        self.assert_title(f"{summary_title} (CDR{summary_id})")
        self.assert_plain_report()
        self.assert_single_table_report()
        self.assert_page_has("<td>CDR ID</td>")
        self.assert_page_has("<td>Summary Title</td>")
        self.assert_page_has("<td>Advisory Board</td>")
        self.assert_page_has("<td>Editorial Board</td>")
        self.assert_page_has("<td>Topics</td>")

        # Test the multi-summary version of the report.
        self.switch_to(form)
        self.click("method-group")
        self.submit_form()
        self.assert_title(
            "English Health Professional Summaries "
            "for Adult Treatment"
        )
        self.assert_plain_report()
        self.assert_multi_table_report()

    def test_non_journal_article_citations_report(self):
        """Test Summaries With Non-Journal Article Citations report."""

        self.navigate_to("SummariesWithNonJournalArticleCitations.py")
        title = "Summaries With Non-Journal Article Citations Report"
        self.assert_title(title)
        self.assert_page_has("Select Citation Type (one or more)")
        self.click("type-abstract-[internet]")
        self.click("type-database")
        self.click("type-database-entry")
        self.click("type-legal-material")
        self.submit_form()
        self.assert_title(title)
        self.assert_plain_report()
        self.assert_single_table_report()
        columns = (
            "Summary ID",
            "Summary Title",
            "Summary Sec Title",
            "Citation Type",
            "Citation ID",
            "Citation Title",
            "Publication Details/Other Publication Info",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

    def test_protocol_ref_links_report(self):
        """Test the ProtocolRef Links in Summaries report."""

        script = "SummaryProtocolRefLinks.py"
        self.navigate_to(script, prompt="yes")
        self.assert_title("ProtocolRef Links in Summaries")
        self.assert_page_has("Instructions")
        self.assert_page_has("The report page has two tables.")
        self.navigate_to(script, limit=2)
        self.assert_title("ProtocolRef Links in Summaries")
        self.assert_page_has("Limited to 2 summaries for testing.")
        self.assert_wide_report()
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        self.assertEqual(tables[0].caption.text, "Total Count by Link Type")
        tables[0].check_headers(("Protocol Links Including ...", "Count"))
        self.assertEqual(len(tables[0].rows[0]), 2)
        self.assertRegex(tables[0].rows[0][1].text, r"^\d+$")
        self.assertEqual(tables[1].caption.text, "Links to Clinical Trials")
        columns = "CDR ID", "Summary Title", "Protocol ID", "Protocol Link"
        self.assertEqual(len(tables[1].rows[0]), len(columns))
        tables[1].check_headers(columns)
        self.assertRegex(tables[1].rows[0][0].text, r"^\d+$")

    def test_publish_preview_report(self):
        """Test publish preview for summaries."""

        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="pp")
        self.assert_title("Publish Preview Report")
        self.set_field_value("DocTitle", f"{summary_title[0]}%")
        self.submit_form()
        self.assert_title("Publish Preview Report")
        expected = f"Multiple matches found for '{summary_title[0]}%'."
        self.assert_page_has(expected)
        self.assert_page_has("Choose Document")
        self.click(f"docid-{summary_id}")
        self.submit_form()
        self.assert_page_has(f"<h1>{summary_title}")
        self.assert_page_has("(PDQ®)–Health Professional Version</h1>")
        self.assert_page_has("On This Page")
        self.assert_page_has("Go to Patient Version")

    def test_qc_reports(self):
        """Test the various flavors of Summary QC reports."""

        # Test the HP Bold/Underline QC report.
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="bu")
        self.assert_title("HP Bold/Underline QC Report")
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("HP Bold/Underline QC Report")
        self.assertTrue(self.select_version("RL"))
        self.click("options-markup-advisory")
        self.click("options-markup-publish")
        self.click("options-markup-proposed")
        self.click("options-com-all")
        self.click("show-options")
        comment_options = (
            "options-com-aud-int",
            "options-com-aud-ext",
            "options-com-src-ed",
            "options-com-src-adv",
            "options-com-dur-perm",
            "options-com-dur-temp",
        )
        for id in comment_options:
            self.assertTrue(self.driver.find_element(By.ID, id).is_selected())
        self.click("options-glossary")
        self.click("options-loes")
        self.submit_form()
        self.assert_page_has(summary_title)
        self.assert_page_has("<h3>Treatment - Health professionals</h3>")
        self.assert_page_has("<h1>Table of Glossary Terms</h1>")
        self.assert_page_has("<h1>Table of Level of Evidence Terms</h1>")
        styles = (
            ("insertproposed", "font-weight", "bold"),
            ("insertproposed", "text-decoration", "underline"),
            ("deleteproposed", "text-decoration", "line-through"),
        )
        for sel, name, value in styles:
            self.assert_regex(rf"\.{sel}[^{{]*{{[^}}]*{name}[^}}]*{value}")

        # Try the HP redline/strikeout version.
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="rs")
        self.assert_title("HP Redline/Strikeout QC Report")
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("HP Redline/Strikeout QC Report")
        self.assertTrue(self.select_version("RL"))
        self.submit_form()
        self.assert_page_has(summary_title)
        self.assert_page_has("<h3>Treatment - Health professionals</h3>")
        styles = (
            ("insertproposed", "color", "green"),
            ("deleteproposed", "color", "green"),
            ("deleteproposed", "text-decoration", "line-through"),
        )
        for selector, name, value in styles:
            self.assert_regex(f".{selector}[^{{]*{{[^}}]*{name}[^}}]*{value}")

        # Make sure the HP report fails with a patient summary.
        # TODO: It would be nice if the software detected the problem sooner.
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="bu")
        summary = self.get_test_summary(audience="Patients")
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("HP Bold/Underline QC Report")
        self.submit_form()
        self.assert_page_has("<h2>ERROR:</h2>")
        self.assert_page_has(
            "<h3>HP Bold/Underline QC Report "
            "not valid for patient summaries</h3>"
        )

        # Now try the bold/underline QC report for a Patient summary.
        # Patient summary QC reports have a Key Points block, which HP
        # report don't have.
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="patbu")
        self.assert_title("PT Bold/Underline QC Report")
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("PT Bold/Underline QC Report")
        version_comment = "Edits to bone scan in RLSO; 1/8/18"
        self.assertTrue(self.select_version(version_comment))
        self.submit_form()
        self.assert_page_has(summary_title)
        self.assert_page_has("<h3>Treatment - Patients</h3>")
        self.assert_page_has("Key Points for This Section")
        styles = (
            ("insertproposed", "font-weight", "bold"),
            ("insertproposed", "text-decoration", "underline"),
            ("deleteproposed", "text-decoration", "line-through"),
        )
        for sel, name, value in styles:
            self.assert_regex(rf"\.{sel}[^{{]*{{[^}}]*{name}[^}}]*{value}")

        # Try the patient redline/strikeout version.
        self.navigate_to("QcReport.py", DocType="Summary", ReportType="pat")
        self.assert_title("PT Redline/Strikeout QC Report")
        self.set_field_value("DocId", summary_id)
        self.submit_form()
        self.assert_title("PT Redline/Strikeout QC Report")
        self.assertTrue(self.select_version(version_comment))
        self.click("options-com-all")
        self.submit_form()
        self.assert_page_has(summary_title)
        self.assert_page_has("<h3>Treatment - Patients</h3>")
        self.assert_page_has("[Comment: ")
        self.assert_page_has("[Permanent-Comment: ")
        self.assert_page_has('<span class="deleteapproved">08/18/2017</span>')
        self.assert_page_has('<span class="insertapproved">01/08/2018</span>')
        styles = (
            ("insertproposed", "color", "green"),
            ("deleteproposed", "color", "green"),
            ("deleteproposed", "text-decoration", "line-through"),
        )
        for selector, name, value in styles:
            self.assert_regex(f".{selector}[^{{]*{{[^}}]*{name}[^}}]*{value}")

    def test_replacement(self):
        """Test replacement of an old summary with a new one."""

        # Do some prep work, creating two test summaries.
        root = etree.Element("Summary", nsmap=self.NSMAP)
        old_title = "Old Test Summary"
        new_title = "New Test Summary"
        title_node = etree.SubElement(root, "SummaryTitle")
        title_node.text = old_title
        xml = etree.tostring(root, encoding="unicode")
        old_doc_id = self.save_doc(xml, "Summary")
        old_cdr_id = f"CDR{old_doc_id:010d}"
        title_node.text = new_title
        xml = etree.tostring(root, encoding="unicode")
        new_doc_id = self.save_doc(xml, "Summary")
        new_cdr_id = f"CDR{new_doc_id:010d}"

        # Verify that the software will refuse the operation
        # without the WillReplace element in place.
        self.navigate_to("ReplaceDocWithNewDoc.py")
        self.assert_title("Replace Old Document With New One")
        self.assert_page_has("This program replaces the XML of a CDR document")
        self.assert_page_has("Enter IDs for Old and New Documents")
        self.set_field_value("old", old_doc_id)
        self.set_field_value("new", new_doc_id)
        self.submit_form()
        self.assert_title("Replace Old Document With New One")
        self.assert_page_has(f"CDR{new_doc_id} has no WillReplace element.")

        # Add the element and try again. This time should succeed.
        will_replace = etree.SubElement(root, "WillReplace")
        will_replace.text = old_title
        will_replace.set(f"{{{self.NS}}}ref", old_cdr_id)
        xml = etree.tostring(root, encoding="unicode")
        self.save_doc(xml, "Summary", id=new_cdr_id)
        self.submit_form(new_tab=False)

        # Confirm the replacement. The validation errors reported are
        # expected (our test Summary documents are pretty bare-bones).
        self.assert_title("Replacement Confirmation Required")
        self.assert_regex(rf"{old_cdr_id} .+ will be replaced by {new_cdr_id}")
        expected = (
            "These errors occurred when validating the new document:",
            "Replacing an existing document with a new one that is invalid "
            "is allowed, but please consider whether you really want to do "
            "that.",
            "The new document will be saved as a non-publishable version.",
        )
        for text in expected:
            self.assert_page_has(text)
        self.click("submit-button-confirm-replacement")
        self.assert_title("Replacement Successful")
        self.assert_page_has(
            f"XML from {new_cdr_id} successfully saved as a new "
            f"version of {old_cdr_id}."
        )

        # Don't leave any dross behind.
        self.delete_doc(old_cdr_id)
        self.delete_doc(new_cdr_id)

    def test_standard_wording_report(self):
        """Test the Summaries Standard Wording report."""

        self.navigate_to("SummaryStandardWording.py")
        self.assert_title("Summaries Standard Wording")
        self.set_field_value("term", "radioactive glucose")
        self.driver.find_element(By.CSS_SELECTOR, "#search-terms img").click()
        self.set_field_value("term-2", "malignant tumor")
        self.submit_form()
        self.assert_title("Standard Wording Report")
        self.assert_plain_report()
        self.assert_single_table_report()
        self.assert_page_has("STANDARD WORDING REPORT (PATIENT)")
        self.assert_page_has(
            "Search Terms: "
            "malignant tumor; radioactive glucose"
        )
        columns = (
            "Doc ID",
            "Doc Title",
            "Match",
            "Context",
            "Standard Wording[?]",
        )
        for column in columns:
            self.assert_regex(f"<th[^>]*>{column}</th>")

    def test_svpc_summaries_report(self):
        """Test the SVPC Summaries report."""

        self.navigate_to("SVPCSummariesReport.py")
        self.assert_title("SVPC Summaries Report")
        self.set_field_value("start", "1/1/2022")
        self.set_field_value("end", "12/31/2022")
        self.submit_form()
        self.assert_title("SVPC Summaries Report")
        self.assert_single_table_report()
        self.assert_plain_report()
        self.assert_regex(r"SVPC Summaries [(]\d+[)]")
        self.assert_page_has("Language: Any")
        self.assert_page_has("Date Range: 2022-01-01 - 2022-12-31")
        self.assert_page_has("Non-publishable summaries: excluded")
        table = self.load_table()
        columns = "CDR ID", "Title", "Summary Type", "Publication Date"
        table.check_headers(columns)

    def test_toc_report(self):
        """Test the Summary Table of Contents Lists report."""

        self.navigate_to("SummariesTocReport.py")
        self.assert_title("Summary TOC Lists")
        self.submit_form()
        self.assert_title("PDQ English Health Professional Summaries")
        self.assert_plain_report()
        self.assert_non_tabular_report()
        board_name = self.get_test_board()["name"]
        short_name = board_name.replace("PDQ ", "")
        really_short_name = short_name.replace(" Editorial Board", "")
        self.assert_page_has(f"<h2>{really_short_name}</h2>")
        summary = self.get_test_summary()
        summary_id = summary["id"]
        summary_title = summary["title"]
        self.assert_page_has(f"<h5>CDR{summary_id} - {summary_title}</h5>")
        nested_list = self.driver.find_element(By.CSS_SELECTOR, "ul ul ul")
        self.assertIsNotNone(nested_list)

    def test_translation(self):
        """Test round-trip of a summary through the translation system."""

        # Get an English summary, which we'll pretend to translate.
        script = "get-english-summary.py"
        self.navigate_to(script)
        self.assert_title("Fetch English Summary For Translation")
        self.assert_page_has("Instructions")
        self.assert_page_has("Document ID is required, and must be an integer")
        english_summary = self.get_test_summary()
        english_summary_id = english_summary["id"]
        self.set_field_value("id", english_summary_id)
        self.set_field_value("version", "pub")
        self.submit_form()
        self.assert_page_has(f"CDR Document CDR{english_summary_id:010d}")
        self.assert_page_has('<span class="tag">Summary</span>')
        params = dict(
            id=english_summary_id,
            version="pub",
            fmt="raw",
            Request="Submit",
            Session=self.session,
        )
        url = f"{self.cgi}/{script}?{urlencode(params)}"
        xml = self.fetch_from_url(url)

        # Modify the summary XML ever so slightly and use it to create a new
        # Spanish summary.
        root = etree.fromstring(xml)
        spanish_title = "Esta es la versión de prueba en español"
        root.find("SummaryTitle").text = spanish_title
        path = Path("spanish-summary.xml").resolve()
        self.logger.debug("writing %s", path)
        path.write_bytes(etree.tostring(root, encoding="utf-8"))
        self.navigate_to("post-translated-summary.py")
        self.assert_title("Create World Server Translated Summary")
        self.set_field_value("file", str(path))
        comment = f"Created by automated translation test {datetime.now()}"
        self.set_field_value("comment", comment)
        self.submit_form(new_tab=False)
        self.assert_title("Create World Server Translated Summary")
        pattern = r"Successfully created (CDR\d+)\."
        re_match = re_search(pattern, self.get_page_source())
        self.assertIsNotNone(re_match)
        spanish_cdr_id = re_match.group(1)

        # Clean up behind ourselves.
        reason = f"Deleted by automated translation test {datetime.now()}"
        self.delete_doc(spanish_cdr_id, reason)
        self.logger.debug("removing %s", path)
        path.unlink()

    def test_translation_job_workflow_report(self):
        """Test the Summary Translation Job Workflow report."""

        self.navigate_to("translation-job-report.py")
        self.assert_title("Translation Job Workflow Report")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.click("type-history")
        self.submit_form()
        self.assert_title("Translation Job History Report")
        self.assert_single_table_report()
        self.assert_wide_report()
        columns = (
            "CDR ID",
            "Title",
            "Audience",
            "Assigned To",
            "Translation Status",
            "Translation Status Date",
            "Type of Change",
            "TRANSLATED DOC CDR ID",
            "Comments",
        )
        table = self.load_table()
        table.check_headers(columns)

    def test_translation_queue(self):
        """Test the Summary translation queue pages.

        This test assumes that the tester account (full name "Regression
        Tester") exists, and that it has an email address associated with
        it (preferably one to which you have acceess).
        """

        # Take care of some preparatory steps first.
        self.add_user_to_group("tester", "Spanish Translators")
        sql = "SELECT id FROM usr WHERE name = 'tester'"
        user_id = int(self.run_query(sql)[0][0])
        self.logger.debug("tester user ID: %d", user_id)
        test_title = "Test English summary for translation queue"
        english_id = self.create_test_summary(title=test_title, svpc=True)
        self.logger.debug("english summary ID: %s", english_id)

        # Test the landing page for the translation queues.
        self.navigate_to("TranslationJobQueues.py", testing=True)
        self.assert_title("Translation Job Queues")
        self.assert_page_has("There are multiple translation job queues")
        self.assert_page_has("Available Queues")
        links = self.find("form fieldset ul li a", all=True)
        labels = (
            "Glossary Translation Job Queue",
            "Media Translation Job Queue",
            "Summary Translation Job Queue",
        )
        self.assertEqual(len(links), len(labels))
        for i, link in enumerate(links):
            self.assertEqual(link.text, labels[i])

        # Clear out any dross left from earlier failed tests.
        links[2].click()
        self.select_new_tab()
        self.assert_title("Summary Translation Job Queue")
        self.assert_single_table_report()

        def find_job_row():
            table = self.load_table()
            for row in table.rows:
                if row[1].text == test_title:
                    return row
            return None
        obsolete_docs = []
        while (job_row := find_job_row()) is not None:
            link = job_row[0].find_element(By.TAG_NAME, "a")
            obsolete_docs.append(link.text)
            self.logger.debug("removing translation job for %s", link.text)
            link.click()
            self.click("submit-button-delete")
            alert = self.wait.until(expected_conditions.alert_is_present())
            self.logger.debug("alert text: %s", alert.text)
            alert.accept()
            sleep(1)
        for obsolete_doc in obsolete_docs:
            self.delete_doc(obsolete_doc)

        # Open the translation job queue for summaries and create a new job.
        self.navigate_to("translation-jobs.py", testing=True)
        comment = "job for automated test"
        self.click("submit-button-add")
        self.assert_title("Summary Translation Job")
        self.assert_page_has("Select English Summary")
        self.select_values("english_id", english_id)
        self.submit_form(new_tab=False)
        self.assert_title("Summary Translation Job")
        statuses = {}
        for option in self.find("select#status option", all=True):
            statuses[option.text] = option.get_attribute("value")
        self.select_values("status", statuses["Ready for Translation"])
        expected = f"English summary: CDR{english_id} ({test_title})"
        self.assert_page_has(expected)
        self.assert_page_has("Create Translation Job")
        self.select_values("assigned_to", user_id)
        self.select_values("change_type", "1")
        self.set_field_value("comments", comment)
        paths = [Path(f"test-{i}.txt") for i in (1, 2)]
        for path in paths:
            path.write_text("test file for automated translation queue test\n")
        files = " \n ".join([f"{path.resolve()}" for path in paths])
        self.set_field_value("files", files)
        self.submit_form(new_tab=False)

        # Check the queue page, make sure the job is there.
        self.driver.set_window_size(2048, 1024)
        self.assert_title("Summary Translation Job Queue")
        table = self.load_table()
        columns = (
            "CDR ID",
            "Title",
            "Audience",
            "Status",
            "Assigned To",
            "Date",
            "Type of Change",
            "Comments",
        )
        table.check_headers(columns)
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.assertEqual(job_row[0].text, f"CDR{english_id}")
        self.assertEqual(job_row[1].text, test_title)
        self.assertEqual(job_row[2].text, "Patients")
        self.assertEqual(job_row[3].text, "Ready for Translation")
        self.assertEqual(job_row[4].text, "Regression Tester")
        self.assertEqual(job_row[5].text, str(date.today()))
        self.assertEqual(job_row[6].text, "New Summary")
        self.assertEqual(job_row[7].text, comment)

        # Test editing the job.
        job_row[0].find_element(By.TAG_NAME, "a").click()
        self.select_values("status", statuses["Translation Made Publishable"])
        checkboxes = self.find("input[name='attachments-to-drop']", all=True)
        self.assertEqual(len(checkboxes), 2)
        for checkbox in checkboxes:
            self.click(checkbox.get_attribute("id"))
        self.submit_form(new_tab=False)

        # Our job has reached the final state, so it should be purgeable.
        job_row = find_job_row()
        self.assertIsNotNone(job_row)
        self.click("submit-button-purge")
        job_row = find_job_row()
        self.assertIsNone(job_row)

        # Clean up behind ourselves.
        self.remove_user_from_group("tester", "Spanish Translators")
        for path in paths:
            path.unlink()
        self.delete_doc(english_id)

    def test_type_of_change_report(self):
        """Test the Summaries Type of Change report."""

        # Test the default report (recent changes only).
        form = self.navigate_to("SummaryTypeChangeReport.py")
        self.assert_title("Summaries Type of Change")
        self.submit_form()
        self.assert_title("Current Summary Changes")
        self.assert_plain_report()
        self.assert_single_table_report()
        self.assert_page_has("Type of Change Report (Most Recent Change)")
        columns = (
            "Summary",
            "Comprehensive revision",
            "Comment",
            "Major change",
            "New summary",
            "Reformat",
        )
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")

        # Test the historical version of the report.
        self.switch_to(form)
        self.click("type-historical-(all-changes-for-a-given-date-range)")
        self.set_field_value("start", "1/1/2020")
        self.set_field_value("end", "1/31/2020")
        self.submit_form()
        self.assert_title("Summary Changes -- 2020-01-01 - 2020-01-31")
        self.assert_page_has("Type of Change Report (All Changes by Summary)")
        columns = "Summary", "Date", "Type of Change", "Comment"
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")
        self.assert_plain_report()
        self.assert_single_table_report()

        # Test the multi-table variant of the historical report.
        self.switch_to(form)
        self.click("organization-one-table-for-each-type-of-change")
        self.submit_form()
        self.assert_title("Summary Changes -- 2020-01-01 - 2020-01-31")
        self.assert_page_has("Type of Change Report")
        columns = "Summary", "Date", "Comment"
        for column in columns:
            self.assert_page_has(f"<th>{column}</th>")
        self.assert_plain_report()
        self.assert_multi_table_report()

    def test_updated_summary_ref_titles_report(self):
        """Test the Updated SummaryRef Titles report."""

        self.navigate_to("UpdatedSummaryRefTitles.py", prompt="yes")
        self.assert_title("Updated SummaryRef Titles")
        self.assert_page_has("Instructions")
        self.assert_page_has("This report shows summary documents in which")
        self.submit_form()
        self.assert_title("Updated SummaryRef Titles")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        self.assertEqual(table.caption.text, "Updates")
        table.check_headers(("Linking Doc ID", "Linked Doc ID", "Updated"))
        self.assertRegex(table.rows[0][0].text, self.CDR_ID_PATTERN)
        self.assertRegex(table.rows[0][1].text, self.CDR_ID_PATTERN)
        self.assertRegex(table.rows[0][2].text, self.DATETIME_PATTERN)


class TerminologyTests(Tester):
    """Tests for CDR Term document reports."""

    def test_ambiguous_drug_concepts_report(self):
        """Test the duplicate concept links report."""

        self.navigate_to("AmbiguousEVSDrugConcepts.py")
        expected = "EVS Drug Concepts Used By More Than One CDR Drug Term"
        self.assert_title(expected)
        self.assert_non_tabular_report()

    def test_cancer_diagnosis_hierarchy_report(self):
        """Test the CDR Cancer Diagnosis Hierarchy report."""

        # First try the full for of the report.
        form = self.navigate_to("DiseaseDiagnosisTerms.py", prompt="yes")
        self.assert_title("CDR Cancer Diagnosis Hierarchy Report")
        self.assert_page_has("Instructions")
        self.assert_page_has("This report represents the hierarchy for the ")
        self.submit_form()
        self.assert_title("CDR Cancer Diagnosis Hierarchy Report")
        self.assert_non_tabular_report()
        self.assert_page_has("malignant neoplasm")
        self.assert_page_has("x Malignant Tumor")

        # Then run the short version (showing only the preferred names).
        self.switch_to(form)
        self.click("flavor-short")
        self.submit_form()
        self.assert_title("CDR Cancer Diagnosis Hierarchy Report")
        self.assert_non_tabular_report()
        self.assert_page_has("malignant neoplasm")
        self.assert_page_not_has("x Malignant Tumor")

    def test_clinical_trials_drug_analysis_report(self):
        """Test the Clinical Trials Drug Analysis Report"""

        self.navigate_to("RecentCTGovProtocols.py")
        self.assert_title("Recent CT.gov Protocols")
        self.click("format-html")
        self.submit_form()
        self.assert_title("Recent CT.gov Protocols")
        self.assert_single_table_report()
        self.assert_plain_report()
        table = self.load_table()
        end = date.today()
        start = end - timedelta(30)
        expected = f"CT.gov Protocols Received Between {start} and {end}"
        caption = table.caption.find_element(By.TAG_NAME, "span").text
        self.assertEqual(caption, expected)
        headers = (
            "NCI ID",
            "Received",
            "Trial Title",
            "Phase",
            "Other IDs",
            "Sponsors",
        )
        table.check_headers(headers)

    def test_drug_review_report(self):
        """Test the Drug Review report."""

        script = "DrugReviewReport.py"
        self.navigate_to(script)
        self.assert_title("Drug Review Report")
        self.assert_page_has("To prepare an Excel format report of Drug/Agent")
        end = date.today()
        start = end - timedelta(365)
        params = dict(start=start, end=end, Request="Submit")
        book = self.fetch_workbook(script, params)
        headers = (
            "CDR ID",
            "Preferred Name",
            "Other Names",
            "Other Name Type",
            "Source",
            "TType",
            "SourceID",
            "Definition",
            "Created",
        )
        # The book.sheetnames property is a list, not a tuple.
        sheetnames = [
            "New Drugs from NCI Thesaurus",
            "New Drugs from the CDR",
            "Drugs to be Reviewed",
        ]
        self.assertEqual(book.sheetnames, sheetnames)
        sheet = book[sheetnames[0]]
        self.assertEqual(sheet["A1"].value, sheetnames[0])
        self.assertEqual(sheet.max_column, len(headers))
        for column, header in enumerate(headers, start=1):
            self.assertEqual(sheet.cell(2, column).value, header)
        sheet = book[sheetnames[2]]
        self.assertEqual(sheet["A1"].value, sheetnames[2])
        self.assertEqual(sheet.max_column, len(headers))
        for column, header in enumerate(headers, start=1):
            self.assertEqual(sheet.cell(2, column).value, header)
        headers = headers[:4] + headers[-1:]
        sheet = book[sheetnames[1]]
        self.assertEqual(sheet["A1"].value, sheetnames[1])
        self.assertEqual(sheet.max_column, len(headers))
        for column, header in enumerate(headers, start=1):
            self.assertEqual(sheet.cell(2, column).value, header)

    def test_hierarchy_tree_report(self):
        """Test the Term Hierarchy Tree report."""

        self.navigate_to("TermHierarchyTree.py", prompt="yes")
        self.assert_title("Term Hierarchy Tree")
        self.assert_page_has("Instructions")
        self.assert_page_has("This report provides an interactive interface")
        self.submit_form()
        self.assert_title("Term Hierarchy Tree")
        self.assert_plain_report()
        self.assert_non_tabular_report()
        top = self.find("ul li")
        nested = self.find("ul li ul")
        self.assertTrue(top.is_displayed())
        self.assertFalse(nested.is_displayed())
        top.find_element(By.TAG_NAME, "span").click()
        self.assertTrue(top.is_displayed())
        self.assertTrue(nested.is_displayed())

    def test_intervention_or_procedural_terms_report(self):
        """Test the Intervention or Procedural Terms report."""

        # The default report includes display of alternate term names.
        form = self.navigate_to("InterventionAndProcedureTerms.py")
        title = "CDR Intervention or Procedure Index Terms"
        self.assert_title(title)
        self.submit_form()
        self.assert_title(title)
        self.assert_non_tabular_report()
        self.assert_page_has("behavioral intervention")
        self.assert_page_has("x Behavior Modification")

        # Test the version which excludes the alternate names.
        self.switch_to(form)
        self.click("includealternatenames-false")
        self.submit_form()
        self.assert_title(f"{title} (without Alternate Names)")
        self.assert_non_tabular_report()
        self.assert_page_has("behavioral intervention")
        self.assert_page_not_has("x Behavior Modification")

    def test_private_concepts_report(self):
        """Test the Concepts Not Marked Public report."""

        self.navigate_to("ocecdr-3588.py", prompt="yes")
        self.assert_title("NCI Thesaurus Links Not Marked Public")
        self.assert_page_has("Instructions")
        self.assert_page_has("NCI Thesaurus concept codes/IDs are added")
        self.submit_form()
        self.assert_title("NCI Thesaurus Links Not Marked Public")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        columns = (
            "CDR ID",
            "Concept ID",
            "Available?",
            "Last Mod",
            "Semantic Types",
        )
        table.check_headers(columns)

    def test_qc_report(self):
        """Test the Terminology QC report."""

        self.navigate_to("TermSearch.py")
        self.assert_title("Term")
        term = self.get_test_term()
        term_id = term["id"]
        term_name = term["name"]
        self.set_field_value("name", term_name)
        self.find('main form input[type="submit"]').click()
        self.select_new_tab()
        self.assert_title("Term")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected = f"1 document matches '{term_name}'"
        self.assertEqual(table.caption.text, expected)
        link = table.rows[0][1].find_element(By.TAG_NAME, "a")
        self.assertIsNotNone(link)
        self.assertEqual(link.text, f"CDR{term_id:010d}")
        self.assertTrue(table.rows[0][2].text.startswith(term_name))
        link.click()
        self.assert_page_has("<center>Terminology<br>QC Report</center>")
        self.assert_page_has(f"CDR{term_id}")
        self.assert_page_has(term_name)
        self.assert_plain_report()

    def test_term_match_utility(self):
        """Test the Match Drug Terms By Name tool."""

        self.driver.set_page_load_timeout(600)
        self.navigate_to("MatchDrugTermsByName.py")
        self.assert_title("Match Drug Terms By Name")
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        caption_patterns = (
            r"Drug Terms Which Can Be Linked With EVS Concepts \((\d+)\)",
            r"Concepts Importable As New CDR Drug Terms \((\d+)\)",
        )
        tables[0].check_headers(("Element", "CDR", "EVS"))
        tables[1].check_headers(("Element", "Values"))
        re_match = re_search(caption_patterns[0], tables[0].caption.text)
        self.assertIsNotNone(re_match)
        linked = False
        linkable = int(re_match.group(1))
        if linkable > 0:
            link_button = self.find('input[name="refreshes"]')
            self.assertIsNotNone(link_button)
            value = link_button.get_attribute("value")
            link_concept_id, link_cdr_id = value.split("-")
            link_cdr_id = int(link_cdr_id)
            self.click(f"refreshes-{value.lower()}")
            linked = True
        re_match = re_search(caption_patterns[1], tables[1].caption.text)
        self.assertIsNotNone(re_match)
        imported = False
        importable = int(re_match.group(1))
        if importable > 0:
            import_button = self.find('input[name="creates"]')
            self.assertIsNotNone(import_button)
            import_concept_id = import_button.get_attribute("value")
            self.click(f"creates-{import_concept_id.lower()}")
            imported = True
        h2_patterns = (
            r"<h2>Concepts Unable To Be Matched Or Imported \(\d+\)</h2>",
            r"<h2>CDR Drug Term Documents With Anomalies \(\d+\)</h2>",
        )
        for pattern in h2_patterns:
            self.assert_regex(pattern)
        self.submit_form(new_tab=False)
        self.assert_title("Match Drug Terms By Name")
        tables = self.load_tables()
        if linked:
            note = "Refreshed from and associated with EVS concept"
            table = tables.pop(0)
            table.check_headers(("CDR ID", "Code", "Name", "Notes"))
            self.assertEqual(table.caption.text, "Actions")
            self.assertEqual(len(table.rows), 1)
            self.assertEqual(table.rows[0][0].text, str(link_cdr_id))
            self.assertEqual(table.rows[0][1].text, link_concept_id)
            self.assertEqual(table.rows[0][3].text, note)
        if imported:
            table = tables.pop(0)
            table.check_headers(("Code", "Name", "CDR ID", "Notes"))
            self.assertEqual(len(table.rows), 1)
            self.assertEqual(table.rows[0][0].text, import_concept_id)
            self.assertRegex(table.rows[0][2].text, r"CDR\d{10}")
            self.assertEqual(table.rows[0][3].text, "Created")
        self.assertEqual(len(tables), 2)
        table = tables.pop(0)
        re_match = re_search(caption_patterns[0], table.caption.text)
        self.assertIsNotNone(re_match)
        expected = linkable - 1 if linked else linkable
        self.assertEqual(int(re_match.group(1)), expected)
        table = tables.pop(0)
        re_match = re_search(caption_patterns[1], table.caption.text)
        self.assertIsNotNone(re_match)
        expected = importable - 1 if imported else importable
        self.assertEqual(int(re_match.group(1)), expected)

    def test_term_refresh_utilities(self):
        """Test the utilities for refreshing Term docs from EVS concepts."""

        # Check the landing page.
        landing_page = self.navigate_to("SyncWithEVS.py")
        self.assert_title("EVS Concept Tools")
        self.assert_page_has("Instructions")
        self.assert_page_has("Choose EVS Utility")
        instructions = self.find("instructions-accordion", method=By.ID)
        self.assertFalse(instructions.is_displayed())
        self.find("form h4").click()
        self.assertTrue(instructions.is_displayed())
        utilities = self.find("#utilities a", all=True)
        labels = (
            "Refresh CDR Drug Term Documents With EVS Concepts",
            "Match CDR Terms to EVS Concepts By Name",
            "EVS concepts used by more than one CDR Drug Term",
        )
        self.assertEqual(len(utilities), len(labels))
        for i, label in enumerate(labels):
            self.assertEqual(label, utilities[i].text)

        # Load the refresh tool.
        self.driver.implicitly_wait(180)
        utilities[0].click()
        refresh_form = self.select_new_tab()
        self.assert_page_has("Drug Terms Which Can Be Refreshed")
        self.driver.implicitly_wait(self.DEFAULT_WAIT)
        self.assert_single_table_report()
        self.assert_title("Refresh Drug Terms")
        table = self.load_table()
        table.check_headers(("Element", "CDR", "EVS"))
        caption = r"Drug Terms Which Can Be Refreshed From the EVS \((\d+)\)"
        re_match = re_search(caption, table.caption.text)
        self.assertIsNotNone(re_match)
        original_refresh_count = int(re_match.group(1))
        selector = 'input[name="actions"]'
        refresh_button = table.node.find_element(By.CSS_SELECTOR, selector)
        self.assertIsNotNone(refresh_button)
        value = refresh_button.get_attribute("value")
        concept_id, refresh_id = value.split("-")
        refresh_id = int(refresh_id)
        selector = 'input[name="suppress"]'
        suppress_buttons = table.node.find_elements(By.CSS_SELECTOR, selector)
        self.assertEqual(len(suppress_buttons), original_refresh_count)
        suppress_button = suppress_buttons[1]
        suppress_id = int(suppress_button.get_attribute("value"))
        self.click(f"actions-{concept_id.lower()}-{refresh_id}")
        self.click(f"suppress-{suppress_id}")
        self.driver.implicitly_wait(300)
        self.submit_form(new_tab=False)
        self.assert_page_has("Drug Terms Which Can Be Refreshed")
        self.driver.implicitly_wait(self.DEFAULT_WAIT)
        tables = self.load_tables()
        self.assertEqual(len(tables), 2)
        self.assertEqual(tables[0].caption.text, "Updates")
        tables[0].check_headers(("CDR ID", "Code", "Name", "Notes"))
        self.assertEqual(len(tables[0].rows), 1)
        self.assertEqual(tables[0].rows[0][0].text, str(refresh_id))
        self.assertEqual(tables[0].rows[0][1].text, concept_id)
        expected = "Refreshed from and associated with EVS concept"
        self.assertEqual(tables[0].rows[0][3].text, expected)
        tables[1].check_headers(("Element", "CDR", "EVS"))
        re_match = re_search(caption, tables[1].caption.text)
        self.assertIsNotNone(re_match)
        new_refresh_count = int(re_match.group(1))
        self.assertEqual(original_refresh_count-2, new_refresh_count)
        self.assert_page_not_has(f"CDR{suppress_id}")

        # This is an excellent time to test the Unsuppress tool.
        self.switch_to(landing_page)
        self.navigate_to("SuppressedDrugTerms.py")
        self.assert_title("Suppressed Drug Terms")
        self.assert_page_has("Click term to unsuppress")
        links = self.find("form ul li a", all=True)
        original_link_count = len(links)
        link = None
        target = f"CDR{suppress_id}"
        for candidate in links:
            if candidate.text.startswith(target):
                link = candidate
                break
        self.assertIsNotNone(link)
        link.click()
        expected = f"Removed {target} from the set of suppressed terms."
        self.assert_page_has(expected)
        self.assert_title("Suppressed Drug Terms")
        links = self.find("form ul li a", all=True)
        self.assertEqual(len(links), original_link_count-1)

        # We should now be able to reload the Refresh Drug Terms form
        # and verify that the unsuppressed term has been restored.
        self.switch_to(refresh_form)
        self.find("submit-button-sort-by-cdr-id", method=By.ID).click()
        self.assert_single_table_report()
        self.assert_title("Refresh Drug Terms")
        table = self.load_table()
        re_match = re_search(caption, table.caption.text)
        self.assertIsNotNone(re_match)
        final_refresh_count = int(re_match.group(1))
        self.assertEqual(original_refresh_count-1, final_refresh_count)
        self.assert_page_has(f"CDR{suppress_id}")

    def test_usage_report(self):
        """Test the Term Usage report."""

        self.navigate_to("TermUsage.py")
        self.assert_title("Report on documents indexed by specified terms")
        self.assert_page_has("Enter Term IDs Separated by Space")
        terms = self.get_test_data("terms")
        ids = " ".join([str(term["id"]) for term in terms["Index term"]])
        self.set_field_value("ids", ids)
        self.submit_form()
        self.assert_title("Report on documents indexed by specified terms")
        self.assert_single_table_report()
        self.assert_tables_in_grid_container()
        table = self.load_table()
        expected = r"Number of documents using specified terms: \d+"
        self.assertRegex(table.caption.text, expected)
        columns = "Doc Type", "Doc ID", "Doc Title", "Term ID", "Term Title"
        table.check_headers(columns)
        self.assertEqual(len(table.rows[0]), 5)
        self.assertRegex(table.rows[0][1].text, r"^CDR\d+$")
        self.assertRegex(table.rows[0][3].text, r"^CDR\d+$")


def setUpModule():
    """Note when we started the test run."""
    Tester.STARTED = datetime.now()


def tearDownModule():
    """Tell how long the run took and summarize the results."""

    elapsed = datetime.now() - Tester.STARTED
    Tester.LOGGER.info("%s elapsed for complete test run", elapsed)
    if Tester.SUCCESSES > 1 and not Tester.ERRORS and not Tester.FAILURES:
        Tester.LOGGER.info("all %d tests passed", Tester.SUCCESSES)
    else:
        args = Tester.SUCCESSES, Tester.ERRORS, Tester.FAILURES
        Tester.LOGGER.info("succeeded=%d errors=%d failures=%d", *args)


class Result(TextTestResult):
    """Customization to capture success/failure statistics."""

    def addSuccess(self, test: TestCase) -> None:
        Tester.SUCCESSES += 1
        return super().addSuccess(test)

    def addError(self, test: TestCase, err) -> None:
        Tester.ERRORS += 1
        return super().addError(test, err)

    def addFailure(self, test: TestCase, err) -> None:
        Tester.FAILURES += 1
        return super().addFailure(test, err)


class Runner(TextTestRunner):
    """Overridden to register our custom Result type."""
    resultclass = Result


if __name__ == "__main__":
    """Don't run as a script if we are loaded as a module.

    We create our own instance of the ArgumentParser class so we
    can get the values we need for the test run, and then we modify
    the array of command-line arguments so that the test harness's
    own argument parser will be happy. Next we clean up the global
    namespace and launch the test suite.
    """

    parser = ArgumentParser()
    parser.add_argument("--host", required=True)
    parser.add_argument("--api", required=True)
    parser.add_argument("--session", required=True)
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument("--tests", "-t", nargs="*")
    opts = parser.parse_args()
    Tester.HOST = opts.host
    Tester.API = opts.api
    Tester.SESSION = opts.session
    Tester.LOGGER.info("-" * 40)
    Tester.LOGGER.info("Tests started using %s", opts.host)
    new_args = ["-v"] if opts.verbose else []
    if opts.verbose:
        Tester.VERBOSE = True
    if opts.tests:
        new_args += opts.tests
    argv[1:] = new_args
    del opts
    del parser
    main(testRunner=Runner)
